# plant_tab.py — 식물상 분석 탭
# 구조: 식물상 > (현지조사/문헌조사) > (전체/1차/2차...) > (식물상 집계/생활형별/귀화율)
import sys, os, re
sys.path.insert(0, os.path.dirname(__file__))

from collections import OrderedDict
from PySide6.QtWidgets import (
    QApplication, QDoubleSpinBox, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QPlainTextEdit, QLineEdit, QComboBox,
    QTabWidget, QScrollArea, QFrame, QSizePolicy, QSplitter,
    QFormLayout, QSpinBox, QPushButton, QSlider, QHeaderView, QTableWidget,
    QGridLayout, QMessageBox, QFileDialog
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor


from analyzer import PlantStats
from parser   import ParsedPlant
from shared import _s, _has as _base_has
from ui_shared import (
    make_tab_qss, make_scroll_widget, bold_row,
    make_outline_btn_qss, _COPY_BTN_QSS, _OUTLINE_BTN_GREEN,
)

from config import (
    _BG, _CARD, _BORDER, _ERR, _TXT, _SUB,
    _ACCENT, _ACCENT_L,
    FF_KR, BD, BD1,
    GRAPH_H, SETTINGS, CHK_INDICATOR_QSS
)
from ui_shared import (
    _Canvas, _apply_mpl,
    CanvasGroupBox, ResizableCanvasFrame, CanvasSizeSlider,
    _make_tbl, _item, _auto_fit_table, _tbl_auto_height, _get_nice_bounds, _apply_graph_scale
)
from graph_widgets import (
    apply_graph_default,
    wrap_canvas as _common_wrap_canvas,
    make_settings_below_graph as _common_make_settings_below_graph,
    make_bar_settings_panel as _common_make_bar_settings_panel,
)
from shared import _prot_grade_str, _prot_list_graded, _apply_title, _resolve_iga, _josa
from ui_shared import _normalized_percentages, _apply_ratio_correction, lit_result_prefix, lit_shi_prefix
try:
    import matplotlib
    matplotlib.use("Agg")
    _MPL = True
except ImportError:
    _MPL = False

_QSS = f"""
QWidget   {{ background:{_BG}; {FF_KR}; font-size:13px; color:{_TXT}; }}
QGroupBox {{
    {FF_KR}; font-size:12px; font-weight:700; color:{_TXT};
    {BD}; border-radius:10px; margin-top:12px; background:{_CARD}; padding-top:8px;
}}
QGroupBox::title {{
    subcontrol-origin:margin; subcontrol-position:top left;
    left:12px; padding:0 6px; color:{_TXT}; font-size:12px;
}}
QTableWidget {{
    {BD}; border-radius:8px; background:{_CARD};
    gridline-color:{_BORDER}; {FF_KR}; font-size:12px;
}}
QTableWidget::item          {{ padding:1px 3px; }}
QTableWidget::item:selected {{ background:{_ACCENT_L}; color:{_ACCENT}; }}
QHeaderView::section {{
    background:{_BG}; {BD1}; border-right:1px solid {_BORDER};
    padding:3px 3px; {FF_KR}; font-size:11px; font-weight:700; color:{_SUB};
}}
QScrollArea {{ border:none; background:transparent; }}
QLabel, QCheckBox, QRadioButton {{ background: transparent; }}
QPlainTextEdit {{ background: #FFFFFF; }}
QScrollBar:vertical         {{ background:{_BG}; width:6px; border-radius:3px; }}
QScrollBar::handle:vertical {{ background:{_BORDER}; border-radius:3px; min-height:30px; }}
QScrollBar::handle:vertical:hover {{ background:#B0B8C8; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}
""" + CHK_INDICATOR_QSS

_LIM_QSS = (f"QLineEdit {{ background:#FFFFFF; border:1px solid {_BORDER}; border-radius:4px;"
            f" {FF_KR}; font-size:12px; padding:2px 6px; max-width:40px; }}")

_VEG_WHITE_QSS = f"""
QWidget {{ background:#FFFFFF; {FF_KR}; font-size:13px; color:{_TXT}; }}
QGroupBox {{ background:#FFFFFF; border:1px solid {_BORDER}; border-radius:8px; margin-top:8px; padding-top:8px; }}
QGroupBox::title {{ subcontrol-origin:margin; subcontrol-position:top left; left:10px; padding:0 4px; color:{_TXT}; font-size:12px; font-weight:700; }}
QTabWidget::pane {{ border:none; background:#FFFFFF; }}
QTabBar::tab {{ background:#FFFFFF; border:1px solid {_BORDER}; border-bottom:1px solid {_BORDER}; border-radius:6px 6px 0 0; padding:5px 12px; margin-right:3px; }}
QTabBar::tab:selected {{ color:{_ACCENT}; font-weight:700; border-bottom:1px solid #FFFFFF; }}
QPlainTextEdit {{ background:#FFFFFF; }}
QTableWidget {{ background:#FFFFFF; gridline-color:{_BORDER}; }}
QHeaderView::section {{ background:#FFFFFF; border:1px solid {_BORDER}; padding:4px; }}
QScrollArea {{ border:none; background:#FFFFFF; }}
"""

_VEG_FLORA_OPTIONS = [
    "한랭 온대 식물상 지역",
    "상량 온대 식물상 지역",
    "온난 온대 식물상 지역",
    "해양성 온난 온대 식물상 지역",
]

_VEG_CLIMATE_OPTIONS = [
    "온대북부 침엽·낙엽활엽수 혼합림(WI < 45)",
    "온대북부 낙엽활엽수림(WI 45~85)",
    "온대중부 식생기후 낙엽활엽수림(WI 85~100)",
    "온대남부 식생기후 낙엽활엽수림(WI > 100)",
    "온대남부 식생기후 상록·낙엽활엽수 혼합림(CI > -10)",
]

_VEG_FLORA_FEATURES = {
    "한랭온대식물상지역": "눈잣나무, 눈측백, 분비나무 분포 범위의 남방한계선으로,",
    "상량온대식물상지역": "따뜻한 해양 기후에 분포하는 목본성 수종이 존재하는 지역으로,",
    "온난온대식물상지역": "우리나라 관속식물종의 약 75%가 존재하는 지역으로,",
    "해양성온난온대식물상지역": "난초과의 핵심 분포 공간으로서 높은 식물 다양성을 나타내는 지역으로,",
}

_VEG_CLIMATE_SPECIES = {
    "온대북부침엽·낙엽활엽수혼합림(WI<45)": "분비나무, 구상나무, 가문비나무, 잣나무, 사스래나무, 신갈나무, 벚나무류, 부게꽃나무, 산겨릅나무, 피나무",
    "온대북부낙엽활엽수림(WI45~85)": "신갈나무, 벚나무류, 서어나무, 피나무, 물푸레나무, 함박꽃나무, 쪽종백나무, 구상나무, 분비나무, 잣나무",
    "온대중부식생기후낙엽활엽수림(WI85~100)": "졸참나무, 신갈나무, 벚나무류, 물푸레나무, 서어나무, 함박꽃나무, 당단풍나무, 소나무, 곰솔, 전나무, 때죽나무",
    "온대남부식생기후낙엽활엽수림(WI>100)": "졸참나무, 서어나무, 개서어나무, 벚나무류, 물푸레나무, 단풍나무, 대팻집나무, 때죽나무, 소나무, 곰솔, 굴거리나무, 동백나무",
    "온대남부식생기후상록·낙엽활엽수혼합림(CI>-10)": "구실잣밤나무, 종가시나무, 붉가시나무, 사스레피나무, 굴거리나무, 동백나무, 졸참나무, 벚나무류, 개서어나무, 때죽나무",
}

def _get_tbl(obj):
    if isinstance(obj, QTableWidget): return obj
    if hasattr(obj, "findChild"):
        return obj.findChild(QTableWidget)
    return None

def _add_tbl_with_btn(lay, tbl, stretch=0, pct_col=None, cnt_col=None, pct_row=None, cnt_row=None, on_normalize=None):
    from ui_shared import normalize_tbl_pct_col, normalize_tbl_pct_row
    w = QWidget()
    wl = QVBoxLayout(w)
    wl.setContentsMargins(0, 0, 0, 0)

    btn_row_lay = QHBoxLayout()
    btn_row_lay.setContentsMargins(0, 0, 0, 0)
    btn_row_lay.addStretch()

    hint_lbl = None
    if pct_col is not None or pct_row is not None:
        tbl._is_normalized = bool(getattr(tbl, "_is_normalized", False))
        if pct_col is not None:
            tbl._pct_col = pct_col
            tbl._cnt_col = cnt_col
        if pct_row is not None:
            tbl._pct_row = pct_row
            tbl._cnt_row = cnt_row
            tbl._pct_start_col = 1
            tbl._pct_end_col_offset = 1
        btn_norm = QPushButton("∑ 비율 보정")
        btn_norm.setFixedHeight(24)
        btn_norm.setStyleSheet(_OUTLINE_BTN_GREEN)
        btn_norm.setToolTip("비율(%) 합계를 100.00으로 자동 보정하고 변경 내역을 표시합니다.")

        hint_lbl = QLabel("")
        hint_lbl.setStyleSheet(
            f"font-size: 11px; color: #B45309; {FF_KR}; "
            "background: #FFFBEB; border: 1px solid #FCD34D; border-radius: 4px; "
            "padding: 3px 8px;"
        )
        hint_lbl.setWordWrap(True)
        hint_lbl.hide()

        def _do_norm():
            hint = None
            if pct_col is not None:
                hint = normalize_tbl_pct_col(tbl, pct_col, cnt_col=cnt_col)
            elif pct_row is not None:
                hint = normalize_tbl_pct_row(tbl, pct_row, cnt_row=cnt_row)
            tbl._is_normalized = True
            if hint:
                hint_lbl.setText(hint)
                hint_lbl.show()
            else:
                hint_lbl.hide()
            if hasattr(tbl, 'on_normalize') and tbl.on_normalize:
                tbl.on_normalize()
            elif on_normalize:
                on_normalize()
        btn_norm.clicked.connect(_do_norm)
        btn_row_lay.addWidget(btn_norm)

    btn_copy = QPushButton("📋 표 복사")
    btn_copy.setFixedHeight(24)
    btn_copy.setStyleSheet(_COPY_BTN_QSS)
    def _copy():
        tbl.copy_selection(include_header=True)
        from ui_shared import apply_button_feedback
        apply_button_feedback(btn_copy)
    btn_copy.clicked.connect(_copy)
    btn_row_lay.addWidget(btn_copy)

    wl.addLayout(btn_row_lay)
    if hint_lbl is not None:
        wl.addWidget(hint_lbl)
    wl.addWidget(tbl)
    lay.addWidget(w, stretch)

def _tab_qss(big=True):
    return make_tab_qss(FF_KR, _SUB, _ACCENT, big)


def _has(v):
    # "○"는 출현 표시이므로 False 목록에서 제외
    return _base_has(v, ("", "None", "-", "x", "X", "×"))


def _bold_row(tbl, row):
    bold_row(tbl, row, _BG)


def _compact_table(tbl, row_h=22):
    tbl.verticalHeader().setDefaultSectionSize(row_h)
    for row in range(tbl.rowCount()):
        tbl.setRowHeight(row, row_h)


def _fit_table_height(tbl, extra=6):
    header_h = tbl.horizontalHeader().height()
    rows_h = sum(tbl.rowHeight(row) for row in range(tbl.rowCount()))
    frame_h = tbl.frameWidth() * 2
    scroll_h = tbl.horizontalScrollBar().sizeHint().height() if tbl.horizontalScrollBar().isVisible() else 0
    tbl.setFixedHeight(header_h + rows_h + frame_h + scroll_h + extra)


def _pad_table_columns(tbl, first_extra=18, other_extra=10):
    if tbl.columnCount() == 0:
        return
    for col in range(tbl.columnCount()):
        extra = first_extra if col == 0 else other_extra
        tbl.setColumnWidth(col, tbl.columnWidth(col) + extra)


def _fit_table_no_scroll(tbl, first_contents=True):
    header = tbl.horizontalHeader()
    tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    tbl.setWordWrap(False)
    tbl.resizeColumnsToContents()
    if tbl.columnCount() == 0:
        return
    if first_contents:
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        for col in range(1, tbl.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.Stretch)
    else:
        for col in range(tbl.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.Stretch)
    header.setStretchLastSection(True)


def _sc(w):
    return make_scroll_widget(w)


def _txtbox(text, h=60):
    t = QPlainTextEdit()
    t.setReadOnly(True)
    t.setFixedHeight(h)
    t.setStyleSheet(
        f"QPlainTextEdit{{background:#FFFFFF;{BD};border-radius:7px;{FF_KR};font-size:12px;}}"
    )
    t.setPlainText(text)
    return t


def _analysis_sentence(mode_label, round_label, ctx, nat_sheet=None, disturb_sheet=None):
    """분석 문장 생성 (단일 문자열 반환)."""
    S = SETTINGS
    species = ctx.get("species", [])
    rounds = ctx.get("rounds", [])
    unit_counts = ctx.get("unit_counts", {})
    nat_metrics = ctx.get("nat_metrics", (0, 0, 0, 0))
    
    total = len(species)
    if total == 0:
        return "분석할 식물 데이터가 없습니다."
    
    unit = unit_counts.get("종합", {})
    counts_str = _plant_counts_str(unit) if unit else f"{total}분류군"
    
    is_field = (mode_label == "현지조사")
    s1_mid = getattr(S, "plant_field_s1_mid", S.field_s1_mid) if is_field else getattr(S, "plant_lit_s1_mid", S.lit_s1_mid)
    s1_end = getattr(S, "plant_field_s1_end", S.field_s1_end) if is_field else getattr(S, "plant_lit_s1_end", S.lit_s1_end)

    s1_mid_josa = _resolve_iga(s1_mid, counts_str)

    # 귀화율 정보
    nat_rate = nat_metrics[2] if len(nat_metrics) >= 3 else 0.0

    order_verb_f = getattr(S, "plant_order_verb_f", S.order_verb_f)
    s1_end_base = s1_end.replace("이 ", "").replace("가 ", "")

    prefix_res = f"{mode_label} 寃곌낵" if is_field else lit_result_prefix(rounds, mode_label)
    sentence = f"{prefix_res} 확인된 소산식물은 {counts_str}{s1_mid_josa} 총 {total}분류군이 {order_verb_f} 귀화율은 {nat_rate:.1f}%로 {s1_end_base}"
    
    return _apply_title(sentence, S.sentence_title)


def _sent_txtbox(mode_label, round_label, ctx, nat_sheet=None, disturb_sheet=None, h=36):
    """문장 QPlainTextEdit — 설정 변경 시 _sent_refresh_fn() 으로 갱신 가능."""
    txt = _txtbox(_analysis_sentence(mode_label, round_label, ctx, nat_sheet, disturb_sheet), h)
    txt._sent_refresh_fn = lambda: txt.setPlainText(
        _analysis_sentence(mode_label, round_label, ctx, nat_sheet, disturb_sheet)
    )
    return txt


def _rnd_label(rn):
    parts = rn.split("_", 2)
    return parts[1] if len(parts) > 1 else rn

def _round_survey_label(label):
    label = str(label or "").strip()
    if not label:
        return "조사"
    if "".join(label.split()).endswith("조사"):
        return label
    return f"{label} 조사"


def _round_groups(rns):
    d = OrderedDict()
    for r in rns:
        lb = _rnd_label(r)
        d.setdefault(lb, []).append(r)
    return d


def _species_key(species):
    return tuple(id(sp) for sp in species)


def _rounds_key(rns):
    return tuple(rns)


def _species_in_rounds(species, rns, cache=None):
    if cache is not None:
        key = (_species_key(species), _rounds_key(rns))
        hit = cache.get(key)
        if hit is not None:
            return hit
    out = [sp for sp in species if any(_has(sp.rounds.get(r)) for r in rns)]
    if cache is not None:
        cache[key] = out
    return out


class _FloatSlider(QWidget):
    valueChanged = Signal(float)

    def __init__(self, min_val, max_val, step, default_val, suffix="", decimals=2, parent=None):
        super().__init__(parent)
        self._min = min_val
        self._max = max_val
        self._step = step
        self._suffix = suffix
        self._decimals = decimals

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)

        self._sld = QSlider(Qt.Horizontal)
        self._sld.setRange(0, int(round((max_val - min_val) / step)))
        self._sld.setFixedHeight(22)
        self._sld.setMinimumWidth(150)
        self._lbl = QLabel()
        self._lbl.setFixedWidth(48)
        self._lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        lay.addWidget(self._sld, 1)
        lay.addWidget(self._lbl)
        self._sld.valueChanged.connect(self._on_change)
        self.setValue(default_val)

    def _to_real(self, raw):
        return self._min + raw * self._step

    def _to_raw(self, value):
        return int(round((value - self._min) / self._step))

    def _on_change(self, raw):
        self._lbl.setText(f"{self._to_real(raw):.{self._decimals}f}{self._suffix}")
        self.valueChanged.emit(self._to_real(raw))

    def value(self):
        return self._to_real(self._sld.value())

    def setValue(self, value):
        self._sld.setValue(self._to_raw(value))
        self._lbl.setText(f"{self.value():.{self._decimals}f}{self._suffix}")

    def setMaximum(self, max_val):
        self._max = max_val
        self._sld.setRange(0, int(round((self._max - self._min) / self._step)))

    def formatted_value(self):
        return f"{self.value():.{self._decimals}f}{self._suffix}"


_FERN_FAMILIES = {
    "OPHIOGLOSSACEAE", "EQUISETACEAE", "OSMUNDACEAE", "DENNSTAEDTIACEAE",
    "PTERIDACEAE", "ASPLENIACEAE", "ATHYRIACEAE", "BLECHNACEAE",
    "DRYOPTERIDACEAE", "WOODSIACEAE", "THELYPTERIDACEAE", "ONOCLEACEAE",
    "POLYPODIACEAE", "MARSILEACEAE", "SALVINIACEAE", "AZOLLACEAE",
    "HYMENOPHYLLACEAE", "LYCOPODIACEAE", "SELAGINELLACEAE", "ISOETACEAE",
    "ADIANTACEAE", "ASPIDIACEAE", "PARKERIACEAE",
}
_GYMNO_FAMILIES = {
    "PINACEAE", "CUPRESSACEAE", "TAXACEAE", "CEPHALOTAXACEAE",
    "GINKGOACEAE", "EPHEDRACEAE", "PODOCARPACEAE", "TAXODIACEAE", "CYCADACEAE",
}
_MONOCOT_FAMILIES = {
    "POACEAE", "CYPERACEAE", "JUNCACEAE", "LILIACEAE", "SMILACACEAE",
    "DIOSCOREACEAE", "IRIDACEAE", "ORCHIDACEAE", "ARACEAE", "TYPHACEAE",
    "SPARGANIACEAE", "POTAMOGETONACEAE", "NAJADACEAE", "HYDROCHARITACEAE",
    "ALISMATACEAE", "BUTOMACEAE", "COMMELINACEAE", "PONTEDERIACEAE",
    "ACORACEAE", "ERIOCAULACEAE", "AMARYLLIDACEAE", "ASPARAGACEAE",
    "MELANTHIACEAE", "COLCHICACEAE", "ALLIACEAE", "CONVALLARIACEAE",
    "AGAVACEAE", "LEMNACEAE", "JUNCAGINACEAE", "TRILLIACEAE",
    "HAEMODORACEAE", "BROMELIACEAE", "MUSACEAE", "ZINGIBERACEAE", "CANNACEAE",
}


def _pct(v):
    """백분율 값 포맷팅 (소수점 자릿수는 SETTINGS.decimal 사용)."""
    return f"{v:.{SETTINGS.decimal}f}"


def _clean_sci_name(sp):
    name = _s(getattr(sp, "sci_name", "")).replace("_", " ")
    return re.sub(r"\s+", " ", name).strip()


def _norm_family(sp):
    return _s(getattr(sp, "family", "")).replace("_", " ").strip().upper()


def _genus_of(sp):
    parts = _clean_sci_name(sp).split()
    return parts[0] if parts else ""


def _is_variety(name):
    n = f" {name.lower()} "
    return " var. " in n or " var " in n


def _is_forma(name):
    n = f" {name.lower()} "
    return " f. " in n or " forma " in n


def _is_subsp(name):
    n = f" {name.lower()} "
    return any(x in n for x in [" subsp. ", " subsp "])


def _species_binomial(name):
    parts = name.split()
    if len(parts) < 2:
        return ""
    if parts[1].lower() in {"sp.", "sp", "spp.", "spp"}:
        return ""
    return f"{parts[0]} {parts[1]}"


# 단자엽식물강 클래스명 (한국 식물 분류 체계)
_MONOCOT_CLASSES = {"백합강", "단자엽식물강", "LILIOPSIDA"}

# 쌍자엽식물강 클래스명
_DICOT_CLASSES = {"목련강", "쌍자엽식물강", "MAGNOLIOPSIDA"}

# 나자식물 클래스명
_GYMNO_CLASSES = {"소나무강", "은행나무강", "소철강",
                  "PINOPSIDA", "GINKGOOPSIDA", "CYCADOPSIDA"}

# 양치식물 클래스명
_FERN_CLASSES = {"고사리강", "속새강", "석송강", "솔잎난강", "관음고사리강",
                 "POLYPODIOPSIDA", "EQUISETOPSIDA", "LYCOPODIOPSIDA",
                 "MARATTIOPSIDA", "PSILOTOPSIDA"}


def _plant_group_of(sp):
    # 1순위: Class 행으로 명시된 클래스명 기반 판별 (가장 신뢰도 높음)
    cls = _s(getattr(sp, "class_name", "")).strip().upper()
    if cls:
        # 백합강 계열 → 단자엽
        if cls in {c.upper() for c in _MONOCOT_CLASSES}:
            return "단자엽식물"
        # 목련강 계열 → 쌍자엽 (양치·나자 제외)
        if cls in {c.upper() for c in _DICOT_CLASSES}:
            return "쌍자엽식물"
        if cls in {c.upper() for c in _GYMNO_CLASSES}:
            return "나자식물문"
        if cls in {c.upper() for c in _FERN_CLASSES}:
            return "양치식물문"

    # 2순위: 과명(Family) 기반 판별 (클래스 정보가 없는 경우 fallback)
    fam = _norm_family(sp)
    if fam in _FERN_FAMILIES:
        return "양치식물문"
    if fam in _GYMNO_FAMILIES:
        return "나자식물문"
    if fam in _MONOCOT_FAMILIES:
        return "단자엽식물"
    if fam.endswith("ACEAE"):
        return "쌍자엽식물"
    return "쌍자엽식물"


def _calc_counts(rows):
    fam = {_s(sp.family) for sp in rows if _s(sp.family)}
    genus = {_genus_of(sp) for sp in rows if _genus_of(sp)}
    var_set, form_set, subsp_set = set(), set(), set()
    for sp in rows:
        nm = _clean_sci_name(sp)
        if not nm:
            continue
        if _is_variety(nm):
            var_set.add(nm)
        elif _is_forma(nm):
            form_set.add(nm)
        elif _is_subsp(nm):
            subsp_set.add(nm)
    var_n = len(var_set)
    form_n = len(form_set)
    subsp_n = len(subsp_set)
    return {
        "과": len(fam),
        "속": len(genus),
        "종": max(0, len(rows) - var_n - form_n - subsp_n),
        "변종": var_n,
        "품종": form_n,
        "아종": subsp_n,
        "분류군": len(rows),
    }


def _unit_counts(species, cache=None):
    if cache is not None:
        skey = _species_key(species)
        if skey in cache:
            return cache[skey]
    buckets = {"양치식물문": [], "나자식물문": [], "단자엽식물": [], "쌍자엽식물": []}
    for sp in species:
        buckets[_plant_group_of(sp)].append(sp)
    data = OrderedDict()
    data["양치식물문"] = _calc_counts(buckets["양치식물문"])
    data["나자식물문"] = _calc_counts(buckets["나자식물문"])
    ang_rows = buckets["단자엽식물"] + buckets["쌍자엽식물"]
    data["피자식물문"] = _calc_counts(ang_rows)
    data["쌍자엽식물"] = _calc_counts(buckets["쌍자엽식물"])
    data["단자엽식물"] = _calc_counts(buckets["단자엽식물"])
    data["종합"] = _calc_counts(species)
    if cache is not None:
        cache[skey] = data
    return data


def _disp(v):
    return "-" if not v else str(v)


def _count_tbl(species, unit_cache=None):
    headers = ["구 분", "과", "속", "종", "변종", "품종", "아종", "분류군"]
    tbl = _make_tbl(headers)
    data = _unit_counts(species, unit_cache)
    tbl.setRowCount(len(data))
    for ri, (label, vals) in enumerate(data.items()):
        tbl.setItem(ri, 0, _item(label, Qt.AlignCenter))
        for ci, key in enumerate(headers[1:], start=1):
            tbl.setItem(ri, ci, _item(_disp(vals[key])))
    _bold_row(tbl, len(data) - 1)
    _compact_table(tbl, 24)
    _auto_fit_table(tbl)
    _fit_table_no_scroll(tbl, True)
    _fit_table_height(tbl, 16)
    return tbl


_LIFE_KEYS = ["M", "N", "E", "CH", "H", "G", "HH", "TH"]
_LIFE_HEADER = ["M", "N", "E", "Ch", "H", "G", "HH", "Th", "합계"]
_LIFE_FIXED_SOUTH = {
    "M": 20.1, "N": 14.8, "E": 7.4, "CH": 1.9,
    "H": 23.0, "G": 12.4, "HH": 1.4, "TH": 19.0,
}
_LIFE_FIXED_RAUNKIAER = {
    "M": 26.0, "N": 15.0, "E": 3.0, "CH": 9.0,
    "H": 28.0, "G": 4.0, "HH": 2.0, "TH": 13.0,
}
_NAT_COMPARE_LABELS = ["언덕주택지", "밭", "시가지", "평지주택지", "논", "냇가", "계단식논", "풀밭", "숲"]
_NAT_COMPARE_VALUES = [48.8, 32.1, 27.7, 18.1, 14.5, 13.3, 7.2, 4.9, 4.4]
_URBANIZATION_INDEX = 16.20


def _life_code_list(sp):
    raw = _s(sp.rounds.get("life"))
    if not raw:
        return []
    parts = raw.replace("/", "+").replace(",", "+").split("+")
    out = []
    for p in parts:
        code = _s(p).upper()
        if code in _LIFE_KEYS:
            out.append(code)
    return out


def _life_counts(species, cache=None):
    if cache is not None:
        skey = _species_key(species)
        if skey in cache:
            return cache[skey]
    cnt = {k: 0 for k in _LIFE_KEYS}
    for sp in species:
        seen = set(_life_code_list(sp))
        for code in seen:
            cnt[code] += 1
    if cache is not None:
        cache[skey] = cnt
    return cnt


def _life_tbl(species, life_counts=None):
    cnt = life_counts if life_counts is not None else _life_counts(species)
    total = len(species)
    tbl = _make_tbl(["구 분"] + _LIFE_HEADER)
    tbl.setRowCount(4)
    tbl.setItem(0, 0, _item("종 수", Qt.AlignCenter))
    tbl.setItem(1, 0, _item("비율", Qt.AlignCenter))
    tbl.setItem(2, 0, _item("남한지역", Qt.AlignCenter))
    tbl.setItem(3, 0, _item("Raunkiaer's N.S.", Qt.AlignCenter))
    s = 0
    
    
    pct_sum = 0.0
    for ci, key in enumerate(_LIFE_KEYS, start=1):
        v = cnt[key]
        s += v
        tbl.setItem(0, ci, _item(_disp(v)))
        pct = (v / total * 100) if total else 0.0
        tbl.setItem(1, ci, _item("-" if v == 0 else f"{pct:.1f}"))
        tbl.setItem(2, ci, _item(f"{_LIFE_FIXED_SOUTH[key]:.1f}"))
        tbl.setItem(3, ci, _item(f"{_LIFE_FIXED_RAUNKIAER[key]:.1f}"))
        pct_sum += round(pct, 1)

    tbl.setItem(0, len(_LIFE_KEYS) + 1, _item(_disp(s)))

    # 총 비율(%)을 동적으로 표시하고 100.0이 아니면 빨간색으로 표시
    from ui_shared import set_total_pct_cell as _shared_set_total_pct_cell
    _shared_set_total_pct_cell(
        tbl, 1, len(_LIFE_KEYS) + 1, pct_sum if s else 0,
        item_factory=_item,
        pct_formatter=lambda x: "-" if s == 0 else f"{float(x):.1f}",
        err_color=_ERR,
        decimals=1
    )
    
    tbl.setItem(2, len(_LIFE_KEYS) + 1, _item("100.0"))
    tbl.setItem(3, len(_LIFE_KEYS) + 1, _item("100.0"))
    _compact_table(tbl, 24)
    _auto_fit_table(tbl)
    _fit_table_no_scroll(tbl, True)
    _fit_table_height(tbl, 16)
    return tbl


def _life_pct_values(species, life_counts=None):
    cnt = life_counts if life_counts is not None else _life_counts(species)
    total = len(species)
    if not total:
        return [0.0 for _ in _LIFE_KEYS]
    return [round(cnt[key] / total * 100, 1) for key in _LIFE_KEYS]


def _build_round_context(species, rounds, nat_sheet=None, disturb_sheet=None, cache=None):
    cache = cache or {}
    life_cache = cache.setdefault("life_counts", {})
    unit_cache = cache.setdefault("unit_counts", {})
    nat_cache = cache.setdefault("nat_metrics", {})
    selected_cache = cache.setdefault("selected_name_sets", {})

    species_key = _species_key(species)
    rounds_key = _rounds_key(rounds)
    life_counts = _life_counts(species, life_cache)
    unit_counts = _unit_counts(species, unit_cache)
    nat_metrics = _naturalized_metrics(
        species,
        rounds,
        nat_sheet,
        disturb_sheet,
        nat_cache=nat_cache,
        selected_cache=selected_cache,
    )
    return {
        "species": species,
        "rounds": rounds,
        "life_counts": life_counts,
        "unit_counts": unit_counts,
        "nat_metrics": nat_metrics,
        "species_key": species_key,
        "rounds_key": rounds_key,
        "cache": cache,
    }


def _draw_life_compare(canvas, tbl_obj):
    if not _MPL: return
    tbl = _get_tbl(tbl_obj)
    if not tbl: return

    survey = []
    for c in range(1, 9):
        txt = tbl.item(1, c).text().replace("%", "").strip()
        survey.append(0.0 if txt == "-" else float(txt))
        
    _apply_mpl()
    canvas._fig.clf()
    ax = canvas._fig.add_subplot(111)

    decimal = canvas._cfg.get("decimal", SETTINGS.decimal)
    fs = canvas._cfg.get("bar_fontsize", SETTINGS.bar_fontsize)
    labels = ["M", "N", "E", "Ch", "H", "G", "HH", "Th"]
    south = [_LIFE_FIXED_SOUTH[key] for key in _LIFE_KEYS]
    raunkiaer = [_LIFE_FIXED_RAUNKIAER[key] for key in _LIFE_KEYS]
    width = min(max(canvas._cfg.get("bar_v_width", 0.40), 0.08), 1.0)
    step = max(2.0, width * 3.55 + 0.20)
    x = [i * step for i in range(len(labels))]
    offsets = [-width * 1.18, 0.0, width * 1.18]

    series = [
        ("조사지역(%)", survey, "#d9d9d9", None),
        ("남한지역", south, "#ffffff", "////"),
        ("Raunkiaer's N.S.", raunkiaer, "#f2f2f2", "xxxx"),
    ]
    for idx, (name, values, color, hatch) in enumerate(series):
        bars = ax.bar(
            [xi + offsets[idx] for xi in x],
            values,
            width=width,
            label=name,
            color=color,
            edgecolor="#6b7280",
            linewidth=0.8,
            hatch=hatch,
            zorder=3,
        )
        for bar, value in zip(bars, values):
            if value <= 0:
                continue
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                value + 0.18 + (idx * 0.22),
                f"{value:.{decimal}f}",
                ha="center",
                va="bottom",
                fontsize=max(6, fs - 3),
            )

    bar_left = min((xi + off) - width / 2 for xi in x for off in offsets)
    bar_right = max((xi + off) + width / 2 for xi in x for off in offsets)
    pad = max(width * 0.9, step * 0.18)

    boundary_ticks = [(left + right) / 2 for left, right in zip(x[:-1], x[1:])]

    axis_min = canvas._cfg.get("axis_min", SETTINGS.axis_min)
    axis_max = canvas._cfg.get("axis_max", SETTINGS.axis_max)
    axis_step = canvas._cfg.get("axis_step", SETTINGS.axis_step)
    
    max_val = max(max(survey), max(south), max(raunkiaer)) if survey else 30.0
    c_min, c_max, c_step = axis_min, axis_max, axis_step
    effective_max = c_max if c_max > 0 else (max_val if max_val > 0 else 1.0)
    auto_max, auto_step = _get_nice_bounds(effective_max)
    if c_max <= 0: c_max = auto_max
    if c_step <= 0: c_step = auto_step

    ax.set_xticks(boundary_ticks)
    ax.set_xticklabels([""] * len(boundary_ticks))
    if c_min < c_max:
        ax.set_ylim(c_min, c_max)
    if c_step > 0:
        import matplotlib.ticker as ticker
        ax.yaxis.set_major_locator(ticker.MultipleLocator(c_step))
    ax.set_xlim(bar_left - pad, bar_right + pad)
    ax.margins(x=0.01)
    ax.tick_params(axis="x", which="major", pad=4, length=4, width=0.8, direction="out")
    ax.tick_params(axis="y", labelsize=max(8, fs))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_linewidth(1.0)
    ax.spines["left"].set_linewidth(1.0)
    for xi, label in zip(x, labels):
        ax.text(
            xi,
            -0.075,
            label,
            transform=ax.get_xaxis_transform(),
            ha="center",
            va="top",
            fontsize=max(8, fs),
        )
    if canvas._cfg.get("grid_on", SETTINGS.grid_on):
        ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35, zorder=0)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.10), ncol=3, frameon=False, fontsize=max(8, fs))
    canvas._fig.subplots_adjust(left=0.15, right=0.95, bottom=0.15, top=0.90)
    _apply_graph_scale(canvas._fig, canvas)
    canvas.draw()


def _life_panel(ctx, on_norm_cb=None, parent_window=None):
    species = ctx["species"]
    life_counts = ctx["life_counts"]
    panel = QWidget()
    lay = QVBoxLayout(panel)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.setSpacing(8)
    tbl = _life_tbl(species, life_counts)
    panel._table_refs = {"life": lambda tbl=tbl: tbl}
    _add_tbl_with_btn(lay, tbl, 0, pct_row=1, cnt_row=0, on_normalize=lambda: (cv._refresh_cb() if _MPL else None, on_norm_cb() if on_norm_cb else None))
    if _MPL:
        cv = _Canvas(6.6, 3.5)
        cv._setting_mode = "bar"
        cv._cfg["bar_v_width"] = 0.4
        cv._cfg["initial_graph_height"] = 350
        apply_graph_default(parent_window, cv, "plant_life")
        cv._preview_draw = lambda target: _draw_life_compare(target, tbl)
        cv._refresh_cb = lambda: _draw_life_compare(cv, tbl)
        cv._refresh_cb()
        tbl.on_normalize = cv._refresh_cb
        cv_grp = _wrap_cv("생활형 그래프", cv)
        cv_grp._initial_height = 350
        panel_w = _make_bar_settings_panel(
            cv,
            cv_grp,
            lambda: _draw_life_compare(cv, tbl),
            title_default="생활형 비교",
            y_default="비율(%)",
            parent=panel,
        )
        lay.addWidget(_make_settings_graph(panel_w, cv_grp), 1)
    return panel


def _selected_name_set(sheet, rounds, selected_cache=None, round_species_cache=None):
    if not sheet:
        return set()
    cache_key = (id(sheet), _rounds_key(rounds))
    if selected_cache is not None and cache_key in selected_cache:
        return selected_cache[cache_key]
    chosen = _species_in_rounds(sheet.species, rounds, round_species_cache)
    names = {_clean_sci_name(sp) for sp in chosen if _clean_sci_name(sp)}
    if selected_cache is not None:
        selected_cache[cache_key] = names
    return names


def _naturalized_metrics(species, rounds, nat_sheet=None, disturb_sheet=None, nat_cache=None, selected_cache=None, round_species_cache=None):
    if nat_cache is not None:
        cache_key = (_species_key(species), _rounds_key(rounds), id(nat_sheet), id(disturb_sheet))
        if cache_key in nat_cache:
            return nat_cache[cache_key]
    selected = {_clean_sci_name(sp) for sp in species if _clean_sci_name(sp)}
    nat_names = _selected_name_set(nat_sheet, rounds, selected_cache, round_species_cache)
    disturb_names = _selected_name_set(disturb_sheet, rounds, selected_cache, round_species_cache)
    nat_cnt = len(selected & nat_names) if nat_names else 0
    disturb_cnt = len(selected & disturb_names) if disturb_names else 0
    total = len(species)
    rate = (nat_cnt / total * 100) if total else 0.0
    result = (total, nat_cnt, rate, disturb_cnt)
    if nat_cache is not None:
        nat_cache[cache_key] = result
    return result


def _naturalized_tbl(nat_metrics):
    total, nat_cnt, rate, disturb_cnt = nat_metrics
    tbl = _make_tbl(["총 분류군", "귀화식물", "생태계교란 생물", "귀화율"])
    tbl.setRowCount(1)
    tbl.setItem(0, 0, _item(total))
    tbl.setItem(0, 1, _item(_disp(nat_cnt)))
    tbl.setItem(0, 2, _item(_disp(disturb_cnt)))
    tbl.setItem(0, 3, _item("-" if nat_cnt == 0 else f"{rate:.{SETTINGS.decimal}f}"))
    _compact_table(tbl, 24)
    _auto_fit_table(tbl)
    _fit_table_no_scroll(tbl, False)
    _fit_table_height(tbl, 16)
    return tbl


def _draw_naturalized_compare(canvas, tbl_obj):
    if not _MPL: return
    tbl = _get_tbl(tbl_obj)
    if not tbl: return

    txt = tbl.item(0, len(_NAT_COMPARE_VALUES)).text()
    rate = 0.0 if txt == "-" else float(txt)

    _apply_mpl()
    canvas._fig.clf()
    ax = canvas._fig.add_subplot(111)

    decimal = canvas._cfg.get("decimal", SETTINGS.decimal)
    fs = canvas._cfg.get("bar_fontsize", SETTINGS.bar_fontsize)
    labels = list(_NAT_COMPARE_LABELS) + ["조사지역"]
    values = list(_NAT_COMPARE_VALUES) + [rate]
    colors = ["#5b9bd5"] * len(_NAT_COMPARE_VALUES) + ["#ff0000"]
    x = list(range(len(labels)))
    width = min(max(canvas._cfg.get("bar_v_width", 0.40), 0.08), 1.0)
    top_val = max(values) if values else 0.0

    bars = ax.bar(x, values, width=width, color=colors, edgecolor="none", zorder=3)
    for bar, value in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            value + 0.7,
            f"{value:.{decimal}f}",
            ha="center",
            va="bottom",
            fontsize=max(8, fs - 1),
        )

    boundary_ticks = [(left + right) / 2 for left, right in zip(x[:-1], x[1:])]

    axis_min = canvas._cfg.get("axis_min", SETTINGS.axis_min)
    axis_max = canvas._cfg.get("axis_max", SETTINGS.axis_max)
    axis_step = canvas._cfg.get("axis_step", SETTINGS.axis_step)
    
    c_min, c_max, c_step = axis_min, axis_max, axis_step
    effective_max = c_max if c_max > 0 else (top_val if top_val > 0 else 1.0)
    auto_max, auto_step = _get_nice_bounds(effective_max)
    if c_max <= 0: c_max = auto_max
    if c_step <= 0: c_step = auto_step

    ax.set_xticks(boundary_ticks)
    ax.set_xticklabels([""] * len(boundary_ticks))
    if c_min < c_max:
        ax.set_ylim(c_min, c_max)
    if c_step > 0:
        import matplotlib.ticker as ticker
        ax.yaxis.set_major_locator(ticker.MultipleLocator(c_step))
    ax.set_xlim(-0.5 - width * 0.2, len(labels) - 0.5 + width * 0.2)
    ax.tick_params(axis="x", which="major", pad=4, length=4, width=0.8, direction="out")
    ax.tick_params(axis="y", labelsize=max(8, fs))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_linewidth(1.0)
    ax.spines["left"].set_linewidth(1.0)
    for xi, label in zip(x, labels):
        ax.text(
            xi,
            -0.075,
            label,
            transform=ax.get_xaxis_transform(),
            ha="center",
            va="top",
            fontsize=max(8, fs),
        )
    if canvas._cfg.get("grid_on", SETTINGS.grid_on):
        ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35, zorder=0)
    canvas._fig.subplots_adjust(left=0.15, right=0.95, bottom=0.15, top=0.90)
    _apply_graph_scale(canvas._fig, canvas)
    canvas.draw()


def _naturalized_compare_tbl(nat_metrics):
    _, nat_cnt, rate, _ = nat_metrics
    headers = list(_NAT_COMPARE_LABELS) + ["조사지역", "도시화지수"]
    tbl = _make_tbl(headers)
    tbl.setRowCount(1)

    for col, value in enumerate(_NAT_COMPARE_VALUES):
        tbl.setItem(0, col, _item(f"{value:.1f}"))
    tbl.setItem(0, len(_NAT_COMPARE_VALUES), _item(f"{rate:.{SETTINGS.decimal}f}"))
    urbanization = nat_cnt / 321 * 100 if nat_cnt else 0.0
    tbl.setItem(0, len(_NAT_COMPARE_VALUES) + 1, _item("-" if not nat_cnt else f"{urbanization:.{SETTINGS.decimal}f}"))

    _compact_table(tbl, 24)
    _auto_fit_table(tbl)
    _fit_table_no_scroll(tbl, False)
    _fit_table_height(tbl, 16)
    return tbl


def _naturalized_panel(ctx, nat_sheet=None, disturb_sheet=None, parent_window=None):
    species = ctx["species"]
    rounds = ctx["rounds"]
    nat_metrics = ctx["nat_metrics"]
    panel = QWidget()
    lay = QVBoxLayout(panel)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.setSpacing(8)
    tbl_nat = _naturalized_tbl(nat_metrics)
    tbl_comp = _naturalized_compare_tbl(nat_metrics)
    panel._table_refs = {
        "naturalized": lambda tbl=tbl_nat: tbl,
        "naturalized_compare": lambda tbl=tbl_comp: tbl,
    }
    _add_tbl_with_btn(lay, tbl_nat, 0)
    _add_tbl_with_btn(lay, tbl_comp, 0)
    if _MPL:
        cv = _Canvas(7.2, 2.5)
        cv._setting_mode = "bar"
        cv._cfg["bar_v_width"] = 0.4
        cv._cfg["initial_graph_height"] = 250
        apply_graph_default(parent_window, cv, "plant_naturalized")
        cv._preview_draw = lambda target: _draw_naturalized_compare(target, tbl_comp)
        cv._refresh_cb = lambda: _draw_naturalized_compare(cv, tbl_comp)
        cv._refresh_cb()
        cv_grp = _wrap_cv("귀화율 그래프", cv)
        cv_grp._initial_height = 250
        panel_w = _make_bar_settings_panel(
            cv,
            cv_grp,
            lambda: _draw_naturalized_compare(cv, tbl_comp),
            title_default="귀화율 비교",
            y_default="귀화율(%)",
            parent=panel,
        )
        lay.addWidget(_make_settings_graph(panel_w, cv_grp), 1)
    return panel


def _wrap_cv(title, canvas):
    # 식물 그래프도 스크롤 영역 안에서 항상 중앙에 배치한다.
    return _common_wrap_canvas(title, canvas, alignment=Qt.AlignCenter)


def _bind_auto_max(grp, sld_w, sld_h):
    def _on_res(w, h):
        mw = max(300, w)
        mh = max(200, h)
        sld_w.setMaximum(mw)
        sld_h.setMaximum(mh)
        if not getattr(grp, "_init_max", False) and w > 150 and h > 150:
            grp._init_max = True
    grp.resized_sig.connect(_on_res)


def _make_settings_graph(panel_w, cv_grp):
    return _common_make_settings_below_graph(panel_w, cv_grp)

def _make_bar_settings_panel(canvas, cv_grp, draw_fn, title_default="", y_default="비율(%)", parent=None):
    canvas._refresh_cb = draw_fn
    return _common_make_bar_settings_panel(
        canvas,
        mode="bar",
        on_change=draw_fn,
        parent=parent,
    )


def _plant_counts_str(unit: dict) -> str:
    """SETTINGS 레벨 설정에 따라 식물 분류군 수 문자열 생성."""
    S = SETTINGS
    parts = []
    parts.append(f"{unit['과']}과")
    parts.append(f"{unit['속']}속")
    parts.append(f"{unit['종']}종")
    if S.sent_show_var   and unit['변종']: parts.append(f"{unit['변종']}변종")
    if S.sent_show_forma and unit['품종']: parts.append(f"{unit['품종']}품종")
    if S.sent_show_subsp and unit['아종']: parts.append(f"{unit['아종']}아종")
    show_taxa = getattr(S, "sent_show_taxa", True)
    base = " ".join(parts)
    if show_taxa and base:
        return f"{base}으로 총 {unit['분류군']}분류군"
    if show_taxa:
        return f"총 {unit['분류군']}분류군"
    if base:
        return base
    return f"{unit['분류군']}분류군"


def _list_limit_for(sentence_key: str, default=None):
    limits = getattr(SETTINGS, "plant_sent_list_limits", {}) or {}
    if sentence_key in limits:
        try:
            return int(limits.get(sentence_key))
        except Exception:
            return default
    if sentence_key in ("생태계교란 생물", "산림청지정 특산식물"):
        return 0
    return getattr(SETTINGS, "sent_species_limit", 3) if default is None else default


def _limit_for_sentence(names_list, sentence_key: str, limit=None):
    if not names_list:
        return ""
    if limit is None:
        limit = _list_limit_for(sentence_key)
    try:
        limit = int(limit)
    except Exception:
        limit = getattr(SETTINGS, "sent_species_limit", 3)
    if limit <= 0 or len(names_list) <= limit:
        return ", ".join(names_list)
    return ", ".join(names_list[:limit]) + " 등"


def _join_all(names_list):
    return ", ".join(names_list or [])


_ROMAN_TO_UNICODE = {"V": "Ⅴ", "IV": "Ⅳ", "III": "Ⅲ", "II": "Ⅱ", "I": "Ⅰ",
                     "5": "Ⅴ", "4": "Ⅳ", "3": "Ⅲ", "2": "Ⅱ", "1": "Ⅰ"}
_FLORISTIC_GRADES  = {"Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ", "Ⅴ"}
_RARE_GRADES       = {"EW", "CR", "EN", "VU", "LC", "DD"}
_RARE_GRADE_ORDER  = ["EW", "CR", "EN", "VU", "LC", "DD"]
_RARE_GRADE_LABELS = {
    "EW": "EW(야생멸종)",
    "CR": "CR(멸종위기종)",
    "EN": "EN(위기종)",
    "VU": "VU(취약종)",
    "LC": "LC(약관심종)",
    "DD": "DD(정보부족종)",
}

def _norm_grade(g: str) -> str:
    """등급 문자열 정규화: ASCII Roman·숫자 → Unicode Roman, '등급'/'급' 접미사 제거."""
    g = str(g or "").strip()
    if g.endswith("등급"): g = g[:-2].strip()
    if g.endswith("급"): g = g[:-1].strip()
    return _ROMAN_TO_UNICODE.get(g, g)


def _rare_grade_label(g: str) -> str:
    return _RARE_GRADE_LABELS.get(g, g)


def _limit(names_list, limit=None):
    """종 나열 개수 제한: limit 초과 시 앞 N개 + ' 등'."""
    if limit is None:
        limit = getattr(SETTINGS, "sent_species_limit", 3)
    if not names_list:
        return ""
    if len(names_list) > limit:
        return ", ".join(names_list[:limit]) + " 등"
    return ", ".join(names_list)


def _build_plant_sentences(mode_label, ctx, prot_sheet=None, aux_sheets=None, force_norm=False, round_label="전체", num_rounds=1, tbl_ref=None) -> dict:
    if aux_sheets is None: aux_sheets = {}
    S = SETTINGS
    species = ctx["species"]
    rounds = ctx["rounds"]
    unit_counts = ctx["unit_counts"]
    life = ctx["life_counts"]
    tbl_ref = tbl_ref or {}

    def _resolve_tbl(name):
        ref = tbl_ref.get(name)
        if callable(ref):
            try:
                return ref()
            except Exception:
                return None
        return ref

    def _int_cell(tbl, row, col, default=0):
        if not tbl:
            return default
        it = tbl.item(row, col)
        if not it:
            return default
        txt = str(it.text() or "").replace(",", "").strip()
        if txt in ("", "-"):
            return default
        try:
            return int(float(txt))
        except ValueError:
            return default

    def _float_cell(tbl, row, col, default=0.0):
        if not tbl:
            return default
        it = tbl.item(row, col)
        if not it:
            return default
        txt = str(it.text() or "").replace(",", "").replace("%", "").strip()
        if txt in ("", "-"):
            return default
        try:
            return float(txt)
        except ValueError:
            return default

    def _unit_from_count_tbl(tbl):
        if not tbl:
            return None
        target_row = -1
        for r in range(tbl.rowCount()):
            it = tbl.item(r, 0)
            if it and str(it.text()).strip() == "종합":
                target_row = r
                break
        if target_row < 0 and tbl.rowCount() > 0:
            target_row = tbl.rowCount() - 1
        if target_row < 0:
            return None
        return {
            "과": _int_cell(tbl, target_row, 1),
            "속": _int_cell(tbl, target_row, 2),
            "종": _int_cell(tbl, target_row, 3),
            "변종": _int_cell(tbl, target_row, 4),
            "품종": _int_cell(tbl, target_row, 5),
            "아종": _int_cell(tbl, target_row, 6),
            "분류군": _int_cell(tbl, target_row, 7),
        }

    count_tbl = _resolve_tbl("count")
    life_tbl = _resolve_tbl("life")
    nat_tbl = _resolve_tbl("naturalized")
    
    is_field = (mode_label == "현지조사")
    intro_mode = getattr(
        SETTINGS,
        "field_intro_mode" if is_field else "lit_intro_mode",
        getattr(SETTINGS, "field_intro_mode", "auto"),
    )
    
    if round_label == "전체":
        if is_field:
            base = "현지조사"
        else:
            base = "문헌조사"
        
        prefix_res = f"{base} 결과"
        prefix_res_comma = f"{base} 결과,"
        prefix_shi = f"{base}시"
    else:
        fixed_intro = intro_mode == "fixed"
        if is_field:
            base = "현지조사" if fixed_intro else _round_survey_label(round_label)
            prefix_res = f"{base} 결과"
            prefix_res_comma = f"{base} 결과,"
            prefix_shi = f"{base}시"
        else:
            if fixed_intro:
                prefix_res = "문헌조사 결과"
                prefix_res_comma = "문헌조사 결과,"
                prefix_shi = "문헌조사 시"
            else:
                base = _round_survey_label(round_label)
                prefix_res = lit_result_prefix(rounds, base, comma=True)
                prefix_res_comma = lit_result_prefix(rounds, base, comma=True)
                prefix_shi = lit_shi_prefix(rounds, base)

    s1_mid = getattr(S, "plant_field_s1_mid", S.field_s1_mid) if is_field else getattr(S, "plant_lit_s1_mid", S.lit_s1_mid)
    s1_end = getattr(S, "plant_field_s1_end", S.field_s1_end) if is_field else getattr(S, "plant_lit_s1_end", S.lit_s1_end)
    order_verb_f = getattr(S, "plant_order_verb_f", S.order_verb_f)
    order_next = getattr(S, "plant_order_next", S.order_next)

    # 식물상 counts 문자열로 이/가 조사 결정
    unit = _unit_from_count_tbl(count_tbl) or unit_counts["종합"]
    counts = _plant_counts_str(unit)
    s1_mid_josa = _resolve_iga(s1_mid, counts)
    s1_end_base = s1_end.replace("이 ", "").replace("가 ", "")
    
    verb_past = order_verb_f.replace('으며', '음')
    if "음" not in verb_past and verb_past.endswith("었으며"): verb_past = verb_past[:-2] + "음"
    if "음" not in verb_past and verb_past.endswith("었어"): verb_past = verb_past[:-2] + "음"
    verb_ongoing = (verb_past[:-1] + "으며,") if verb_past.endswith("음") else (verb_past + "으며,")
    
    prot_none = S.prot_none_field if is_field else S.prot_none_lit

    sentences = OrderedDict()

    # 1. 식물상 집계
    _sp_names_main = {sp.kor_name for sp in species if sp.kor_name}
    _extinct_sheet = aux_sheets.get("멸종위기종")
    _rare_sheet_s1    = aux_sheets.get("희귀식물") or aux_sheets.get("희귀·특산식물")
    _endemic_sheet_s1 = aux_sheets.get("특산식물") or aux_sheets.get("희귀·특산식물")
    _prot_mode = getattr(SETTINGS, "prot_grade_mode", "short")
    _mode_key = "현지" if is_field else "문헌"

    def _ext_grade_label(sp):
        g = _norm_grade(getattr(sp, "grade", ""))
        if _prot_mode == "none": return ""
        if _prot_mode == "full":
            if g == "Ⅰ": return "멸종위기 야생생물 Ⅰ급"
            if g == "Ⅱ": return "멸종위기 야생생물 Ⅱ급"
            return ""
        # short
        if g == "Ⅰ": return "멸Ⅰ"
        if g == "Ⅱ": return "멸Ⅱ"
        return ""

    # 멸종위기 단독 문장용 목록
    _ext_sp = []
    if _extinct_sheet:
        _ext_sp = [
            sp for sp in _extinct_sheet.species
            if sp.kor_name in _sp_names_main
            and any(_has(v) for k, v in sp.rounds.items() if _mode_key in k)
        ]
        if not _ext_sp:
            _ext_sp = [sp for sp in _extinct_sheet.species if sp.kor_name in _sp_names_main]
        if _ext_sp:
            e_names = [sp.kor_name for sp in _ext_sp]
            sentences["멸종위기 야생생물"] = f"{prefix_res} 멸종위기 야생생물은 {_limit_for_sentence(e_names, '멸종위기 야생생물')} 총 {len(_ext_sp)}분류군이 {s1_end_base}"
        else:
            sentences["멸종위기 야생생물"] = f"{prefix_res} 멸종위기 야생생물은 {prot_none}"

    # ── 통합 보호종(멸종위기+희귀+특산) 파트 계산 ──────────────────────────────
    from collections import OrderedDict as _OD_s1
    _sp_labels_s1 = _OD_s1()

    def _s1_add(name, label):
        if name not in _sp_labels_s1: _sp_labels_s1[name] = []
        if label and label not in _sp_labels_s1[name]: _sp_labels_s1[name].append(label)

    if _extinct_sheet:
        for sp in (_ext_sp if _extinct_sheet and _ext_sp else []):
            _s1_add(sp.kor_name, _ext_grade_label(sp))
    elif prot_sheet:
        from shared import _prot_grade_str as _pgs_s1
        for sp in prot_sheet.species:
            if sp.kor_name in _sp_names_main:
                _s1_add(sp.kor_name, _pgs_s1(sp))

    _rare_lbl    = "" if _prot_mode == "none" else ("희귀" if _prot_mode == "short" else "희귀식물")
    _endemic_lbl = "" if _prot_mode == "none" else ("특산" if _prot_mode == "short" else "특산식물")

    if _rare_sheet_s1:
        _is_comb_r = "특산" in (_rare_sheet_s1.name or "") and "희귀" in (_rare_sheet_s1.name or "")
        for sp in _rare_sheet_s1.species:
            if not sp.kor_name or sp.kor_name not in _sp_names_main: continue
            if not any(_has(sp.rounds.get(r)) for r in rounds): continue
            cat = _s(getattr(sp, "category", ""))
            if _is_comb_r and "특산" in cat and "희귀" not in cat and not any(rg in cat.upper() for rg in _RARE_GRADES):
                continue
            _s1_add(sp.kor_name, _rare_lbl)

    if _endemic_sheet_s1:
        _is_comb_e = "희귀" in (_endemic_sheet_s1.name or "")
        for sp in _endemic_sheet_s1.species:
            if not sp.kor_name or sp.kor_name not in _sp_names_main: continue
            if not any(_has(sp.rounds.get(r)) for r in rounds): continue
            cat = _s(getattr(sp, "category", ""))
            if _is_comb_e and "특산" not in cat: continue
            _s1_add(sp.kor_name, _endemic_lbl)

    _comb_parts = [
        f"{n}({', '.join(lbls)})" if lbls else n
        for n, lbls in _sp_labels_s1.items()
    ]
    _comb_total = len(_sp_labels_s1)

    if _comb_total > 0:
        disp = _limit_for_sentence(_comb_parts, "식물상")
        prot_part = f"{disp} 총 {_comb_total}분류군{s1_end}"
    else:
        prot_part = prot_none

    s_main = f"{prefix_res} 확인된 소산식물은 {counts}{s1_mid_josa} 법정보호종은 {prot_part}"

    sentences["식물상"] = s_main

    if is_field:
        # 2. 생활형
        total_life = sum(life.values())
        if total_life > 0:
            disp_map = {"M":"대형지상식물(M)", "N":"소형육상식물(N)", "E":"착생식물(E)", "CH":"지표식물(Ch)", "H":"반지중식물(H)", "G":"지중식물(G)", "HH":"수생식물(HH)", "TH":"일년생식물(Th)"}
            
            counts_orig = [life.get(k, 0) for k in _LIFE_KEYS]
            if life_tbl:
                pcts_orig = [_float_cell(life_tbl, 1, ci, 0.0) for ci in range(1, len(_LIFE_KEYS) + 1)]
            elif force_norm:
                pcts_orig, _ = _apply_ratio_correction(_normalized_percentages(counts_orig))
            else:
                pcts_orig = [c / total_life * 100 if total_life else 0 for c in counts_orig]
            life_pct_map = dict(zip(_LIFE_KEYS, pcts_orig))

            def _life_value_text(cnt, pct, shared=False):
                fmt = getattr(SETTINGS, "plant_life_item_format", "count_pct")
                if fmt == "pct_count":
                    return f"{_pct(pct)}%({cnt}분류군)"
                return f"{cnt}분류군({_pct(pct)}%)"

            grouped = OrderedDict()
            for code in sorted(_LIFE_KEYS, key=lambda k: (-life.get(k, 0), _LIFE_KEYS.index(k))):
                cnt = life.get(code, 0)
                pct = life_pct_map.get(code, 0.0)
                if cnt <= 0:
                    continue
                grouped.setdefault((cnt, round(pct, 1)), []).append(code)

            parts = []
            for (cnt, pct), codes in grouped.items():
                names = [disp_map.get(code, code) for code in codes]
                val_txt = _life_value_text(cnt, pct, shared=len(codes) > 1)
                if len(codes) > 1:
                    parts.append(f"{' 및 '.join(names)} 각 {val_txt}")
                else:
                    parts.append(f"{names[0]} {val_txt}")

            zero_names = [disp_map.get(code, code) for code in _LIFE_KEYS if life.get(code, 0) <= 0]

            if parts:
                s_life_1 = f"조사지역에서 확인된 관속식물의 생활형을 분석한 결과 {', '.join(parts)}의 순으로 조사되었"
                if zero_names:
                    s_life_1 += f"으며, {', '.join(zero_names)}는 확인되지 않았음"
                else:
                    s_life_1 += "음"
                
                high_parts, low_parts = [], []
                south_vals = _LIFE_FIXED_SOUTH
                for code in _LIFE_KEYS:
                    if life.get(code, 0) <= 0:
                        continue
                    pct = life_pct_map[code.upper()]
                    s_pct = south_vals.get(code, 0)
                    name = disp_map.get(code.upper(), code)
                    if pct - s_pct > 0: high_parts.append(name)
                    elif pct - s_pct < 0: low_parts.append(name)
                
                s_life_2 = "본 조사지역의 생태계 구조와 입지환경을 파악하기 위해 남한지역의 생활형 Spectrum과 비교한 결과, "
                if high_parts: s_life_2 += f"{', '.join(high_parts)}은 높게 {verb_ongoing} "
                if low_parts: s_life_2 += f"{', '.join(low_parts)}은 낮게 {s1_end_base}"
                s_life_2 = s_life_2.rstrip(", ") + "." if s_life_2.endswith(" ") else s_life_2
                
                sentences["소산식물 생활형 현황"] = s_life_1 + "\n" + s_life_2.rstrip(", ")
            
        # 3. 양치식물
        fern_unit = unit_counts.get("양치식물문", {})
        if fern_unit.get("분류군", 0) > 0:
            f_cnt = fern_unit["분류군"]
            f_fam = fern_unit["과"]
            s_fern_1 = f"{prefix_res_comma} 조사지역의 양치식물은 {f_fam}과 {f_cnt}분류군이 {verb_past}"
            
            gymno_cnt = unit_counts.get("나자식물문", {}).get("분류군", 0)
            angio_cnt = unit_counts.get("피자식물문", {}).get("분류군", 0)
            seed_cnt = gymno_cnt + angio_cnt
            if seed_cnt > 0:
                pte_q = 25 * f_cnt / seed_cnt
                comp_str = "높게" if pte_q > 1.4 else "낮게"
                _pte_str = f"{pte_q:.2f}"
                s_fern_2 = f"환경변화에 매우 민감하게 생리·생태적인 반응을 하여 자연의 파괴정도를 나타내는 지수로 사용되는 양치식물계수(Pte·Q=25×양치식물종수/종자식물종수)는 {_pte_str}{_josa(_pte_str, '으로로')} 한반도 전체 1.4(임과 이, 1976)와 비교해 볼 때 {comp_str} {s1_end_base}"
                sentences["양치식물 분포현황"] = s_fern_1 + "\n" + s_fern_2
            else:
                sentences["양치식물 분포현황"] = s_fern_1

    # 4. 귀화식물
    nat_sheet = aux_sheets.get("귀화식물")
    if nat_sheet:
        nat_species = _species_in_rounds(nat_sheet.species, rounds, ctx["cache"].get("round_species"))
        if nat_species:
            nat_fams = len(set(sp.family for sp in nat_species if sp.family))
            nat_cnt = len(nat_species)
            nat_rate_from_tbl = None
            disturb_cnt_from_tbl = None
            if nat_tbl:
                nat_cnt = _int_cell(nat_tbl, 0, 1, nat_cnt)
                disturb_cnt_from_tbl = _int_cell(nat_tbl, 0, 2, None)
                nat_rate_from_tbl = _float_cell(nat_tbl, 0, 3, None)
            
            if is_field:
                ui = nat_cnt / 321 * 100
                ni = nat_rate_from_tbl if nat_rate_from_tbl is not None else (nat_cnt / unit["분류군"] * 100 if unit["분류군"] else 0)

                s_nat_1 = f"{prefix_res_comma} 조사지역의 귀화식물은 총 {nat_fams}과 {nat_cnt}분류군이 확인되었으며, 도시화지수는 {_pct(ui)}%, 귀화율은 {_pct(ni)}%로 {S.end_ana}"

                # 동사: end_field 어미에 맞춰 나타남/나타났음/나타났다 결정
                _ef = (S.end_field or "").strip()
                if _ef.endswith("되었다") or _ef.endswith("었다") or _ef.endswith("다"):
                    _appear = "나타났다"
                elif _ef.endswith("되었음") or _ef.endswith("었음") or _ef.endswith("음"):
                    _appear = "나타났음"
                else:
                    _appear = "나타남"

                from bisect import bisect_right
                rev_vals = _NAT_COMPARE_VALUES[::-1]
                rev_labels = _NAT_COMPARE_LABELS[::-1]
                idx = bisect_right(rev_vals, ni)
                if idx == 0:
                    comp_s = f"{rev_labels[0]}({rev_vals[0]}%)보다 낮게 {_appear}"
                elif idx == len(rev_vals):
                    comp_s = f"{rev_labels[-1]}({rev_vals[-1]}%)보다 높게 {_appear}"
                else:
                    comp_s = f"{rev_labels[idx]}({rev_vals[idx]}%)보다 낮고, {rev_labels[idx-1]}({rev_vals[idx-1]}%)보다 높게 {_appear}"

                s_nat_2 = f"입지별 평균 귀화율과 비교하면 {comp_s}"
                sentences["귀화식물"] = s_nat_1 + "\n" + s_nat_2
            else:
                names = [sp.kor_name for sp in nat_species if sp.kor_name]
                sentences["귀화식물"] = f"{prefix_res} 귀화식물은 {_limit_for_sentence(names, '귀화식물')} 총 {nat_cnt}분류군이 {s1_end_base}"

    # 5. 교란생물
    disturb_sheet = aux_sheets.get("교란생물")
    if disturb_sheet:
        disturb_species = _species_in_rounds(disturb_sheet.species, rounds, ctx["cache"].get("round_species"))
        if disturb_species:
            sp_names = [sp.kor_name for sp in disturb_species if sp.kor_name]
            dis_cnt = len(disturb_species)
            if nat_tbl:
                dis_cnt = _int_cell(nat_tbl, 0, 2, dis_cnt)
            s_dis = f"{prefix_res} 생태계교란 생물은 {_join_all(sp_names)} 총 {dis_cnt}분류군이 {s1_end_base}"
            sentences["생태계교란 생물"] = s_dis

    # 6. 희귀·특산식물
    # 데이터 소스: "희귀식물" 시트 우선, 없으면 "희귀·특산식물" 합본 폴백
    # 구계학적 특정식물 시트는 사용하지 않음
    special_sheet_rare    = aux_sheets.get("희귀식물")
    _rare_is_combined     = False
    if not special_sheet_rare:
        special_sheet_rare = aux_sheets.get("희귀·특산식물")
        _rare_is_combined  = True

    special_sheet_endemic = aux_sheets.get("특산식물")
    _end_is_combined      = False
    if not special_sheet_endemic:
        special_sheet_endemic = aux_sheets.get("희귀·특산식물")
        _end_is_combined      = True

    from collections import defaultdict

    def _normalize_rare_grade(cat: str) -> str:
        """A열 carry-forward 등급 값을 표준 코드로 정규화 (CR/EN/VU/NT/LC/DD → 그대로, 기타는 '기타')."""
        c = _s(cat).strip()
        for g in _RARE_GRADE_ORDER:
            if c == g or c.upper().startswith(g):
                return g
        return "기타"

    # ── 희귀식물 등급별 수집 (A열 carry-forward 구조 기반) ──────────────────────
    r_cat_names: dict = defaultdict(list)
    seen_rare: set    = set()

    if special_sheet_rare:
        for sp in special_sheet_rare.species:
            if not sp.kor_name: continue
            # 현재 차수(현지/문헌) 출현 여부
            if not any(_has(sp.rounds.get(r)) for r in rounds): continue
            cat = _s(sp.category).strip()
            # 합본 시트인 경우: "특산" 카테고리 행은 희귀식물 처리에서 제외
            if _rare_is_combined and "특산" in cat:
                continue
            grade_key = _normalize_rare_grade(cat)
            if sp.kor_name not in seen_rare:
                seen_rare.add(sp.kor_name)
                r_cat_names[grade_key].append(sp.kor_name)

    if r_cat_names:
        sorted_cats = sorted(
            r_cat_names.keys(),
            key=lambda x: _RARE_GRADE_ORDER.index(x) if x in _RARE_GRADE_ORDER else len(_RARE_GRADE_ORDER)
        )
        cat_strs   = [f"{_rare_grade_label(c)} {len(r_cat_names[c])}분류군({_limit_for_sentence(r_cat_names[c], '산림청지정 희귀식물')})" for c in sorted_cats]
        total_rare = sum(len(v) for v in r_cat_names.values())
        sentences["산림청지정 희귀식물"] = (
            f"{prefix_res} 산림청지정 희귀식물은 {', '.join(cat_strs)} 등 "
            f"총 {total_rare}분류군이 {s1_end_base}"
        )
    elif special_sheet_rare:
        sentences["산림청지정 희귀식물"] = f"{prefix_res} 산림청지정 희귀식물은 {prot_none}"

    # ── 특산식물 (특산식물 시트 기준, 합본 시트는 "특산" 카테고리만 필터) ────────
    if special_sheet_endemic:
        seen_end: set       = set()
        endemic_names: list = []
        # 순수 특산식물 전용 시트: 카테고리 필터 불필요
        _endemic_only = (
            "특산" in special_sheet_endemic.name
            and "희귀" not in special_sheet_endemic.name
        )
        for sp in special_sheet_endemic.species:
            if not sp.kor_name: continue
            if not any(_has(sp.rounds.get(r)) for r in rounds): continue
            # 합본 또는 혼합 시트: "특산" 카테고리 행만 포함
            if not _endemic_only and _end_is_combined:
                if "특산" not in _s(sp.category):
                    continue
            if sp.kor_name not in seen_end:
                seen_end.add(sp.kor_name)
                endemic_names.append(sp.kor_name)
        if endemic_names:
            sentences["산림청지정 특산식물"] = (
                f"{prefix_res} 산림청지정 특산식물은 {_limit_for_sentence(endemic_names, '산림청지정 특산식물')} "
                f"총 {len(endemic_names)}분류군이 {s1_end_base}"
            )
        else:
            sentences["산림청지정 특산식물"] = f"{prefix_res} 산림청지정 특산식물은 {prot_none}"

    # 7. 식물구계학적 특정종 (Ⅰ-Ⅴ등급만, CR/EN/VU 등 희귀식물 등급 제외)
    floristic_sheet = aux_sheets.get("구계학적 특정식물")
    if floristic_sheet:
        _sp_names = {sp.kor_name for sp in species if sp.kor_name}
        # 구계학적 등급(Ⅰ-Ⅴ)만 필터 — 희귀식물 등급(CR/EN/VU 등)은 제외
        fl_species = [sp for sp in floristic_sheet.species
                      if sp.kor_name in _sp_names
                      and _norm_grade(getattr(sp, "grade", "")) in _FLORISTIC_GRADES]
        if fl_species:
            fl_cat_names = defaultdict(list)
            for sp in fl_species:
                g = _norm_grade(getattr(sp, "grade", ""))
                fl_cat_names[g + "등급"].append(sp.kor_name or _clean_sci_name(sp) or "?")

            grade_parts = []
            for g in ["Ⅴ등급", "Ⅳ등급", "Ⅲ등급", "Ⅱ등급", "Ⅰ등급"]:
                names = fl_cat_names.get(g, [])
                cnt = len(names)
                if cnt:
                    disp = _limit_for_sentence(names, "식물구계학적 특정종")
                    grade_parts.append(f"{g} {disp} {cnt}분류군")

            v_cnt  = len(fl_cat_names.get("Ⅴ등급", []))
            iv_cnt = len(fl_cat_names.get("Ⅳ등급", []))
            notice = ""
            missing_high = []
            legacy_notice = getattr(SETTINGS, "plant_floristic_high_grade_notice", True)
            if legacy_notice and getattr(SETTINGS, "plant_floristic_notice_grade_v", True) and v_cnt == 0:
                missing_high.append("Ⅴ등급")
            if legacy_notice and getattr(SETTINGS, "plant_floristic_notice_grade_iv", True) and iv_cnt == 0:
                missing_high.append("Ⅳ등급")
            if missing_high:
                notice = f"상대적으로 보호 가치가 높은 {', '.join(missing_high)} 분류군은 조사되지 않았으며, "
            s_fl = f"{prefix_res} {notice}식물 구계학적 특정식물은 {', '.join(grade_parts)}으로 총 {len(fl_species)}분류군이 {s1_end_base}"
            sentences["식물구계학적 특정종"] = s_fl
        else:
            sentences["식물구계학적 특정종"] = f"{prefix_res} 식물구계학적 특정종은 {prot_none}"
            
    for k, v in sentences.items():
        sentences[k] = _apply_title(v, S.sentence_title)
        
    return sentences



def _section_page(title, widget):
    page = QWidget()
    lay = QVBoxLayout(page)
    lay.setContentsMargins(6, 6, 6, 6)
    lay.setSpacing(8)
    grp = QGroupBox(title)
    gl = QVBoxLayout(grp)
    gl.setContentsMargins(6, 12, 6, 6)
        
    if hasattr(widget, "copy_selection"):
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0,0,0,0)
        btn_row.addStretch()
        btn_copy = QPushButton("📋 표 복사")
        btn_copy.setFixedHeight(24)
        btn_copy.setStyleSheet(_COPY_BTN_QSS)
        def _do_copy():
            widget.copy_selection(include_header=True)
            from ui_shared import apply_button_feedback
            apply_button_feedback(btn_copy)
        btn_copy.clicked.connect(_do_copy)
        btn_row.addWidget(btn_copy)
        gl.addLayout(btn_row)
    gl.addWidget(widget)
    lay.addWidget(grp)
    lay.addStretch()
    return page


def _combined_trait_page(ctx, nat_sheet=None, disturb_sheet=None, on_norm_cb=None, parent_window=None):
    page = QWidget()
    lay = QHBoxLayout(page)
    lay.setContentsMargins(6, 6, 6, 6)
    lay.setSpacing(8)

    left_grp = QGroupBox("생활형별")
    left_lay = QVBoxLayout(left_grp)
    left_lay.setContentsMargins(6, 18, 6, 6)
    life_panel = _life_panel(ctx, on_norm_cb=on_norm_cb, parent_window=parent_window)
    left_lay.addWidget(life_panel)

    right_grp = QGroupBox("귀화율")
    right_lay = QVBoxLayout(right_grp)
    right_lay.setContentsMargins(6, 18, 6, 6)
    nat_panel = _naturalized_panel(ctx, nat_sheet, disturb_sheet, parent_window=parent_window)
    right_lay.addWidget(nat_panel)

    spl = QSplitter(Qt.Horizontal)
    spl.setChildrenCollapsible(False)
    spl.addWidget(left_grp)
    spl.addWidget(right_grp)
    spl.setStretchFactor(0, 1)
    spl.setStretchFactor(1, 1)
    spl.setSizes([500, 500])
    lay.addWidget(spl)
    page._table_refs = {}
    page._table_refs.update(getattr(life_panel, "_table_refs", {}))
    page._table_refs.update(getattr(nat_panel, "_table_refs", {}))
    return page

def _lit_panel_plant(species_all, rounds_all, prot_sheet=None, nat_sheet=None, disturb_sheet=None, aux_sheets=None, parent_window=None):
    """문헌조사 패널: 식물상 / 귀화식물 / 생태계교란 생물 / 희귀식물 / 구계특정종 탭."""
    from ui_shared import make_copy_button as _mcb
    from shared import _prot_grade_str as _pgs
    aux_sheets = aux_sheets or {}
    rare_sheet    = aux_sheets.get("희귀·특산식물") or aux_sheets.get("희귀식물")
    endemic_sheet = aux_sheets.get("희귀·특산식물") or aux_sheets.get("특산식물")
    floristic_sheet = aux_sheets.get("구계학적 특정식물")
    extinct_sheet = aux_sheets.get("멸종위기종")

    w = QWidget(); w.setStyleSheet(_QSS)
    lay = QVBoxLayout(w); lay.setContentsMargins(6, 6, 6, 6); lay.setSpacing(0)

    tabs = QTabWidget()
    tabs.setDocumentMode(True)
    tabs.setStyleSheet(_tab_qss(False))

    ind_rns = [r for r in rounds_all if "_합계" not in r]

    def _sp_of(rk):
        return species_all if rk == "_all" else [sp for sp in species_all if _has(sp.rounds.get(rk))]

    def _sp_in_any_rns(rns):
        return [sp for sp in species_all if any(_has(sp.rounds.get(r)) for r in rns)]

    def _row_data():
        rows = []
        group_defs = parent_window.get_active_groups(rounds_all) if parent_window and hasattr(parent_window, "get_active_groups") else []
        if group_defs:
            for group_name, group_rns in group_defs:
                sp_g = _sp_in_any_rns(group_rns)
                if sp_g:
                    rows.append((group_name, sp_g, True))
            return rows
        if ind_rns:
            for rk in ind_rns:
                rows.append((_rnd_label(rk), _sp_of(rk), False))
            rows.append(("전체", species_all, True))
        else:
            rows.append(("문헌조사", species_all, False))
        return rows

    def _combined_prot_info(sp_rn):
        """(parts, total) — 멸종위기+희귀+특산 각 종별 라벨 통합 (prot_grade_mode 적용)."""
        from collections import OrderedDict as _OD
        _mode = getattr(SETTINGS, "prot_grade_mode", "short")
        sp_names = {sp.kor_name for sp in sp_rn if sp.kor_name}
        sp_labels = _OD()

        def _add(name, label):
            if name not in sp_labels: sp_labels[name] = []
            if label and label not in sp_labels[name]: sp_labels[name].append(label)

        def _ext_lbl(g):
            if _mode == "none": return ""
            if _mode == "full":
                return "멸종위기 야생생물 Ⅰ급" if g == "Ⅰ" else ("멸종위기 야생생물 Ⅱ급" if g == "Ⅱ" else "")
            return "멸Ⅰ" if g == "Ⅰ" else ("멸Ⅱ" if g == "Ⅱ" else "")

        _rare_lbl_l    = "" if _mode == "none" else ("희귀" if _mode == "short" else "희귀식물")
        _endemic_lbl_l = "" if _mode == "none" else ("특산" if _mode == "short" else "특산식물")

        if extinct_sheet:
            for sp in extinct_sheet.species:
                if sp.kor_name in sp_names:
                    g = _norm_grade(getattr(sp, "grade", ""))
                    _add(sp.kor_name, _ext_lbl(g))
        elif prot_sheet:
            for sp in prot_sheet.species:
                if sp.kor_name in sp_names:
                    lbl = _pgs(sp) if _mode != "none" else ""
                    _add(sp.kor_name, lbl)

        if rare_sheet:
            _is_comb_r = "특산" in (rare_sheet.name or "") and "희귀" in (rare_sheet.name or "")
            for sp in rare_sheet.species:
                if not sp.kor_name or sp.kor_name not in sp_names: continue
                cat = _s(getattr(sp, "category", ""))
                if _is_comb_r and "특산" in cat and "희귀" not in cat and not any(rg in cat.upper() for rg in _RARE_GRADES):
                    continue
                _add(sp.kor_name, _rare_lbl_l)

        if endemic_sheet:
            _is_comb_e = "희귀" in (endemic_sheet.name or "")
            for sp in endemic_sheet.species:
                if not sp.kor_name or sp.kor_name not in sp_names: continue
                cat = _s(getattr(sp, "category", ""))
                if _is_comb_e and "특산" not in cat: continue
                _add(sp.kor_name, _endemic_lbl_l)

        parts = [
            f"{n}({', '.join(lbls)})" if lbls else n
            for n, lbls in sp_labels.items()
        ]
        return parts, len(sp_labels)

    def _combined_prot_text(sp_rn):
        """법정보호종+희귀+특산 통합 표시 문자열 (테이블 셀용)."""
        parts, total = _combined_prot_info(sp_rn)
        if not total:
            return "-"
        lim = getattr(SETTINGS, "sent_species_limit", 3)
        disp = ", ".join(parts[:lim]) + (" 등" if len(parts) > lim else "")
        return f"{disp} 총 {total}분류군"

    # ── 탭 1: 식물상 ──────────────────────────────────────────────────────────
    page_plant = QWidget(); page_plant.setStyleSheet(_QSS)
    lay_plant = QVBoxLayout(page_plant); lay_plant.setContentsMargins(6, 6, 6, 6); lay_plant.setSpacing(8)

    grp_t = QGroupBox("문헌별 요약")
    gt = QVBoxLayout(grp_t); gt.setContentsMargins(6, 12, 6, 6); gt.setSpacing(4)

    tbl_sum = _make_tbl(["구분", "분류군 현황", "법정보호종"])
    tbl_sum.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

    def _fill_tbl():
        rows = _row_data()
        tbl_sum.setRowCount(max(len(rows), 1))
        for ri, (lbl, sp_rn, is_total) in enumerate(rows):
            unit   = _unit_counts(sp_rn).get("종합", {})
            counts = _plant_counts_str(unit) if unit else f"{len(sp_rn)}분류군"
            prot_text = _combined_prot_text(sp_rn)
            tbl_sum.setItem(ri, 0, _item(lbl, Qt.AlignCenter))
            tbl_sum.setItem(ri, 1, _item(counts, Qt.AlignCenter))
            tbl_sum.setItem(ri, 2, _item(prot_text, Qt.AlignLeft | Qt.AlignVCenter))
            if is_total:
                _bold_row(tbl_sum, ri)
        _tbl_auto_height(tbl_sum, 26)

    _fill_tbl()
    _auto_fit_table(tbl_sum)
    _compact_table(tbl_sum, 26)
    page_plant._table_refs = {"count": lambda tbl=tbl_sum: tbl}
    gt.addWidget(_mcb(tbl_sum))
    gt.addWidget(tbl_sum)
    lay_plant.addWidget(grp_t, 0)

    def _build_lit_lines():
        S = SETTINGS
        s1_mid = getattr(S, "plant_lit_s1_mid", S.lit_s1_mid)
        s1_end = getattr(S, "plant_lit_s1_end", S.lit_s1_end)
        lines = []

        def _prot_part_for(sp_rn):
            parts, total = _combined_prot_info(sp_rn)
            if not total: return S.prot_none_lit
            lim = getattr(S, "sent_species_limit", 3)
            disp = ", ".join(parts[:lim]) + (" 등" if len(parts) > lim else "")
            return f"{disp} 총 {total}분류군{s1_end}"

        num_rounds = len(ind_rns)
        if num_rounds > 1:
            unit   = _unit_counts(species_all).get("종합", {})
            counts = _plant_counts_str(unit) if unit else f"{len(species_all)}분류군"
            prefix = "문헌조사"
            lines.append(f"{prefix} 결과 {counts}{s1_mid} 법정보호종은 {_prot_part_for(species_all)}")

        for lbl, sp_rn, is_total in _row_data():
            if is_total: continue
            unit   = _unit_counts(sp_rn).get("종합", {})
            counts = _plant_counts_str(unit) if unit else f"{len(sp_rn)}분류군"
            if getattr(S, "lit_intro_mode", getattr(S, "field_intro_mode", "auto")) == "fixed":
                prefix = "문헌조사 결과,"
            else:
                prefix = lit_result_prefix([lbl], _round_survey_label(lbl), comma=True)
            lines.append(f"{prefix} {counts}{s1_mid} 법정보호종은 {_prot_part_for(sp_rn)}")
        return _apply_title("\n".join(lines), S.sentence_title)

    grp_s = QGroupBox("문장")
    gs = QVBoxLayout(grp_s); gs.setContentsMargins(8, 20, 8, 8); gs.setSpacing(6)
    n_lit = max(len(ind_rns) + (1 if len(ind_rns) > 1 else 0), 1)
    txt = QPlainTextEdit(); txt.setReadOnly(True)
    txt.setFixedHeight(min(n_lit * 42 + 20, 320))
    txt.setStyleSheet(f"QPlainTextEdit{{background:#FFFFFF;border:1.5px solid {_BORDER};border-radius:8px;{FF_KR};font-size:12px;padding:6px;}}")
    txt.setPlainText(_build_lit_lines())

    btn_s = QPushButton("📋  복사"); btn_s.setFixedHeight(30); btn_s.setStyleSheet(_COPY_BTN_QSS)
    def _copy_s():
        QApplication.clipboard().setText(txt.toPlainText())
        from ui_shared import apply_button_feedback; apply_button_feedback(btn_s)
    btn_s.clicked.connect(_copy_s)

    def _refresh_all():
        _fill_tbl()
        txt.setPlainText(_build_lit_lines())
    txt._sent_refresh_fn = _refresh_all

    br = QHBoxLayout(); br.addWidget(btn_s); br.addStretch()
    gs.addLayout(br); gs.addWidget(txt)
    lay_plant.addWidget(grp_s, 0)
    lay_plant.addStretch()
    tabs.addTab(_sc(page_plant), "식물상")

    lay.addWidget(tabs, 1)
    return w


def _survey_tab(mode_label, species_all, rounds_all, aux_sheets=None, parent_window=None, sheet_name="", sentence_table_registry=None, progress_share=1.0):
    wrap = QWidget()
    lay = QVBoxLayout(wrap)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.setSpacing(0)

    main_tabs = QTabWidget()
    main_tabs.setDocumentMode(True)
    main_tabs.setStyleSheet(_tab_qss(False))

    tab_count = QTabWidget()
    tab_count.setDocumentMode(True)
    tab_count.setStyleSheet(_tab_qss(False))

    tab_trait = QTabWidget()
    tab_trait.setDocumentMode(True)
    tab_trait.setStyleSheet(_tab_qss(False))

    round_data = []
    group_defs = parent_window.get_active_groups(rounds_all) if parent_window and hasattr(parent_window, "get_active_groups") else []
    round_data.append(("전체", species_all, rounds_all))
    for group_name, group_rns in group_defs:
        chosen = _species_in_rounds(species_all, group_rns)
        if chosen:
            round_data.append((group_name, chosen, group_rns))

    if not group_defs:
        for rk, rr in _round_groups(rounds_all).items():
            chosen = _species_in_rounds(species_all, rr)
            if chosen:
                round_data.append((rk, chosen, rr))

    aux_sheets = aux_sheets or {}
    nat_sheet = aux_sheets.get("귀화식물")
    disturb_sheet = aux_sheets.get("교란생물")
    num_rounds = max(len(round_data), 1)

    # 전체 탭용 회차별 ctx 목록 (extra_round_data)
    for rk, species, rounds in round_data:
        if parent_window and hasattr(parent_window, "step_sub_progress"):
            parent_window.step_sub_progress(f"화면(UI) 생성 중 ({sheet_name}: {mode_label} - {rk})", progress_share / num_rounds)

        cache = {"life_counts": {}, "unit_counts": {}, "nat_metrics": {}, "selected_name_sets": {}, "round_species": {}}
        ctx = _build_round_context(species, rounds, nat_sheet, disturb_sheet, cache)

        is_norm_sent = [False]
        panel_sent_ref = [None]

        def _on_norm(ref=panel_sent_ref):
            is_norm_sent[0] = True
            if ref[0]: ref[0]._sent_refresh_fn()
            # 식물상은 문장 화면이 별도 탭("✏  분석 문장")에 있으므로
            # 비율 보정 시 전역 문장 새로고침 시그널을 함께 올린다.
            if parent_window and hasattr(parent_window, "sig_sent_settings"):
                try:
                    parent_window.sig_sent_settings.emit()
                except Exception:
                    pass

        page_count = QWidget(); page_count.setStyleSheet(_QSS)
        lay_count = QVBoxLayout(page_count); lay_count.setContentsMargins(6, 6, 6, 6); lay_count.setSpacing(8)
        count_tbl = _count_tbl(species, cache["unit_counts"])
        lay_count.addWidget(_sc(_section_page("식물상 집계표", count_tbl)))
        tab_count.addTab(page_count, rk)

        page_trait = QWidget(); page_trait.setStyleSheet(_QSS)
        lay_trait = QVBoxLayout(page_trait); lay_trait.setContentsMargins(6, 6, 6, 6); lay_trait.setSpacing(8)
        trait_page = _combined_trait_page(ctx, nat_sheet, disturb_sheet, on_norm_cb=_on_norm, parent_window=parent_window)
        lay_trait.addWidget(_sc(trait_page))
        tab_trait.addTab(page_trait, rk)

        if sentence_table_registry is not None:
            refs = {"count": (lambda tbl=count_tbl: tbl)}
            refs.update(getattr(trait_page, "_table_refs", {}))
            sentence_table_registry[(mode_label, rk)] = refs

    main_tabs.addTab(tab_count, "식물상 집계")
    main_tabs.addTab(tab_trait, "생활형/귀화율")

    lay.addWidget(main_tabs)
    return wrap


def _combined_sent_view(fld_species, fld_rounds, lit_species, lit_rounds,
                         prot_sheet, aux_sheets, parent_window=None, sentence_table_registry=None):
    aux_sheets = aux_sheets or {}
    nat_sheet     = aux_sheets.get("귀화식물")
    disturb_sheet = aux_sheets.get("교란생물")

    def _mk_ctx(sp, rns):
        cache = {"life_counts": {}, "unit_counts": {}, "nat_metrics": {}, "selected_name_sets": {}, "round_species": {}}
        return _build_round_context(sp, rns, nat_sheet, disturb_sheet, cache)

    def _collect(sp_list, rns):
        data = []
        if sp_list:
            data.append(("전체", sp_list, rns))
            group_defs = parent_window.get_active_groups(rns) if parent_window and hasattr(parent_window, "get_active_groups") else []
            for group_name, group_rns in group_defs:
                chosen = _species_in_rounds(sp_list, group_rns)
                if chosen:
                    data.append((group_name, chosen, group_rns))
            if not group_defs:
                for rk, rr in _round_groups(rns).items():
                    chosen = _species_in_rounds(sp_list, rr)
                    if chosen:
                        data.append((rk, chosen, rr))
        return data

    fld_data = _collect(fld_species, fld_rounds)
    lit_data  = _collect(lit_species,  lit_rounds)

    def _sent_text(mode_label, data_list, rk, cat_key):
        entry = next(((sp, rns) for lbl, sp, rns in data_list if lbl == rk), None)
        if not entry:
            return ""
        num_rounds = len([x for x in data_list if x[0] != "전체"])
        tbl_ref = (sentence_table_registry or {}).get((mode_label, rk), {})
        sents = _build_plant_sentences(mode_label, _mk_ctx(*entry), prot_sheet, aux_sheets, round_label=rk, num_rounds=num_rounds, tbl_ref=tbl_ref)
        text = sents.get(cat_key, "")
        
        if rk == "전체":
            for sub_rk, sub_sp, sub_rns in data_list:
                if sub_rk == "전체":
                    continue
                sub_tbl_ref = (sentence_table_registry or {}).get((mode_label, sub_rk), {})
                sub_s = _build_plant_sentences(mode_label, _mk_ctx(sub_sp, sub_rns), prot_sheet, aux_sheets, round_label=sub_rk, num_rounds=num_rounds, tbl_ref=sub_tbl_ref)
                chunk = sub_s.get(cat_key, "")
                if chunk:
                    text = (text + "\n\n" if text else "") + chunk
        return text

    def _has_sentence(mode_label, rk, sp, rns, sent_key):
        tbl_ref = (sentence_table_registry or {}).get((mode_label, rk), {})
        sents = _build_plant_sentences(
            mode_label,
            _mk_ctx(sp, rns),
            prot_sheet,
            aux_sheets,
            round_label=rk,
            tbl_ref=tbl_ref,
        )
        actual_key = sent_key
        if sent_key == "멸종위기 야생생물" and actual_key not in sents:
            actual_key = "법정보호종"
        return actual_key in sents, actual_key

    _TE_QSS = (f"QPlainTextEdit{{background:#FFFFFF;border:1.5px solid {_BORDER};"
               f"border-radius:8px;{FF_KR};font-size:12px;padding:6px;}}")

    def _mk_plain(text: str) -> QPlainTextEdit:
        te = QPlainTextEdit(); te.setReadOnly(True)
        te.setPlainText(text or ""); te.setStyleSheet(_TE_QSS)
        return te

    outer = QWidget(); outer.setStyleSheet(_QSS)
    outer_lay = QVBoxLayout(outer); outer_lay.setContentsMargins(0, 0, 0, 0)

    rk_tabs = QTabWidget()
    rk_tabs.setDocumentMode(True)
    rk_tabs.setStyleSheet(_tab_qss(False))

    te_map: dict = {}
    _lim_edits: dict = {}

    tab_configs = []
    for rk, sp, rns in fld_data:
        tab_configs.append((rk, "현지조사", fld_data, sp, rns))
        
    lit_entry = next((x for x in lit_data if x[0] == "전체"), None)
    if lit_entry:
        tab_configs.append(("문헌조사", "문헌조사", lit_data, lit_entry[1], lit_entry[2]))

    # 고정 탭 순서: (표시명, 문장 키, 현지조사 전용 여부)
    _SENT_TAB_DEFS = [
        ("식물상",           "식물상",                  False),
        ("생활형",           "소산식물 생활형 현황",      True),
        ("양치식물",         "양치식물 분포현황",         True),
        ("귀화식물",         "귀화식물",                 False),
        ("교란생물",         "생태계교란 생물",       False),
        ("특산식물",         "산림청지정 특산식물",        False),
        ("구계학적 특정식물", "식물구계학적 특정종",       False),
        ("희귀식물",         "산림청지정 희귀식물",        False),
        ("멸종위기종",       "멸종위기 야생생물",          False),  # 없으면 "법정보호종"
    ]

    for tab_label, mode, data_list, sp, rns in tab_configs:
        cat_tabs = QTabWidget()
        cat_tabs.setDocumentMode(True)
        cat_tabs.setStyleSheet(_tab_qss(False))

        is_field_mode = (mode == "현지조사")
        for disp_name, sent_key, field_only in _SENT_TAB_DEFS:
            if field_only and not is_field_mode:
                continue
            exists, actual_key = _has_sentence(mode, tab_label if is_field_mode else "전체", sp, rns, sent_key)
            if not exists:
                continue

            rk_for_text = tab_label if is_field_mode else "전체"
            text = _sent_text(mode, data_list, rk_for_text, actual_key)
            te = _mk_plain(text)

            cat_page = QWidget(); cat_page.setStyleSheet(_QSS)
            cat_lay = QVBoxLayout(cat_page); cat_lay.setContentsMargins(6, 6, 6, 6); cat_lay.setSpacing(8)

            btn_copy = QPushButton("📋  복사"); btn_copy.setFixedHeight(30)
            btn_copy.setStyleSheet(_COPY_BTN_QSS)
            def _copy_fn(_=False, t=te, b=btn_copy):
                QApplication.clipboard().setText(t.toPlainText())
                from ui_shared import apply_button_feedback; apply_button_feedback(b)
            btn_copy.clicked.connect(_copy_fn)

            limits = getattr(SETTINGS, "plant_sent_list_limits", {}) or {}
            if not isinstance(limits, dict):
                limits = {}
                SETTINGS.plant_sent_list_limits = limits

            lbl_lim = QLabel("나열")
            cur_lim = _list_limit_for(actual_key)
            edt_lim = QLineEdit("전체" if cur_lim <= 0 else str(cur_lim))
            edt_lim.setStyleSheet(_LIM_QSS)
            edt_lim.setFixedWidth(48)
            _lim_edits[(tab_label, disp_name)] = edt_lim

            def _on_lim_changed(text, key=actual_key):
                t = str(text or "").strip()
                if t in ("전체", "all", "ALL"):
                    val = 0
                else:
                    try:
                        val = max(1, int(t))
                    except ValueError:
                        return
                limits = getattr(SETTINGS, "plant_sent_list_limits", {}) or {}
                if not isinstance(limits, dict):
                    limits = {}
                limits[key] = val
                SETTINGS.plant_sent_list_limits = limits
                _refresh()
            edt_lim.textChanged.connect(_on_lim_changed)

            btn_all = QPushButton("전체나열")
            btn_all.setFixedHeight(24)
            btn_all.setStyleSheet(_COPY_BTN_QSS)
            def _all_fn(_=False, e=edt_lim, key=actual_key):
                limits = getattr(SETTINGS, "plant_sent_list_limits", {}) or {}
                if not isinstance(limits, dict):
                    limits = {}
                limits[key] = 0
                SETTINGS.plant_sent_list_limits = limits
                e.blockSignals(True); e.setText("전체"); e.blockSignals(False)
                _refresh()
            btn_all.clicked.connect(_all_fn)

            br = QHBoxLayout()
            br.addWidget(btn_copy)
            br.addWidget(lbl_lim)
            br.addWidget(edt_lim)
            br.addWidget(btn_all)
            if disp_name == "생활형":
                cb_life_fmt = QComboBox()
                cb_life_fmt.addItem("분류군(%)", "count_pct")
                cb_life_fmt.addItem("%(분류군)", "pct_count")
                cur_fmt = getattr(SETTINGS, "plant_life_item_format", "count_pct")
                cb_life_fmt.setCurrentIndex(1 if cur_fmt == "pct_count" else 0)
                cb_life_fmt.setFixedHeight(24)
                cb_life_fmt.currentIndexChanged.connect(lambda *_args, cb=cb_life_fmt: (setattr(SETTINGS, "plant_life_item_format", cb.currentData()), _refresh()))
                br.addWidget(cb_life_fmt)
            if actual_key == "식물구계학적 특정종":
                def _make_notice_btn(label, attr):
                    checked = getattr(SETTINGS, attr, True)
                    btn = QPushButton(f"{label} 표현" if checked else f"{label} 미표현")
                    btn.setCheckable(True)
                    btn.setChecked(checked)
                    btn.setFixedHeight(24)
                    btn.setStyleSheet(_COPY_BTN_QSS)
                    def _notice_fn(_=False, b=btn, a=attr, lab=label):
                        setattr(SETTINGS, a, b.isChecked())
                        b.setText(f"{lab} 표현" if b.isChecked() else f"{lab} 미표현")
                        _refresh()
                    btn.clicked.connect(_notice_fn)
                    return btn
                br.addWidget(_make_notice_btn("Ⅴ등급", "plant_floristic_notice_grade_v"))
                br.addWidget(_make_notice_btn("Ⅳ등급", "plant_floristic_notice_grade_iv"))
            br.addStretch()

            cat_lay.addLayout(br)
            cat_lay.addWidget(te, 1)
            cat_tabs.addTab(cat_page, disp_name)

            te_map[(tab_label, disp_name)] = (te, mode, data_list, rk_for_text, actual_key)

        rk_tabs.addTab(cat_tabs, tab_label)

    def _refresh():
        for (tab_label, disp_name), (te, mode, data_list, rk, actual_key) in te_map.items():
            te.setPlainText(_sent_text(mode, data_list, rk, actual_key))

    outer._sent_refresh_fn = _refresh

    outer_lay.addWidget(rk_tabs, 1)
    return outer


def _veg_clean_text(txt: str) -> str:
    txt = str(txt or "").replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    txt = txt.replace("；", ";").replace("ㆍ", "·").strip()
    while "  " in txt:
        txt = txt.replace("  ", " ")
    return txt


def _veg_norm_key(txt: str) -> str:
    txt = _veg_clean_text(txt)
    for a, b in ((" ", ""), ("　", ""), ("?", "-"), ("－", "-"), ("＜", "<"), ("＞", ">"), ("（", "("), ("）", ")")):
        txt = txt.replace(a, b)
    return txt


def _veg_join_korean(items) -> str:
    vals = [str(x) for x in items if str(x).strip()]
    if len(vals) <= 1:
        return vals[0] if vals else ""
    if len(vals) == 2:
        return " 및 ".join(vals)
    return ", ".join(vals[:-1]) + f" 및 {vals[-1]}"


def _veg_flora_feature(name: str) -> str:
    return _VEG_FLORA_FEATURES.get(_veg_norm_key(name), "")


def _veg_climate_species(name: str) -> str:
    return _VEG_CLIMATE_SPECIES.get(_veg_norm_key(name), "")


def _veg_format_flora_feature(feature: str, idx: int, total: int) -> str:
    feature = _veg_clean_text(feature)
    if feature.endswith(","):
        feature = feature[:-1].strip()
    if total == 1:
        return feature if feature.endswith(",") else feature + ","
    if idx == 0:
        if feature.endswith("으로"):
            feature = feature[:-2]
        return feature + "이며,"
    if idx < total - 1:
        if feature.endswith("으로"):
            feature = feature[:-2]
        return feature + ","
    return feature if feature.endswith(",") else feature + ","


def _veg_flora_feature_sentence(flora_items) -> str:
    parts = []
    total = len(flora_items)
    for idx, item in enumerate(flora_items):
        feature = _veg_flora_feature(item)
        if feature:
            parts.append(f"{item}은 {_veg_format_flora_feature(feature, idx, total)}")
    return " ".join(parts).strip()


def _veg_species_list(species_text: str) -> list[str]:
    out = []
    seen = set()
    for raw in str(species_text or "").split(","):
        item = _veg_clean_text(raw)
        if item and item not in seen:
            out.append(item)
            seen.add(item)
    return out


def _veg_common_species_text(climate_items) -> str:
    if not climate_items:
        return ""
    first = _veg_species_list(_veg_climate_species(climate_items[0]))
    others = [set(_veg_species_list(_veg_climate_species(item))) for item in climate_items[1:]]
    common = [sp for sp in first if all(sp in s for s in others)]
    return ", ".join(common)


def _veg_climate_fallback_sentence(climate_items) -> str:
    parts = []
    for item in climate_items:
        species = _veg_climate_species(item)
        if species:
            parts.append(f"{item}은 {species} 등")
    return "각 식생기후별 주요 수종은 " + "; ".join(parts) + "등이 분포함"


def _veg_climate_species_sentence(climate_items) -> str:
    if len(climate_items) == 1:
        item = climate_items[0]
        return f"{item}의 주요 교목 수종은 {_veg_climate_species(item)} 등이 분포함"
    common = _veg_common_species_text(climate_items)
    if common:
        return f"주요 교목 수종은 {common} 등이 분포함"
    return _veg_climate_fallback_sentence(climate_items)


def _build_vegetation_sentence(location: str, flora_items, climate_items) -> str:
    location = _veg_clean_text(location)
    sentence_1 = f"사업지구는 {location}에 위치하며, {_veg_join_korean(flora_items)}, {_veg_join_korean(climate_items)}에 해당됨"
    sentence_2 = f"{_veg_flora_feature_sentence(flora_items)} {_veg_climate_species_sentence(climate_items)}".strip()
    return sentence_1 + "\n\n" + sentence_2


_VEG_GRADE_ORDER = ["Ⅰ", "Ⅱ", "Ⅲ", "Ⅳ", "Ⅴ"]
_VEG_GRADE_MAP = {
    "1": "Ⅰ", "2": "Ⅱ", "3": "Ⅲ", "4": "Ⅳ", "5": "Ⅴ",
    "I": "Ⅰ", "II": "Ⅱ", "III": "Ⅲ", "IV": "Ⅳ", "V": "Ⅴ",
}


def _veg_grade(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    if s.endswith("등급"):
        s = s[:-2].strip()
    return _VEG_GRADE_MAP.get(s.upper(), s)


def _veg_num(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0


def _veg_fmt_area(v) -> str:
    n = _veg_num(v)
    return "-" if n <= 0 else f"{int(round(n)):,}"


def _veg_fmt_pct(v) -> str:
    n = _veg_num(v)
    decimals = max(0, int(getattr(SETTINGS, "decimal", 1)))
    return "-" if n <= 0 else f"{n:.{decimals}f}"


def _veg_sum_pct(rows: list[dict]) -> str:
    decimals = max(0, int(getattr(SETTINGS, "decimal", 1)))
    s = sum(_veg_num(r.get("pct")) for r in rows)
    return f"{s:.{decimals}f}"


def _veg_pct_value(v) -> float:
    n = _veg_num(v)
    return n * 100.0 if 0 < n <= 1.0001 else n


def _veg_join_comma(items) -> str:
    return ", ".join([str(x) for x in items if str(x).strip()])



def _veg_top_type_items(rows, n=4):
    grouped = OrderedDict()
    for r in rows:
        name = str(r.get("name") or "").strip()
        if not name or name == "합계" or _veg_num(r.get("area")) <= 0:
            continue
        item = grouped.setdefault(name, {"name": name, "area": 0.0, "pct": 0.0})
        item["area"] += _veg_num(r.get("area"))
        item["pct"] += _veg_num(r.get("pct"))
    return sorted(grouped.values(), key=lambda r: (-_veg_num(r.get("area")), r.get("name", "")))[:n]


def _veg_top_distinct_row_items(rows, n=3, exclude_names=None):
    exclude_names = set(exclude_names or [])
    picked = []
    seen = set()
    for r in sorted([r for r in rows if _veg_num(r.get("area")) > 0], key=lambda x: (-_veg_num(x.get("area")), x.get("name", ""))):
        name = str(r.get("name") or "").strip()
        if not name or name in exclude_names or name in seen:
            continue
        picked.append(r)
        seen.add(name)
        if len(picked) >= n:
            break
    return picked


def _veg_grade_values(grade) -> list[str]:
    vals = []
    for part in re.split(r"[,/·\s]+", str(grade or "")):
        g = _veg_grade(part)
        if g and g not in vals:
            vals.append(g)
    return vals




def _veg_has_batchim(text: str) -> bool:
    for ch in reversed(str(text or "").strip()):
        if "가" <= ch <= "힣":
            return (ord(ch) - 0xAC00) % 28 != 0
    return False


def _veg_iga(text: str) -> str:
    return "이" if _veg_has_batchim(text) else "가"


def _veg_adapt_end(s: str) -> str:
    """SETTINGS.end_field 어미 스타일에 맞춰 종결어미 변환."""
    import re
    base = str(getattr(SETTINGS, "end_field", "") or "").strip()
    if base.endswith("됨") or base.endswith("함"):
        s = re.sub(r'하였음$', '함', s)
        s = re.sub(r'되었음$', '됨', s)
        s = re.sub(r'였음$', '였음', s)  # 보였음 등은 그대로
    elif base.endswith("되었다") or base.endswith("하였다"):
        s = re.sub(r'하였음$', '하였다', s)
        s = re.sub(r'되었음$', '되었다', s)
        s = re.sub(r'였음$', '였다', s)
    return s


def _veg_classified_end() -> str:
    return _veg_adapt_end("구분되었음")


def _merge_vegetation_rows(rows: list[dict]) -> list[dict]:
    merged = OrderedDict()
    for row in rows:
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        grade_key = tuple(_veg_grade_values(row.get("grade")))
        key = (name, grade_key)
        item = merged.setdefault(key, {"name": name, "area": 0.0, "pct": 0.0, "grade": ""})
        item["area"] += _veg_num(row.get("area"))
        item["pct"] += _veg_num(row.get("pct"))
        item["grade"] = ", ".join(grade_key)
    return list(merged.values())


def _worksheet_values_with_merged(ws) -> list[list]:
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            if r - 1 >= len(rows):
                continue
            row = rows[r - 1]
            for c in range(min_col, max_col + 1):
                if c - 1 < len(row):
                    row[c - 1] = val
    return rows


_VEG_HEADER_SKIP = frozenset({
    "식생유형", "면적", "구성비", "유형", "비율", "등급",
    "구분", "구 분", "합계", "합", "조사지역", "구간", "현황",
})


def _extract_route_label(rows, col_start: int = 8, col_end: int = 12) -> str:
    for row in rows[:4]:
        for ci in range(col_start, min(col_end + 1, len(row))):
            val = _veg_clean_text(row[ci] if ci < len(row) else "")
            if val and val not in _VEG_HEADER_SKIP:
                return val
    return ""


def _new_vegetation_area_block(label: str, route_label: str = "") -> dict:
    return {
        "label": label or "전체",
        "route_label": route_label,
        "study": {"rows": [], "grades": [], "total_area": 0.0, "total_pct": 100.0},
        "route": {"rows": [], "grades": [], "total_area": 0.0, "total_pct": 100.0},
    }


def _finalize_vegetation_area_block(data: dict) -> dict:
    for key in ("study", "route"):
        data[key]["rows"] = _merge_vegetation_rows(data[key]["rows"])
        if not data[key]["total_area"]:
            data[key]["total_area"] = sum(_veg_num(r.get("area")) for r in data[key]["rows"])
        seen = set()
        dedup = []
        for g in data[key]["grades"]:
            if g["grade"] in seen or _veg_num(g.get("area")) <= 0:
                continue
            seen.add(g["grade"])
            dedup.append(g)
        data[key]["grades"] = dedup
    return data


def _find_merged_label_col(rows, start: int, stop: int) -> int:
    """지정 범위 열(start~stop) 중 병합 셀 패턴(연속 동일 텍스트)인 첫 열 반환.
    숫자(면적·구성비 등급 병합값)는 라벨로 취급하지 않음."""
    for ci in range(start, stop):
        prev = None
        for row in rows:
            v = _veg_clean_text(row[ci] if ci < len(row) else "")
            if v and v not in _VEG_HEADER_SKIP:
                try:
                    float(v.replace(",", "").replace("%", ""))
                    prev = None  # 숫자는 건너뜀
                    continue
                except ValueError:
                    pass
                if v == prev:
                    return ci
                prev = v
            else:
                prev = None
    return start


def _find_block_label_col(rows) -> int:
    return _find_merged_label_col(rows, 0, 6)


def _find_route_label_col(rows, bc: int) -> int:
    max_col = max((len(row) for row in rows if row), default=bc + 15)
    return _find_merged_label_col(rows, bc + 5, max_col)


def _count_block_labels(rows, bc: int) -> dict:
    """bc 열의 텍스트 등장 횟수를 집계. 합/합계 제외."""
    cnt = {}
    for row in rows:
        c = _veg_clean_text(row[bc] if bc < len(row) else "")
        if c and c not in _VEG_HEADER_SKIP and c not in ("합", "합계"):
            cnt[c] = cnt.get(c, 0) + 1
    return cnt


def _is_segmented_vegetation_rows(rows) -> bool:
    bc = _find_block_label_col(rows)
    # 병합 셀 라벨은 2행 이상 반복됨 — 1회만 등장하는 타이틀은 제외
    block_labels = {k for k, v in _count_block_labels(rows, bc).items() if v >= 2}
    return len(block_labels) >= 2


def _parse_segmented_vegetation_area_rows(rows) -> list[dict]:
    bc = _find_block_label_col(rows)
    rc = _find_route_label_col(rows, bc)
    # 1회만 등장하는 값은 타이틀/헤더 — 진짜 블럭 라벨만 허용
    valid_labels = {k for k, v in _count_block_labels(rows, bc).items() if v >= 2}
    blocks = []
    current = None

    def _close_current():
        nonlocal current
        if current is not None:
            blocks.append(_finalize_vegetation_area_block(current))
            current = None

    for row in rows:
        cb = _veg_clean_text(row[bc] if bc < len(row) else "")

        if cb in ("합", "합계"):
            if current is not None:
                current["study"]["total_area"] = _veg_num(row[bc+2] if len(row) > bc+2 else 0)
                current["route"]["total_area"] = _veg_num(row[rc+2] if len(row) > rc+2 else 0)
                _close_current()
            continue
        elif cb and cb in valid_labels:
            if current is None or current.get("label") != cb:
                _close_current()
                rl = _veg_clean_text(row[rc] if rc < len(row) else "")
                if rl in _VEG_HEADER_SKIP:
                    rl = ""
                current = _new_vegetation_area_block(cb, rl)

        if current is None:
            continue

        s_name = _veg_clean_text(row[bc+1] if len(row) > bc+1 else "")
        if s_name and s_name not in _VEG_HEADER_SKIP and s_name != cb:
            current["study"]["rows"].append({
                "name": s_name,
                "area": _veg_num(row[bc+2] if len(row) > bc+2 else 0),
                "pct": _veg_pct_value(row[bc+3] if len(row) > bc+3 else 0),
                "grade": _veg_grade(row[bc+4] if len(row) > bc+4 else ""),
            })
        r_name = _veg_clean_text(row[rc+1] if len(row) > rc+1 else "")
        if r_name and r_name not in _VEG_HEADER_SKIP and r_name != cb:
            current["route"]["rows"].append({
                "name": r_name,
                "area": _veg_num(row[rc+2] if len(row) > rc+2 else 0),
                "pct": _veg_pct_value(row[rc+3] if len(row) > rc+3 else 0),
                "grade": _veg_grade(row[rc+4] if len(row) > rc+4 else ""),
            })

        s_grade = _veg_grade(row[bc+4] if len(row) > bc+4 else "")
        if s_grade in _VEG_GRADE_ORDER:
            current["study"]["grades"].append({
                "grade": s_grade,
                "area": _veg_num(row[bc+5] if len(row) > bc+5 else 0),
                "pct": _veg_pct_value(row[bc+6] if len(row) > bc+6 else 0),
            })
        r_grade = _veg_grade(row[rc+4] if len(row) > rc+4 else "")
        if r_grade in _VEG_GRADE_ORDER:
            current["route"]["grades"].append({
                "grade": r_grade,
                "area": _veg_num(row[rc+5] if len(row) > rc+5 else 0),
                "pct": _veg_pct_value(row[rc+6] if len(row) > rc+6 else 0),
            })

    _close_current()
    return blocks


def _parse_vegetation_area_xlsx(path: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[-1]]
    rows = _worksheet_values_with_merged(ws)
    route_label = _extract_route_label(rows)
    if _is_segmented_vegetation_rows(rows):
        return _parse_segmented_vegetation_area_rows(rows)

    data = _new_vegetation_area_block("전체", route_label)
    study_done = False
    route_done = False

    for row in rows:
        s_name = _veg_clean_text(row[2] if len(row) > 2 else "")
        r_name = _veg_clean_text(row[9] if len(row) > 9 else "")
        if s_name and s_name != "식생유형" and not study_done:
            if s_name == "합계":
                data["study"]["total_area"] = _veg_num(row[3] if len(row) > 3 else 0)
                study_done = True
            else:
                data["study"]["rows"].append({
                    "name": s_name,
                    "area": _veg_num(row[3] if len(row) > 3 else 0),
                    "pct": _veg_num(row[4] if len(row) > 4 else 0),
                    "grade": _veg_grade(row[5] if len(row) > 5 else ""),
                })
        if r_name and r_name != "식생유형" and not route_done:
            if r_name == "합계":
                data["route"]["total_area"] = _veg_num(row[10] if len(row) > 10 else 0)
                route_done = True
            else:
                data["route"]["rows"].append({
                    "name": r_name,
                    "area": _veg_num(row[10] if len(row) > 10 else 0),
                    "pct": _veg_num(row[11] if len(row) > 11 else 0),
                    "grade": _veg_grade(row[12] if len(row) > 12 else ""),
                })

        s_grade = _veg_grade(row[5] if len(row) > 5 else "")
        if s_grade in _VEG_GRADE_ORDER:
            data["study"]["grades"].append({
                "grade": s_grade,
                "area": _veg_num(row[6] if len(row) > 6 else 0),
                "pct": _veg_num(row[7] if len(row) > 7 else 0),
            })
        r_grade = _veg_grade(row[12] if len(row) > 12 else "")
        if r_grade in _VEG_GRADE_ORDER:
            data["route"]["grades"].append({
                "grade": r_grade,
                "area": _veg_num(row[13] if len(row) > 13 else 0),
                "pct": _veg_num(row[14] if len(row) > 14 else 0),
            })

    return [_finalize_vegetation_area_block(data)]


def _veg_type_distribution_sentence(label: str, rows: list[dict]) -> str:
    present = [r for r in rows if _veg_num(r.get("area")) > 0 and r.get("name") != "합계"]
    if not present:
        return f"{label}에 분포하는 식생 유형을 확인할 수 없음"
    top = _veg_top_type_items(present, 3)
    parts = ", ".join(f"{r['name']}({_veg_fmt_pct(r.get('pct'))}%)" for r in top)
    suffix = f", {parts} 등의 순으로 " + _veg_adapt_end("분포하였음") if top else ""
    return f"{label}에 분포하는 식생 유형은 {len(present)}개로 구분되었고{suffix}"




def _veg_grade_distribution_sentence(label: str, grades: list[dict]) -> str:
    present = [g for g in grades if _veg_num(g.get("area")) > 0]
    if not present:
        return f"{label}의 식생보전등급 분포를 확인할 수 없음"
    sorted_g = sorted(present, key=lambda g: _veg_num(g.get("pct")), reverse=True)
    n = len(sorted_g)
    parts = [f"{g['grade']}등급({_veg_fmt_pct(g.get('pct'))}%)" for g in sorted_g]
    intro = f"{label}에 분포하는 식생보전등급은 총 {n}개 유형으로 " + _veg_adapt_end("구분되었으며")
    return intro + ", " + ", ".join(parts) + " 순으로 " + _veg_adapt_end("확인되었음")


def _build_veg_type_sentences(data: dict) -> str:
    rl = data.get("route_label", "")
    lines = [_veg_type_distribution_sentence("조사지역", data["study"]["rows"])]
    if rl:
        lines.append(_veg_type_distribution_sentence(rl, data["route"]["rows"]))
    return "\n".join(lines)


def _build_veg_grade_sentences(data: dict) -> str:
    rl = data.get("route_label", "")
    lines = [_veg_grade_distribution_sentence("조사지역", data["study"]["grades"])]
    if rl:
        lines.append(_veg_grade_distribution_sentence(rl, data["route"]["grades"]))
    return "\n".join(lines)


def _set_table_item(tbl, row, col, text, align=Qt.AlignCenter):
    tbl.setItem(row, col, _item(text, align))


def _fit_vegetation_result_table(tbl: QTableWidget):
    tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    tbl.setWordWrap(False)
    header = tbl.horizontalHeader()
    header.setStretchLastSection(False)
    for col in range(tbl.columnCount()):
        header.setSectionResizeMode(col, QHeaderView.Stretch)
    _fit_table_height(tbl, 8)


def _veg_grade_sort_key(r) -> int:
    g = _veg_grade(str(r.get("grade") or "").split(",")[0].strip())
    try:
        return _VEG_GRADE_ORDER.index(g)
    except ValueError:
        return 99


def _apply_grade_span(tbl: QTableWidget, grade_col: int, data_rows_count: int):
    """grade_col 열에서 같은 등급값이 연속된 행을 setSpan으로 병합."""
    i = 0
    while i < data_rows_count:
        val = tbl.item(i, grade_col)
        cur = val.text() if val else ""
        j = i + 1
        while j < data_rows_count:
            nxt = tbl.item(j, grade_col)
            if (nxt.text() if nxt else "") != cur:
                break
            j += 1
        span = j - i
        if span > 1:
            tbl.setSpan(i, grade_col, span, 1)
        i = j


def _fill_vegetation_area_table(tbl: QTableWidget, data: dict):
    rl = data.get("route_label", "")
    rows = sorted(data["study"]["rows"], key=_veg_grade_sort_key)
    tbl.setRowCount(0)
    if rl:
        tbl.setColumnCount(6)
        tbl.setHorizontalHeaderLabels(["등급", "구 분", f"{rl} 면적(㎡)", f"{rl} 구성비(%)", "조사지역 면적(㎡)", "조사지역 구성비(%)"])
        route_by_key = {(r["name"], tuple(_veg_grade_values(r.get("grade")))): r for r in data["route"]["rows"]}
        for r in rows:
            ri = tbl.rowCount()
            tbl.insertRow(ri)
            rr = route_by_key.get((r["name"], tuple(_veg_grade_values(r.get("grade")))), {})
            g = _veg_grade(str(r.get("grade") or "").split(",")[0].strip())
            _set_table_item(tbl, ri, 0, f"{g}등급" if g else "")
            _set_table_item(tbl, ri, 1, r["name"])
            _set_table_item(tbl, ri, 2, _veg_fmt_area(rr.get("area")))
            _set_table_item(tbl, ri, 3, _veg_fmt_pct(rr.get("pct")))
            _set_table_item(tbl, ri, 4, _veg_fmt_area(r.get("area")))
            _set_table_item(tbl, ri, 5, _veg_fmt_pct(r.get("pct")))
        _apply_grade_span(tbl, 0, tbl.rowCount())
        ri = tbl.rowCount()
        tbl.insertRow(ri)
        _set_table_item(tbl, ri, 0, "")
        _set_table_item(tbl, ri, 1, "합계")
        _set_table_item(tbl, ri, 2, _veg_fmt_area(data["route"]["total_area"]))
        _set_table_item(tbl, ri, 3, _veg_sum_pct(data["route"]["rows"]))
        _set_table_item(tbl, ri, 4, _veg_fmt_area(data["study"]["total_area"]))
        _set_table_item(tbl, ri, 5, _veg_sum_pct(data["study"]["rows"]))
    else:
        tbl.setColumnCount(4)
        tbl.setHorizontalHeaderLabels(["등급", "구 분", "조사지역 면적(㎡)", "조사지역 구성비(%)"])
        for r in rows:
            ri = tbl.rowCount()
            tbl.insertRow(ri)
            g = _veg_grade(str(r.get("grade") or "").split(",")[0].strip())
            _set_table_item(tbl, ri, 0, f"{g}등급" if g else "")
            _set_table_item(tbl, ri, 1, r["name"])
            _set_table_item(tbl, ri, 2, _veg_fmt_area(r.get("area")))
            _set_table_item(tbl, ri, 3, _veg_fmt_pct(r.get("pct")))
        _apply_grade_span(tbl, 0, tbl.rowCount())
        ri = tbl.rowCount()
        tbl.insertRow(ri)
        _set_table_item(tbl, ri, 0, "")
        _set_table_item(tbl, ri, 1, "합계")
        _set_table_item(tbl, ri, 2, _veg_fmt_area(data["study"]["total_area"]))
        _set_table_item(tbl, ri, 3, _veg_sum_pct(data["study"]["rows"]))
    _bold_row(tbl, ri)
    _fit_vegetation_result_table(tbl)


def _fill_vegetation_grade_table(tbl: QTableWidget, data: dict):
    rl = data.get("route_label", "")
    tbl.setRowCount(0)
    study = {g["grade"]: g for g in data["study"]["grades"]}
    route = {g["grade"]: g for g in data["route"]["grades"]}
    grades = [g for g in _VEG_GRADE_ORDER if g in study or g in route]
    if rl:
        tbl.setHorizontalHeaderLabels(["구 분", f"{rl} 면적(㎡)", f"{rl} 구성비(%)", "조사지역 면적(㎡)", "조사지역 구성비(%)"])
        for grade in grades:
            ri = tbl.rowCount()
            tbl.insertRow(ri)
            sg = study.get(grade, {})
            rg = route.get(grade, {})
            _set_table_item(tbl, ri, 0, f"{grade}등급")
            _set_table_item(tbl, ri, 1, _veg_fmt_area(rg.get("area")))
            _set_table_item(tbl, ri, 2, _veg_fmt_pct(rg.get("pct")))
            _set_table_item(tbl, ri, 3, _veg_fmt_area(sg.get("area")))
            _set_table_item(tbl, ri, 4, _veg_fmt_pct(sg.get("pct")))
        ri = tbl.rowCount()
        tbl.insertRow(ri)
        _set_table_item(tbl, ri, 0, "합계")
        _set_table_item(tbl, ri, 1, _veg_fmt_area(data["route"]["total_area"]))
        _set_table_item(tbl, ri, 2, _veg_sum_pct(data["route"]["grades"]))
        _set_table_item(tbl, ri, 3, _veg_fmt_area(data["study"]["total_area"]))
        _set_table_item(tbl, ri, 4, _veg_sum_pct(data["study"]["grades"]))
    else:
        tbl.setHorizontalHeaderLabels(["구 분", "조사지역 면적(㎡)", "조사지역 구성비(%)"])
        for grade in grades:
            ri = tbl.rowCount()
            tbl.insertRow(ri)
            sg = study.get(grade, {})
            _set_table_item(tbl, ri, 0, f"{grade}등급")
            _set_table_item(tbl, ri, 1, _veg_fmt_area(sg.get("area")))
            _set_table_item(tbl, ri, 2, _veg_fmt_pct(sg.get("pct")))
        ri = tbl.rowCount()
        tbl.insertRow(ri)
        _set_table_item(tbl, ri, 0, "합계")
        _set_table_item(tbl, ri, 1, _veg_fmt_area(data["study"]["total_area"]))
        _set_table_item(tbl, ri, 2, _veg_sum_pct(data["study"]["grades"]))
    _bold_row(tbl, ri)
    _fit_vegetation_result_table(tbl)


def _make_vegetation_tab(parent_window=None):
    page = QWidget()
    page.setStyleSheet(_QSS)
    root = QVBoxLayout(page)
    root.setContentsMargins(8, 8, 8, 8)
    root.setSpacing(8)

    tabs = QTabWidget()
    tabs.setDocumentMode(True)
    tabs.setStyleSheet(_tab_qss(False))
    root.addWidget(tabs, 1)

    geo_page = QWidget()
    geo_page.setStyleSheet(_QSS)
    geo_lay = QVBoxLayout(geo_page)
    geo_lay.setContentsMargins(8, 8, 8, 8)
    geo_lay.setSpacing(8)

    area_page = QWidget()
    area_page.setStyleSheet(_VEG_WHITE_QSS)
    area_root = QVBoxLayout(area_page)
    area_root.setContentsMargins(8, 8, 8, 8)
    area_root.setSpacing(8)

    grp_input = QGroupBox("식생 문장 생성")
    lay = QVBoxLayout(grp_input)
    lay.setContentsMargins(12, 14, 12, 12)
    lay.setSpacing(10)

    loc_row = QHBoxLayout()
    loc_lbl = QLabel("조사지역")
    loc_lbl.setFixedWidth(72)
    edit_location = QLineEdit()
    edit_location.setPlaceholderText("예: ○○시 ○○구 ○○동 일원")
    edit_location.setFixedHeight(30)
    edit_location.setStyleSheet(f"QLineEdit{{background:#FFFFFF;{BD};border-radius:6px;{FF_KR};font-size:12px;padding:4px 8px;}}")
    loc_row.addWidget(loc_lbl)
    loc_row.addWidget(edit_location, 1)
    lay.addLayout(loc_row)

    toggle_qss = (
        f"QPushButton{{background:#FFFFFF;color:{_TXT};border:1px solid {_BORDER};border-radius:7px;"
        f"padding:7px 10px;text-align:left;{FF_KR};font-size:12px;}}"
        f"QPushButton:hover{{background:{_ACCENT_L};border-color:{_ACCENT};}}"
        f"QPushButton:checked{{background:{_ACCENT};color:#FFFFFF;border-color:{_ACCENT};font-weight:700;}}"
    )

    def _make_toggle_group(title, options, columns):
        grp = QGroupBox(title)
        gl = QGridLayout(grp)
        gl.setContentsMargins(10, 14, 10, 10)
        gl.setSpacing(8)
        buttons = []
        for idx, text in enumerate(options):
            btn = QPushButton(text)
            btn.setCheckable(True)
            btn.setMinimumHeight(34)
            btn.setStyleSheet(toggle_qss)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            gl.addWidget(btn, idx // columns, idx % columns)
            buttons.append(btn)
        return grp, buttons

    flora_grp, flora_buttons = _make_toggle_group("식물상 지역", _VEG_FLORA_OPTIONS, 2)
    QApplication.processEvents()
    climate_grp, climate_buttons = _make_toggle_group("식생기후", _VEG_CLIMATE_OPTIONS, 1)
    lay.addWidget(flora_grp)
    lay.addWidget(climate_grp)
    QApplication.processEvents()

    btn_row = QHBoxLayout()
    btn_generate = QPushButton("문장 생성")
    btn_reset = QPushButton("초기화")
    btn_copy = QPushButton("📋 문장 복사")
    for btn in (btn_generate, btn_reset, btn_copy):
        btn.setFixedHeight(32)
    btn_generate.setStyleSheet(f"QPushButton{{background:{_ACCENT};color:#FFFFFF;border:none;border-radius:7px;font-weight:700;{FF_KR};}}QPushButton:hover{{background:#1D4ED8;}}")
    btn_reset.setStyleSheet(make_outline_btn_qss(_SUB, "#F3F4F6"))
    btn_copy.setStyleSheet(_COPY_BTN_QSS)
    btn_row.addStretch()
    btn_row.addWidget(btn_generate)
    btn_row.addWidget(btn_reset)
    btn_row.addWidget(btn_copy)
    lay.addLayout(btn_row)

    result = QPlainTextEdit()
    result.setReadOnly(True)
    result.setPlaceholderText("문장 생성 결과")
    result.setMinimumHeight(150)
    result.setStyleSheet(f"QPlainTextEdit{{background:#FFFFFF;{BD};border-radius:8px;{FF_KR};font-size:12px;padding:8px;}}")
    lay.addWidget(result)

    geo_lay.addWidget(grp_input)
    geo_lay.addStretch()

    grp_area = QGroupBox("현존식생 면적")
    grp_area.setStyleSheet(_VEG_WHITE_QSS)
    area_lay = QVBoxLayout(grp_area)
    area_lay.setContentsMargins(12, 14, 12, 12)
    area_lay.setSpacing(8)

    area_btn_row = QHBoxLayout()
    btn_load_area = QPushButton("면적 엑셀 불러오기")
    btn_load_area.setFixedHeight(32)
    btn_load_area.setStyleSheet(f"QPushButton{{background:{_ACCENT};color:#FFFFFF;border:none;border-radius:7px;font-weight:700;{FF_KR};}}QPushButton:hover{{background:#1D4ED8;}}")
    area_btn_row.addWidget(btn_load_area)
    area_btn_row.addStretch()
    area_lay.addLayout(area_btn_row)

    area_result_wrap = QWidget()
    area_result_wrap.setStyleSheet(_VEG_WHITE_QSS)
    area_result_lay = QVBoxLayout(area_result_wrap)
    area_result_lay.setContentsMargins(0, 0, 0, 0)
    area_result_lay.setSpacing(8)
    area_hint = QLabel("면적 엑셀을 불러오면 현존식생/식생보전등급 문장과 표가 생성됩니다.")
    area_hint.setAlignment(Qt.AlignCenter)
    area_hint.setStyleSheet(f"color:{_SUB};background:#FFFFFF;{BD};border-radius:8px;padding:18px;")
    area_result_lay.addWidget(area_hint)
    area_lay.addWidget(area_result_wrap)
    current_area_blocks = []

    area_root.addWidget(grp_area)
    area_root.addStretch()
    QApplication.processEvents()

    tabs.addTab(_sc(geo_page), "지리·기후")
    tabs.addTab(_sc(area_page), "면적")
    QApplication.processEvents()

    def _selected(buttons):
        return [btn.text() for btn in buttons if btn.isChecked()]

    def _generate():
        location = _veg_clean_text(edit_location.text())
        flora = _selected(flora_buttons)
        climate = _selected(climate_buttons)
        if not location or not flora or not climate:
            QMessageBox.warning(page, "식생 문장 생성", "조사지역, 식물상 지역, 식생기후를 모두 입력 또는 선택해 주세요.")
            return
        result.setPlainText(_build_vegetation_sentence(location, flora, climate))

    def _reset():
        edit_location.clear()
        for btn in flora_buttons + climate_buttons:
            btn.setChecked(False)
        result.clear()

    def _copy():
        text = result.toPlainText().strip()
        if not text:
            QMessageBox.warning(page, "문장 복사", "복사할 문장이 없습니다.")
            return
        QApplication.clipboard().setText(text)

    def _clear_area_results():
        while area_result_lay.count():
            item = area_result_lay.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _make_area_block_widget(data: dict):
        block_page = QWidget()
        block_page.setStyleSheet(_VEG_WHITE_QSS)
        block_lay = QVBoxLayout(block_page)
        block_lay.setContentsMargins(12, 10, 12, 12)
        block_lay.setSpacing(10)

        def _add_area_table_section(title: str, tbl: QTableWidget):
            section = QWidget()
            section.setStyleSheet(_VEG_WHITE_QSS)
            section_lay = QVBoxLayout(section)
            section_lay.setContentsMargins(0, 0, 0, 0)
            section_lay.setSpacing(3)
            title_row = QHBoxLayout()
            title_row.setContentsMargins(0, 0, 0, 0)
            title_lbl = QLabel(title)
            title_lbl.setStyleSheet(f"font-weight:700;color:{_TXT};")
            btn_norm_tbl = QPushButton("∑ 비율 보정")
            btn_norm_tbl.setFixedHeight(24)
            btn_norm_tbl.setStyleSheet(_OUTLINE_BTN_GREEN)
            btn_copy_tbl = QPushButton("📋 표 복사")
            btn_copy_tbl.setFixedHeight(24)
            btn_copy_tbl.setStyleSheet(_COPY_BTN_QSS)
            title_row.addWidget(title_lbl)
            title_row.addStretch()
            title_row.addWidget(btn_norm_tbl)
            title_row.addWidget(btn_copy_tbl)
            section_lay.addLayout(title_row)
            section_lay.addWidget(tbl)

            def _normalize_tbl():
                from ui_shared import normalize_tbl_pct_col, apply_button_feedback
                d = max(0, int(getattr(SETTINGS, 'decimal', 1)))
                fmt = lambda v: "-" if v == 0 else f"{v:.{d}f}"
                nc = tbl.columnCount()
                if nc == 6:  # 등급+구분+rl면적+rl구성비+조사지역면적+조사지역구성비
                    normalize_tbl_pct_col(tbl, 3, cnt_col=2, pct_formatter=fmt)
                    normalize_tbl_pct_col(tbl, 5, cnt_col=4, pct_formatter=fmt)
                elif nc == 4:  # 등급+구분+조사지역면적+조사지역구성비
                    normalize_tbl_pct_col(tbl, 3, cnt_col=2, pct_formatter=fmt)
                else:  # 5: 구분+rl면적+rl구성비+조사지역면적+조사지역구성비
                    normalize_tbl_pct_col(tbl, 2, cnt_col=1, pct_formatter=fmt)
                    normalize_tbl_pct_col(tbl, 4, cnt_col=3, pct_formatter=fmt)
                _fit_vegetation_result_table(tbl)
                apply_button_feedback(btn_norm_tbl)

            def _copy_tbl():
                tbl.copy_selection(include_header=True)
                from ui_shared import apply_button_feedback
                apply_button_feedback(btn_copy_tbl)

            btn_norm_tbl.clicked.connect(_normalize_tbl)
            btn_copy_tbl.clicked.connect(_copy_tbl)
            block_lay.addWidget(section)

        def _make_sent_widget(text: str) -> QWidget:
            w = QWidget()
            w.setStyleSheet(_VEG_WHITE_QSS)
            wl = QVBoxLayout(w)
            wl.setContentsMargins(0, 0, 0, 0)
            wl.setSpacing(4)
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            btn = QPushButton("📋 문장 복사")
            btn.setFixedHeight(24)
            btn.setStyleSheet(_COPY_BTN_QSS)
            row.addStretch()
            row.addWidget(btn)
            wl.addLayout(row)
            te = QPlainTextEdit()
            te.setReadOnly(True)
            te.setLineWrapMode(QPlainTextEdit.WidgetWidth)
            te.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            te.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            te.setPlainText(text)
            te.setStyleSheet(f"QPlainTextEdit{{background:#FFFFFF;{BD};border-radius:8px;{FF_KR};font-size:12px;padding:8px;}}")
            def _sync_height():
                doc_h = te.document().size().height()
                te.setFixedHeight(int(doc_h) + 20)
            te.document().documentLayout().documentSizeChanged.connect(lambda _: _sync_height())
            wl.addWidget(te)
            def _copy():
                t = te.toPlainText().strip()
                if t:
                    QApplication.clipboard().setText(t)
            btn.clicked.connect(_copy)
            return w

        rl = data.get("route_label", "")

        # ── 식생유형별 면적 분포 현황 ──
        block_lay.addWidget(_make_sent_widget(_build_veg_type_sentences(data)))
        area_tbl = _make_tbl(["등급", "구 분", "면적(㎡)", "구성비(%)"])
        _fill_vegetation_area_table(area_tbl, data)
        _add_area_table_section("조사지역의 식생유형별 면적 분포 현황", area_tbl)

        note = QLabel("주) 면적은 지형도(1:5,000)를 QGIS상에서 구적하였으며, 정확한 실측 면적과는 다소 상이할 수 있음")
        note.setWordWrap(True)
        note.setStyleSheet(f"color:{_SUB};font-size:11px;")
        block_lay.addWidget(note)

        # ── 식생보전등급 분포 현황 ──
        block_lay.addWidget(_make_sent_widget(_build_veg_grade_sentences(data)))
        rl_hdr = rl or "조사지역"
        grade_tbl = _make_tbl(["구 분", f"{rl_hdr} 면적(㎡)", f"{rl_hdr} 구성비(%)", "조사지역 면적(㎡)", "조사지역 구성비(%)"])
        _fill_vegetation_grade_table(grade_tbl, data)
        _add_area_table_section("조사지역의 식생보전등급 분포 현황", grade_tbl)
        block_lay.addStretch()

        return block_page

    def _render_area_blocks(blocks: list[dict]):
        nonlocal current_area_blocks
        current_area_blocks = list(blocks or [])
        _clear_area_results()
        if not blocks:
            area_result_lay.addWidget(area_hint)
            return
        if len(blocks) == 1:
            area_result_lay.addWidget(_make_area_block_widget(blocks[0]))
            return
        sub_tabs = QTabWidget()
        sub_tabs.setDocumentMode(True)
        sub_tabs.setStyleSheet(_VEG_WHITE_QSS)
        for block in blocks:
            sub_tabs.addTab(_make_area_block_widget(block), str(block.get("label") or "구간"))
        area_result_lay.addWidget(sub_tabs)

    def _load_area_excel():
        path, _ = QFileDialog.getOpenFileName(page, "식생 면적 엑셀 선택", "", "Excel Files (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            blocks = _parse_vegetation_area_xlsx(path)
            _render_area_blocks(blocks)
        except Exception as e:
            QMessageBox.critical(page, "식생 면적 불러오기 실패", f"식생 면적 엑셀을 읽을 수 없습니다.\n{e}")

    def _refresh_area_blocks():
        if current_area_blocks:
            _render_area_blocks(current_area_blocks)

    btn_generate.clicked.connect(_generate)
    btn_reset.clicked.connect(_reset)
    btn_copy.clicked.connect(_copy)
    btn_load_area.clicked.connect(_load_area_excel)
    if parent_window is not None and hasattr(parent_window, "sig_chart_settings"):
        parent_window.sig_chart_settings.connect(_refresh_area_blocks)
    if parent_window is not None and hasattr(parent_window, "sig_sent_settings"):
        parent_window.sig_sent_settings.connect(_refresh_area_blocks)

    return page


def make_vegetation_tab(parent_window=None):
    return _make_vegetation_tab(parent_window)


def build_plant_tab(stats: PlantStats, sheet: ParsedPlant, prot_sheet=None, parent_window=None, aux_sheets=None):
    aux_sheets = aux_sheets or {}
    round_species_cache = {}
    nat_sheet = aux_sheets.get("귀화식물")
    disturb_sheet = aux_sheets.get("교란생물")
    sentence_table_registry = {}

    outer = QWidget()
    outer.setStyleSheet(_QSS)
    lay = QVBoxLayout(outer)
    lay.setContentsMargins(0, 0, 0, 0)

    top = QTabWidget()
    top.setDocumentMode(True)
    top.setStyleSheet(_tab_qss())

    rns_fld = [r for r in sheet.meta.round_names if r.startswith("현지_") and r != "현지_합계"]
    rns_lit = [r for r in sheet.meta.round_names if r.startswith("문헌_") and r != "문헌_합계"]
    if not rns_lit:
        rns_lit = [k for k in sheet.meta.field_cols if k.startswith("문헌_") and "_합계" not in k]

    fld_all = _species_in_rounds(sheet.species, rns_fld, round_species_cache)
    lit_all = _species_in_rounds(sheet.species, rns_lit, round_species_cache)

    def _plant_progress(msg: str, ratio: float):
        if parent_window and hasattr(parent_window, "step_sub_progress"):
            parent_window.step_sub_progress(msg, ratio)

    if fld_all:
        top.addTab(_survey_tab("현지조사", fld_all, rns_fld, aux_sheets, parent_window, sheet.name, sentence_table_registry, progress_share=0.78), "🔭  현지조사")
    else:
        _plant_progress(f"화면(UI) 생성 중 ({sheet.name}: 현지조사 없음)", 0.05)
        e = QWidget(); l = QVBoxLayout(e); l.addWidget(QLabel("현지조사 데이터 없음")); l.addStretch()
        top.addTab(e, "🔭  현지조사")

    if lit_all:
        _plant_progress(f"화면(UI) 생성 중 ({sheet.name}: 문헌조사)", 0.08)
        lit_panel = _lit_panel_plant(lit_all, rns_lit, prot_sheet, nat_sheet, disturb_sheet, aux_sheets=aux_sheets, parent_window=parent_window)
        top.addTab(lit_panel, "📚  문헌조사")
        sentence_table_registry[("문헌조사", "전체")] = getattr(lit_panel, "_table_refs", {})
    else:
        _plant_progress(f"화면(UI) 생성 중 ({sheet.name}: 문헌조사 없음)", 0.05)
        e = QWidget(); l = QVBoxLayout(e); l.addWidget(QLabel("문헌조사 데이터 없음")); l.addStretch()
        top.addTab(e, "📚  문헌조사")

    _plant_progress(f"화면(UI) 생성 중 ({sheet.name}: 분석 문장)", 0.07)
    top.addTab(
        _combined_sent_view(fld_all, rns_fld, lit_all, rns_lit, prot_sheet, aux_sheets, parent_window, sentence_table_registry),
        "✏  분석 문장",
    )

    lay.addWidget(top)

    if parent_window:
        def _refresh_sents():
            for w in outer.findChildren(QWidget):
                fn = getattr(w, "_sent_refresh_fn", None)
                if callable(fn):
                    fn()
        if hasattr(parent_window, "sig_sent_settings"):
            parent_window.sig_sent_settings.connect(_refresh_sents)
        if hasattr(parent_window, "sig_chart_settings"):
            parent_window.sig_chart_settings.connect(_refresh_sents)

    return outer
