# parser.py — 분류군별 xlsx 파싱

from openpyxl import load_workbook
from dataclasses import dataclass, field
from typing import Optional
import re
from sheet_classifier import classify_sheet
from shared import _s, _has


def _norm_header(v) -> str:
    # 공백/밑줄 제거 후 소문자로 맞춰 헤더 변형 표기를 흡수한다.
    return _s(v).replace(" ", "").replace("_", "").lower()

def _is_order(raw: str) -> bool:
    r = raw.strip()
    return r.startswith("Order ") or r.startswith("Phylum ") or r.startswith("Class ")

def _is_family(raw: str) -> bool:
    r = raw.strip()
    return r.startswith("Family ") or r.startswith("___Family_")

def _is_species(raw: str) -> bool:
    if not raw: return False
    if raw[0] == " ": return False
    if raw.startswith("출현종수") or raw.startswith("개체수"): return False
    if _is_order(raw) or _is_family(raw): return False
    return True

def _clean_taxon(raw: str) -> str:
    r = raw.strip()
    for prefix in ("Order ", "Phylum ", "Class ", "Family ", "___Family_"):
        if r.startswith(prefix):
            return r[len(prefix):]
    return r

def _count_int(v) -> int:
    if isinstance(v, str):
        m = re.search(r"\d+(?:\.\d+)?", v.replace(",", ""))
        if not m:
            return 0
        v = m.group(0)
    try:
        return max(0, int(float(v or 0)))
    except Exception:
        return 0

def _trace_codes(v) -> str:
    s = _s(v)
    if not s or s in ("None", "-"):
        return s
    s = re.sub(r"\d+(?:\.\d+)?", "", s)
    s = s.replace("(", ",").replace(")", ",")
    s = s.replace("/", ",").replace("+", ",").replace(";", ",")
    parts = [p.strip() for p in s.split(",") if p.strip()]
    out = []
    seen = set()
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return ",".join(out)


# ── 결과 데이터클래스 ─────────────────────────────────────────────────────────
@dataclass
class SpeciesRow:
    sci_name: str
    kor_name: str
    order: str
    family: str
    rounds: dict          # {열이름: 값}
    remark: str = ""
    order_kor: str = ""
    family_kor: str = ""

@dataclass
class SurveyMeta:
    field_cols: dict      # 열이름→col index
    round_names: list     # 순서 보존된 회차 이름 목록
    has_field: bool
    has_lit: bool

@dataclass
class ParsedSheet:
    name: str
    taxon: str
    meta: SurveyMeta
    species: list         # list[SpeciesRow]

@dataclass
class ProtectedRow:
    group: str
    kor_name: str
    extinction: str
    monument: str
    rounds: dict
    remark: str = ""
    region: str = ""

@dataclass
class ParsedProtected:
    name: str
    species: list         # list[ProtectedRow]
    round_names: list

@dataclass
class PlantRow:
    sci_name: str
    kor_name: str
    family: str
    grade: str            # 구계학적/습생 등급 (없으면 "")
    rounds: dict
    remark: str = ""
    class_name: str = ""  # 클래스명 (예: 백합강, 목련강) — 단자엽/쌍자엽 판별에 사용

@dataclass
class ParsedPlant:
    name: str
    taxon: str
    meta: SurveyMeta
    species: list         # list[PlantRow]

@dataclass
class PlantSpecialRow:
    kor_name: str
    category: str         # 희귀식물 / 특산식물
    rounds: dict

@dataclass
class ParsedPlantSpecial:
    name: str
    meta: SurveyMeta
    species: list         # list[PlantSpecialRow]

@dataclass
class AquaticRow:
    sci_name: str
    kor_name: str
    order: str
    family: str
    family_kor: str
    rounds: dict          # {회차_지점: 개체수}
    total: int
    ra_pct: float
    phylum: str = ""
    class_name: str = ""
    remark: str = ""
    order_kor: str = ""
    score_tesb: float = 0.0   # Qi(TESB) — 수환경평가 TESB 점수
    score_aesb: float = 0.0   # Qi(AESB) — 수환경평가 AESB 점수

@dataclass
class ParsedAquatic:
    name: str
    taxon: str
    meta: SurveyMeta
    species: list         # list[AquaticRow]


# ── 헤더 파서 (육상 공통: 헤더 2행) ─────────────────────────────────────────
def _parse_land_header(row1, row2):
    """
    row1: 학명 / 국명 / 문헌조사(구간) / 현지조사(구간) / 생활형 / 비고
    row2: None / None / 1,2,...,N / 지점명1,지점명2,총합,우점도 / None / None
    반환: SurveyMeta
    """
    section_ranges = []  # (시작col, "문헌"|"현지")
    field_cols = {}

    for ci, v in enumerate(row1):
        s = _s(v)
        nh = _norm_header(v)
        if nh == "학명": field_cols["sci"] = ci
        elif nh == "국명": field_cols["kor"] = ci
        elif "문헌" in s: section_ranges.append((ci, "문헌"))
        elif "현지" in s: section_ranges.append((ci, "현지"))
        elif s == "생활형": field_cols["life"] = ci
        elif s == "비고": field_cols["remark"] = ci

    def _sec(ci):
        sec = None
        for start, name in section_ranges:
            if ci >= start: sec = name
        return sec

    round_names = []
    for ci, v in enumerate(row2):
        if v is None: continue
        s = _s(v)
        if s in ("", "None"): continue
        sec = _sec(ci)
        if not sec: continue
        if s in ("총합", "합계"):
            key = f"{sec}_합계"
            field_cols[key] = ci
        elif s == "우점도":
            field_cols["우점도"] = ci
        else:
            key = f"{sec}_{s}"
            field_cols[key] = ci
            round_names.append(key)

    has_field = any("현지" in k for k in field_cols)
    has_lit   = any("문헌" in k for k in field_cols)

    # 현지합계가 있으면 round_names 맨앞에
    if "현지_합계" in field_cols and "현지_합계" not in round_names:
        round_names = ["현지_합계"] + round_names

    return SurveyMeta(field_cols=field_cols, round_names=round_names,
                      has_field=has_field, has_lit=has_lit)


def _parse_aquatic_header_fixed(row1, row2, row3):
    section_ranges = []
    field_cols = {}

    for ci, v in enumerate(row1):
        s = _s(v)
        hs = _norm_header(v)
        if hs in ("학명", "scientificname") or "학명" in hs:
            field_cols["sci"] = ci
        elif hs in ("국명", "한글명", "국문명") or "국명" in hs:
            field_cols["kor"] = ci
        elif hs in ("과명", "family") or "과명" in hs:
            field_cols["family"] = ci
        elif hs in ("목명", "order") or "목명" in hs:
            field_cols["order"] = ci
        elif any(tok in s for tok in ("현지조사", "현지 조사", "현지")):
            section_ranges.append((ci, "현지"))
        elif any(tok in s for tok in ("문헌조사", "문헌 조사", "기존문헌", "문헌")):
            section_ranges.append((ci, "문헌"))
        elif s == "비고":
            field_cols["remark"] = ci
        elif "tesb" in hs:
            field_cols["qi_tesb"] = ci
        elif "aesb" in hs:
            field_cols["qi_aesb"] = ci

    def _sec(ci):
        sec = None
        for start, name in section_ranges:
            if ci >= start:
                sec = name
        return sec

    def _is_total_label(text):
        return _s(text) in ("합계", "종합", "총합")

    def _is_ra_label(text):
        return _s(text).upper() in ("RA(%)", "RA", "상대풍부도", "상대풍부도(%)")

    round_names = []
    total_col = None
    ra_col = None
    cur_round = ""
    prev_sec = None   # 섹션 전환 감지용

    for ci, v in enumerate(row2):
        s2 = _s(v)
        v3 = _s(row3[ci]) if ci < len(row3) else ""
        sec = _sec(ci)

        # 현지 → 문헌 등 섹션이 바뀌면 차수 초기화 (이전 섹션 차수 오염 방지)
        if sec != prev_sec:
            prev_sec = sec
            cur_round = ""

        if s2 not in ("", "None") and not _is_total_label(s2) and not _is_ra_label(s2):
            cur_round = s2

        if _is_ra_label(s2) or _is_ra_label(v3) or _is_ra_label(_s(row1[ci]) if ci < len(row1) else ""):
            ra_col = ci
            continue

        if _is_total_label(s2) or _is_total_label(v3):
            if sec:
                key = f"{sec}_{cur_round}_합계" if cur_round else f"{sec}_합계"
                field_cols[key] = ci
                if sec == "현지":
                    round_names.append(key)
            else:
                total_col = ci
            continue

        if not sec or v3 in ("", "None"):
            continue

        # 현지 섹션: round 없으면 "1차" 기본값 (3단계 키 보장 → _detail_rns 필터 통과)
        # 문헌 섹션: round 없으면 2단계 키 허용 (문헌제목 앞에 "1차_" 붙지 않게)
        if sec == "현지":
            round_label = cur_round if cur_round else "1차"
            key = f"{sec}_{round_label}_{v3}"
        else:
            key = f"{sec}_{cur_round}_{v3}" if cur_round else f"{sec}_{v3}"
        field_cols[key] = ci
        if sec == "현지":
            round_names.append(key)

    field_cols["_total"] = total_col
    field_cols["_ra"] = ra_col
    has_field = any("현지" in k for k in field_cols)
    has_lit = any("문헌" in k for k in field_cols)
    return SurveyMeta(field_cols=field_cols, round_names=round_names,
                      has_field=has_field, has_lit=has_lit)


def _has_aquatic_lit_mark(v):
    s = _s(v)
    return s not in ("", "None", "-", "x", "X")


# ── 육상 동물 파싱 ────────────────────────────────────────────────────────────
def parse_land_sheet(rows: list, sheet_name: str, taxon: str) -> ParsedSheet:
    if len(rows) < 2:
        return ParsedSheet(name=sheet_name, taxon=taxon,
                           meta=SurveyMeta({}, [], False, False), species=[])

    meta = _parse_land_header(rows[0], rows[1])
    fc   = meta.field_cols
    sci_c = fc.get("sci", 0)
    kor_c = fc.get("kor", 1)
    rem_c = fc.get("remark")

    species = []
    cur_order = cur_order_kor = cur_family = cur_family_kor = ""

    for row in rows[2:]:
        raw = _s(row[sci_c]) if sci_c < len(row) else ""
        if not raw: continue
        if raw.startswith("출현종수") or raw.startswith("개체수"): continue

        if _is_order(raw):
            cur_order  = _clean_taxon(raw)
            cur_order_kor = _s(row[kor_c]) if kor_c < len(row) else ""
            cur_family = ""
            cur_family_kor = ""
            continue
        if _is_family(raw):
            cur_family = _clean_taxon(raw)
            cur_family_kor = _s(row[kor_c]) if kor_c < len(row) else ""
            continue
        if not _is_species(raw): continue

        rounds = {}
        for rname, ci in fc.items():
            if rname in ("sci","kor","remark","우점도"): continue
            val = row[ci] if ci < len(row) else None
            if taxon in ("bird", "insect") and rname.startswith("현지_"):
                n = _count_int(val)
                val = n if n else val
            elif taxon in ("mammal", "herp") and rname.startswith("현지_"):
                val = _trace_codes(val)
            rounds[rname] = val

        species.append(SpeciesRow(
            sci_name=raw,
            kor_name=_s(row[kor_c]) if kor_c < len(row) else "",
            order=cur_order, order_kor=cur_order_kor,
            family=cur_family, family_kor=cur_family_kor,
            rounds=rounds,
            remark=_s(row[rem_c]) if rem_c is not None and rem_c < len(row) else "",
        ))

    return ParsedSheet(name=sheet_name, taxon=taxon, meta=meta, species=species)


# ── 수생 동물 파싱 (헤더 3행) ────────────────────────────────────────────────
def _parse_aquatic_header(row1, row2, row3):
    """
    row1: 학명 / 국명 / 현지조사(구간) / 문헌조사(구간) / 비고
    row2: None / None / 1차,None,..,2차,.. / ... / ...  (차수)
    row3: None / None / Wa.1,Wa.2,종합,.. / ...         (지점)
    """
    section_ranges = []
    field_cols = {}

    for ci, v in enumerate(row1):
        s = _s(v)
        if s in ("학__명","학명"): field_cols["sci"] = ci
        elif s in ("국__명","국명"): field_cols["kor"] = ci
        elif any(tok in s for tok in ("현지조사", "현지 조사", "현지")):
            section_ranges.append((ci, "현지"))
        elif any(tok in s for tok in ("문헌조사", "문헌 조사", "기존문헌", "문헌")):
            section_ranges.append((ci, "문헌"))
        elif s == "비고": field_cols["remark"] = ci

    def _sec(ci):
        sec = None
        for start, name in section_ranges:
            if ci >= start: sec = name
        return sec

    # row2: 차수 구간 (None이면 직전 차수 유지)
    cur_round = ""
    round_starts = {}  # col → 차수명
    for ci, v in enumerate(row2):
        if v is None: continue
        s = _s(v)
        if s in ("","None"): continue
        if s in ("합계","RA(%)"): continue
        cur_round = s
        round_starts[ci] = s

    # row3: 지점명
    round_names = []
    ra_col = total_col = None
    # row2에서 합계/RA 컬럼 먼저 처리
    for ci, v in enumerate(row1):
        s = _s(v)
        hs = _norm_header(v)

        if hs in ("학명", "scientificname") or "학명" in hs:
            field_cols["sci"] = ci
        elif hs in ("국명", "한글명", "국문명") or "국명" in hs:
            field_cols["kor"] = ci
        elif hs in ("과명", "family") or "과명" in hs:
            field_cols["family"] = ci
        elif hs in ("목명", "order") or "목명" in hs:
            field_cols["order"] = ci
        elif "현지" in s:
            section_ranges.append((ci, "현지"))
        elif "문헌" in s:
            section_ranges.append((ci, "문헌"))
        elif s == "비고":
            field_cols["remark"] = ci

    cur_r = ""
    for ci, v in enumerate(row2):
        s2 = _s(v)
        if s2 not in ("","None"): cur_r = s2
        v3 = _s(row3[ci]) if ci < len(row3) else ""
        if v3 in ("","None"): continue
        sec = _sec(ci)
        if not sec: continue
        if _s(row2[ci]) in ("합계","RA(%)"): continue  # 이미 처리됨
        if sec == "현지":
            round_lbl = cur_r if cur_r else "1차"
            key = f"{sec}_{round_lbl}_종합" if v3 == "종합" else f"{sec}_{round_lbl}_{v3}"
        else:
            key = f"{sec}_{cur_r}_종합" if v3 == "종합" else (f"{sec}_{cur_r}_{v3}" if cur_r else f"{sec}_{v3}")
        field_cols[key] = ci
        if sec == "현지":
            round_names.append(key)

    field_cols["_total"] = total_col
    field_cols["_ra"]    = ra_col

    has_field = any("현지" in k for k in field_cols)
    has_lit   = any("문헌" in k for k in field_cols)
    return SurveyMeta(field_cols=field_cols, round_names=round_names,
                      has_field=has_field, has_lit=has_lit)


def parse_aquatic_sheet(rows: list, sheet_name: str, taxon: str) -> ParsedAquatic:
    if len(rows) < 3:
        return ParsedAquatic(name=sheet_name, taxon=taxon,
                             meta=SurveyMeta({}, [], False, False), species=[])

    meta  = _parse_aquatic_header_fixed(rows[0], rows[1], rows[2])
    fc      = meta.field_cols
    sci_c   = fc.get("sci", 0)
    kor_c   = fc.get("kor", 1)
    ord_c   = fc.get("order")
    fam_c   = fc.get("family")
    rem_c   = fc.get("remark")
    tot_c   = fc.get("_total")
    ra_c    = fc.get("_ra")
    qi_tesb_c = fc.get("qi_tesb")
    qi_aesb_c = fc.get("qi_aesb")

    species = []
    cur_phylum = cur_class = ""
    cur_order = cur_order_kor = cur_family = cur_family_kor = ""

    for row in rows[3:]:
        raw = _s(row[sci_c]) if sci_c < len(row) else ""
        if not raw: continue
        if raw.startswith("출현종수") or raw.startswith("개체수"): continue

        # 과/목이 별도 컬럼으로 제공되는 서식(저서상 템플릿) 대응
        if fam_c is not None and fam_c < len(row):
            fam_cell = _s(row[fam_c])
            if fam_cell and fam_cell not in ("출현종수", "개체수"):
                cur_family = fam_cell
                # 따로 과명 컬럼이 있을 때, 동일 행의 국명 컬럼에 과 국명이 있을 수도 있음
                cur_family_kor = _s(row[kor_c]) if kor_c < len(row) and _s(row[kor_c]) else cur_family_kor
        if ord_c is not None and ord_c < len(row):
            ord_cell = _s(row[ord_c])
            if ord_cell and ord_cell not in ("출현종수", "개체수"):
                cur_order = ord_cell
                cur_order_kor = _s(row[kor_c]) if kor_c < len(row) and _s(row[kor_c]) else cur_order_kor

        if raw.startswith("Phylum "):
            cur_phylum = _clean_taxon(raw)
            cur_class = ""
            cur_order = ""
            cur_order_kor = ""
            cur_family = ""
            cur_family_kor = ""
            continue
        if raw.startswith("Class "):
            cur_class = _clean_taxon(raw)
            cur_order = ""
            cur_order_kor = ""
            cur_family = ""
            cur_family_kor = ""
            continue
        if _is_order(raw):
            cur_order  = _clean_taxon(raw)
            cur_order_kor = _s(row[kor_c]) if kor_c < len(row) else ""
            cur_family = ""
            cur_family_kor = ""
            continue
        if _is_family(raw):
            cur_family = _clean_taxon(raw)
            cur_family_kor = _s(row[kor_c]) if kor_c < len(row) else ""
            continue
        if raw[0] == " ": continue

        rounds = {}
        _skip_keys = {"sci", "kor", "remark", "family", "order", "qi_tesb", "qi_aesb"}
        for rname, ci_r in fc.items():
            if rname.startswith("_") or rname in _skip_keys: continue
            val = row[ci_r] if ci_r < len(row) else None
            if rname.startswith("문헌_"):
                val = 1 if _has_aquatic_lit_mark(val) else 0
            elif rname.startswith("현지_"):
                n = _count_int(val)
                val = n if n else val
            rounds[rname] = val

        total = 0
        if tot_c is not None and tot_c < len(row):
            total = _count_int(row[tot_c])
        if total == 0:
            total = sum(_count_int(v) for k, v in rounds.items() if k.startswith("현지_"))
        ra = 0.0
        if ra_c is not None and ra_c < len(row):
            try: ra = float(row[ra_c] or 0)
            except: ra = 0.0

        score_tesb = 0.0
        if qi_tesb_c is not None and qi_tesb_c < len(row):
            try: score_tesb = float(row[qi_tesb_c] or 0)
            except: score_tesb = 0.0

        score_aesb = 0.0
        if qi_aesb_c is not None and qi_aesb_c < len(row):
            try: score_aesb = float(row[qi_aesb_c] or 0)
            except: score_aesb = 0.0

        species.append(AquaticRow(
            sci_name=raw,
            kor_name=_s(row[kor_c]) if kor_c < len(row) else "",
            phylum=cur_phylum,
            class_name=cur_class,
            order=cur_order, order_kor=cur_order_kor,
            family=cur_family,
            family_kor=cur_family_kor,
            rounds=rounds, total=total, ra_pct=ra,
            remark=_s(row[rem_c]) if rem_c is not None and rem_c < len(row) else "",
            score_tesb=score_tesb,
            score_aesb=score_aesb,
        ))

    return ParsedAquatic(name=sheet_name, taxon=taxon, meta=meta, species=species)


# ── 식물 파싱 ─────────────────────────────────────────────────────────────────
def parse_plant_sheet(rows: list, sheet_name: str, taxon: str) -> ParsedPlant:
    """식물상 / 귀화식물 / 교란생물 / 멸종위기종 / 습생식물 공통"""
    if len(rows) < 2:
        return ParsedPlant(name=sheet_name, taxon=taxon,
                           meta=SurveyMeta({}, [], False, False), species=[])

    # 등급 열이 있는지 확인 (구계학적·습생·멸종위기 시트)
    has_grade = _s(rows[0][0]) == "등급"
    offset = 1 if has_grade else 0

    # 헤더 파싱
    r0 = list(rows[0])
    r1 = list(rows[1])
    meta = _parse_land_header(r0, r1)
    fc   = meta.field_cols
    # has_grade 시: r0[0]=등급, r0[1]=학명, r0[2]=국명 → 헤더 파서가 직접 인식
    sci_c = fc.get("sci", 1 if has_grade else 0)
    kor_c = fc.get("kor", 2 if has_grade else 1)
    rem_c = fc.get("remark")

    species = []
    cur_class  = ""
    cur_family = ""
    cur_grade  = ""   # 등급 carry-forward: A열 등급은 그룹 첫 행에만 있음

    for row in rows[2:]:
        if not row: continue
        row = list(row)

        # 등급 carry-forward: A열에 값이 있을 때만 갱신
        if has_grade:
            g_val = _s(row[0]) if row else ""
            if g_val:
                cur_grade = g_val
        grade = cur_grade if has_grade else ""

        raw_sci = _s(row[sci_c]) if sci_c < len(row) else ""
        if not raw_sci: continue
        # 출현종수·합계 행 스킵
        if raw_sci.startswith("출현종수"): continue
        if "합계" in raw_sci: continue

        # Class 행 추적 — 단자엽(백합강) / 쌍자엽(목련강) 판별에 사용
        if raw_sci.startswith("Class "):
            cur_class  = _clean_taxon(raw_sci)
            cur_family = ""
            continue

        if _is_family(raw_sci):
            cur_family = _clean_taxon(raw_sci)
            continue
        if raw_sci[0] in (" ", "_") and not raw_sci.startswith("___Family"):
            continue

        rounds = {}
        for rname, ci in fc.items():
            if rname in ("sci","kor","remark","우점도"): continue
            rounds[rname] = row[ci] if ci < len(row) else None

        species.append(PlantRow(
            sci_name=raw_sci.replace("_", " "),
            kor_name=_s(row[kor_c]) if kor_c < len(row) else "",
            family=cur_family, grade=grade,
            rounds=rounds,
            remark=_s(row[rem_c]) if rem_c is not None and rem_c < len(row) else "",
            class_name=cur_class,
        ))

    return ParsedPlant(name=sheet_name, taxon=taxon, meta=meta, species=species)


def parse_plant_special(rows: list, sheet_name: str) -> ParsedPlantSpecial:
    """희귀·특산식물: 등급(A carry-forward) / 학명(B) / 국명(C) / 현지... / 문헌... 구조"""
    if len(rows) < 2:
        return ParsedPlantSpecial(name=sheet_name,
                                  meta=SurveyMeta({}, [], False, False), species=[])

    r0 = list(rows[0])
    r1 = list(rows[1])

    # 구계학적 특정식물과 동일 구조: A=등급, B=학명, C=국명
    has_grade = _s(r0[0]) == "등급"

    meta = _parse_land_header(r0, r1)
    fc   = meta.field_cols

    if has_grade:
        kor_c = fc.get("kor", 2)   # 국명: C열 기본
    else:
        fc.setdefault("sci", 0)
        fc.setdefault("kor", 1)
        kor_c = fc["kor"]          # 헤더 파싱으로 국명 열 결정

    species = []
    cur_grade = ""

    for row in rows[2:]:
        if not row: continue
        row = list(row)

        if has_grade:
            # 등급 carry-forward
            g_val = _s(row[0]) if row else ""
            if g_val:
                cur_grade = g_val

            sci_val = _s(row[1]) if len(row) > 1 else ""
            if not sci_val: continue
            if "합계" in sci_val: continue   # "CR등급 합계" 등 스킵

            kor = _s(row[kor_c]) if kor_c < len(row) else ""
            cat = cur_grade
        else:
            kor = _s(row[kor_c]) if kor_c < len(row) else ""
            # category: kor 열이 아닌 열 중 두 번째 열 (옛 A=국명, B=구분 형식)
            _cat_c = 1 if kor_c != 1 else 0
            cat = _s(row[_cat_c]) if _cat_c < len(row) and _cat_c != kor_c else ""
            if not kor or kor.startswith("출현종수"): continue

        rounds = {}
        for rname, ci in fc.items():
            if rname in ("sci", "kor"): continue
            rounds[rname] = row[ci] if ci < len(row) else None
        species.append(PlantSpecialRow(kor_name=kor, category=cat, rounds=rounds))

    return ParsedPlantSpecial(name=sheet_name, meta=meta, species=species)


# ── 법정보호종 파싱 ───────────────────────────────────────────────────────────
def parse_protected(rows: list, sheet_name: str) -> ParsedProtected:
    if len(rows) < 2:
        return ParsedProtected(name=sheet_name, species=[], round_names=[])

    r0, r1 = list(rows[0]), list(rows[1])
    field_cols = {}
    section_ranges = []

    for ci, v in enumerate(r0):
        s = _s(v)
        if s in ("구__분", "구 분", "구분"): field_cols["group"] = ci
        elif s in ("종__명", "종  명", "종명"): field_cols["name"] = ci
        elif "멸종" in s: field_cols["extinction"] = ci
        elif "천연" in s: field_cols["monument"] = ci
        elif any(tok in s for tok in ("지역", "시도", "시·도", "지자체")):
            field_cols["region"] = ci
        elif "현지" in s: section_ranges.append((ci, "현지"))
        elif "문헌" in s: section_ranges.append((ci, "문헌"))
        elif s == "비고": field_cols["remark"] = ci

    def _sec(ci):
        sec = None
        for start, name in section_ranges:
            if ci >= start: sec = name
        return sec

    round_names = []
    for ci, v in enumerate(r1):
        s = _s(v)
        if not s or s in ("None",): continue
        sec = _sec(ci)
        if not sec: continue
        if s in ("합계", "총합"):
            field_cols[f"{sec}_합계"] = ci
        else:
            key = f"{sec}_{s}"
            field_cols[key] = ci
            round_names.append(key)

    species = []
    cur_group = ""
    gc = field_cols.get("group", 0)
    nc = field_cols.get("name", 1)
    ec = field_cols.get("extinction", 2)
    mc = field_cols.get("monument", 3)
    rc = field_cols.get("remark")
    reg_c = field_cols.get("region")

    for row in rows[2:]:
        g = _s(row[gc]) if gc < len(row) else ""
        if g and g != "출현종수": cur_group = g
        if g == "출현종수": continue
        kn = _s(row[nc]) if nc < len(row) else ""
        if not kn or kn == "출현종수": continue
        rounds = {}
        for rname, ci in field_cols.items():
            if rname in ("group","name","extinction","monument","remark"): continue
            rounds[rname] = row[ci] if ci < len(row) else None
        species.append(ProtectedRow(
            group=cur_group, kor_name=kn,
            extinction=_s(row[ec]) if ec < len(row) else "",
            monument=_s(row[mc]) if mc < len(row) else "",
            rounds=rounds,
            remark=_s(row[rc]) if rc is not None and rc < len(row) else "",
            region=_s(row[reg_c]) if reg_c is not None and reg_c < len(row) else "",
        ))

    return ParsedProtected(name=sheet_name, species=species, round_names=round_names)


# ── 전체 로드 ─────────────────────────────────────────────────────────────────
def load_xlsx(path: str) -> dict:
    """
    반환: {시트명: ParsedSheet | ParsedAquatic | ParsedPlant |
                   ParsedPlantSpecial | ParsedProtected}
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {}

    for name in wb.sheetnames:
        taxon = classify_sheet(name)
        if taxon is None:
            n = "".join(str(name or "").strip().split())
            if "저서" in n:
                taxon = "benthos"
            elif "어류" in n:
                taxon = "fish"
        if taxon is None:
            continue
        rows = list(wb[name].iter_rows(values_only=True))

        if taxon == "protected":
            result[name] = parse_protected(rows, name)
        elif taxon in ("fish", "benthos"):
            result[name] = parse_aquatic_sheet(rows, name, taxon)
        elif taxon in ("plant", "plant_sub"):
            result[name] = parse_plant_sheet(rows, name, taxon)
        elif taxon == "plant_special":
            result[name] = parse_plant_special(rows, name)
        else:  # mammal / bird / herp / insect
            result[name] = parse_land_sheet(rows, name, taxon)

    return result
