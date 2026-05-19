# ui_shared.py — UI 공통 컴포넌트 · 헬퍼 함수
# ══════════════════════════════════════════════════════════════════
# 모든 탭에서 import해서 쓰는 공유 UI 요소입니다.
# ══════════════════════════════════════════════════════════════════
import math
import colorsys
import io
import json
import os
import re
import unicodedata

from PySide6.QtCore    import Qt, QMimeData, Signal, QPoint, QSize, QRect, QSettings
from PySide6.QtGui     import QColor, QKeySequence, QImage, QCursor, QBrush, QPen, QFont, QPainter
from PySide6.QtWidgets import (
    QApplication, QFrame, QHeaderView, QRadioButton, QScrollArea,
    QTableWidget, QTableWidgetItem, QPushButton,
    QDialog, QVBoxLayout, QHBoxLayout, QComboBox, QCheckBox,
    QButtonGroup,
    QDoubleSpinBox, QSpinBox, QFormLayout, QDialogButtonBox, QLabel,
    QSlider, QColorDialog, QSizePolicy, QWidget, QGroupBox, QLineEdit, QMenu,
    QRubberBand, QTabWidget, QListWidget, QStyledItemDelegate, QStyle, QStyleOptionViewItem,
    QProxyStyle, QFileDialog, QMessageBox
)

from config import (
    SETTINGS, _ACCENT, _ACCENT_D, _ACCENT_L, _BORDER, _SUCCESS, _WARN, _ERR, _PATH,
    _BG, _CARD, _TXT, _SUB, FF_KR, FF_EN, BD, BD1, PIE_COLORS, GRAPH_H, CHK_INDICATOR_QSS
)

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    _MPL = True
except ImportError:
    _MPL = False


# ══════════════════════════════════════════════════════════════════
# 탭 바 QSS
# ══════════════════════════════════════════════════════════════════
def make_tab_qss(ff_kr: str, sub_color: str, accent_color: str, big=True) -> str:
    py = "7px" if big else "5px"
    px = "18px" if big else "14px"
    return f"""
        QTabBar {{ qproperty-drawBase: 0; }}
        QTabBar::tab {{
            {ff_kr}; font-size:{"12px" if big else "11px"}; font-weight:600;
            min-height:{"18px" if big else "16px"}; padding:{py} {px};
            background:#EEF2F7; color:{sub_color};
            border:1px solid transparent; border-radius:10px; margin-right:6px;
        }}
        QTabBar::tab:selected {{ background:#FFFFFF; color:{accent_color}; font-weight:700; border-color:#DCE5F0; }}
        QTabBar::tab:hover    {{ background:#F8FBFF; color:{accent_color}; }}
        QTabWidget::pane      {{ border:none; background:transparent; }}
    """


# ══════════════════════════════════════════════════════════════════
# 스크롤 래퍼
# ══════════════════════════════════════════════════════════════════
def make_scroll_widget(widget):
    sc = QScrollArea()
    sc.setWidgetResizable(True)
    sc.setFrameShape(QFrame.NoFrame)
    sc.setWidget(widget)
    return sc


# ══════════════════════════════════════════════════════════════════
# CopyableTableWidget — Ctrl+C 및 버튼 복사 지원
# ══════════════════════════════════════════════════════════════════
class _NoFocusRectDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        opt = QStyleOptionViewItem(option)
        opt.state &= ~QStyle.State_HasFocus
        super().paint(painter, opt, index)


class _NoFocusRectStyle(QProxyStyle):
    def drawPrimitive(self, element, option, painter, widget=None):
        if element == QStyle.PE_FrameFocusRect:
            return
        super().drawPrimitive(element, option, painter, widget)


def install_global_no_focus_rect_style(app=None):
    app = app or QApplication.instance()
    if app is None or getattr(app, "_no_focus_rect_style_installed", False):
        return
    proxy = _NoFocusRectStyle(app.style())
    app.setStyle(proxy)
    app._no_focus_rect_style = proxy
    app._no_focus_rect_style_installed = True


class CopyableTableWidget(QTableWidget):
    """선택 영역을 TSV + HTML(Excel 호환)로 클립보드에 복사."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setItemDelegate(_NoFocusRectDelegate(self))

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            self.copy_selection()
            event.accept()
            return
        super().keyPressEvent(event)

    def copy_selection(self, include_header=False):
        ranges = self.selectedRanges()
        if ranges:
            r    = ranges[0]
            rows = range(r.topRow(),    r.bottomRow()    + 1)
            cols = range(r.leftColumn(), r.rightColumn() + 1)
        else:
            rows = range(self.rowCount())
            cols = range(self.columnCount())

        # 헤더 행 (기본값: 제외)
        headers = [
            (self.horizontalHeaderItem(c).text()
             if self.horizontalHeaderItem(c) else "")
            for c in cols
        ]
        html  = "<html><body><table border='1' style='border-collapse:collapse;'>"
        if include_header and any(headers):
            html += "<tr>"
            html += "".join(
                f"<th style='padding:4px;'>{h}</th>"
                for h in headers
            )
            html += "</tr>"

        lines = []
        if include_header and any(headers):
            lines.append("\t".join(headers))

        for row in rows:
            if self.isRowHidden(row):
                continue
            html += "<tr>"
            vals = []
            for col in cols:
                it   = self.item(row, col)
                text = it.text() if it else ""
                html += f"<td style='padding:4px;'>{text}</td>"
                vals.append(text)
            html += "</tr>"
            lines.append("\t".join(vals))
        html += "</table></body></html>"

        mime = QMimeData()
        mime.setHtml(html)
        mime.setText("\n".join(lines))
        QApplication.clipboard().setMimeData(mime)


# ══════════════════════════════════════════════════════════════════
# 테이블 헬퍼
# ══════════════════════════════════════════════════════════════════
def _item(txt, align=Qt.AlignCenter) -> QTableWidgetItem:
    """기본 QTableWidgetItem 생성 (편집 불가, 정렬 설정)."""
    it = QTableWidgetItem("" if txt is None else str(txt))
    it.setTextAlignment(align)
    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
    return it


def _int(v) -> int:
    """안전한 정수 변환 (실패 시 0)."""
    try:
        return max(0, int(float(v)))
    except Exception:
        return 0


def _fmt(x) -> str:
    """SETTINGS.decimal 자릿수 소수 문자열."""
    try:
        return f"{float(x):.{SETTINGS.decimal}f}"
    except Exception:
        return str(x)


def _pct(x) -> str:
    return _fmt(x) + "%"


def _make_tbl(cols: list, row_h: int = 24) -> CopyableTableWidget:
    """표준 CopyableTableWidget 생성 (헤더 설정 포함)."""
    t = CopyableTableWidget(0, len(cols))
    t.setHorizontalHeaderLabels(cols)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setAlternatingRowColors(True)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(row_h)
    t.horizontalHeader().setStretchLastSection(True)
    t.setSelectionMode(QTableWidget.ExtendedSelection)
    t.setSelectionBehavior(QTableWidget.SelectItems)
    t.setToolTip("셀 선택 후 Ctrl+C 로 복사")
    return t


def _auto_fit_table(tbl: QTableWidget):
    """열을 내용에 맞게 조절하되, 표의 내용이 모두 보일 수 있도록 최소 너비를 보장합니다."""
    h = tbl.horizontalHeader()
    n = tbl.columnCount()
    if n == 0:
        return

    tbl.setWordWrap(False)  # 글자 잘림/줄바꿈 방지

    h.setStretchLastSection(False)
    for c in range(n):
        h.setSectionResizeMode(c, QHeaderView.ResizeToContents)
    tbl.resizeColumnsToContents()

    # columnWidth()는 미표시 상태에서 0 → fontMetrics로 직접 측정
    fm = tbl.fontMetrics()
    hfm = tbl.horizontalHeader().fontMetrics()
    col_widths = []
    for c in range(n):
        hdr = tbl.horizontalHeaderItem(c)
        w = hfm.horizontalAdvance(hdr.text() if hdr else "") + 18
        for r in range(tbl.rowCount()):
            item = tbl.item(r, c)
            if item:
                w = max(w, fm.horizontalAdvance(item.text()) + 18)
        col_widths.append(max(w, 30))

    for c in range(n - 1):
        h.setSectionResizeMode(c, QHeaderView.ResizeToContents)
    if n > 0:
        h.setSectionResizeMode(n - 1, QHeaderView.Stretch)
    h.setStretchLastSection(True)

    content_w = tbl.verticalHeader().width() if tbl.verticalHeader().isVisible() else 0
    content_w += sum(col_widths)
    content_w += tbl.frameWidth() * 2 + 32

    tbl.setMinimumWidth(content_w)
    tbl.setMaximumWidth(16777215)
    tbl.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)


def _tbl_auto_height(tbl: QTableWidget, row_h: int = 24, extra: int = 6) -> None:
    """표 높이를 행 수에 맞게 정확히 고정 (세로 스크롤 제거).
    _fill_tbl() 내부에서 호출하면 내용 변경 시 자동 갱신됩니다."""
    tbl.verticalHeader().setDefaultSectionSize(row_h)
    for r in range(tbl.rowCount()):
        tbl.setRowHeight(r, row_h)
    tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    h_hdr = tbl.horizontalHeader().height()
    tbl.setFixedHeight(h_hdr + row_h * max(tbl.rowCount(), 1) + extra)


# ══════════════════════════════════════════════════════════════════
# 버튼 클릭 피드백
# ══════════════════════════════════════════════════════════════════
def apply_button_feedback(btn: QPushButton, success_text: str = "✔ 완료"):
    from PySide6.QtCore import QTimer
    orig_text = btn.text()
    if orig_text == success_text: return
    orig_style = btn.styleSheet()
    btn.setText(success_text)
    btn.setStyleSheet(orig_style + " color: #059669; border-color: #059669; background: #D1FAE5;")
    QTimer.singleShot(1200, lambda: (btn.setText(orig_text), btn.setStyleSheet(orig_style)))

# ══════════════════════════════════════════════════════════════════
# 소형 윤곽선 버튼 QSS 팩토리 + 색상 변형 상수
# ══════════════════════════════════════════════════════════════════
def make_outline_btn_qss(color_hex: str, hover_bg_hex: str) -> str:
    """소형 윤곽선(Outline) 버튼 QSS 반환."""
    return (
        f"QPushButton{{background:transparent;color:{color_hex};font-size:11px;"
        f"font-weight:bold;border:1px solid {color_hex};border-radius:4px;padding:2px 8px;}}"
        f"QPushButton:hover{{background:{hover_bg_hex};}}"
    )


_COPY_BTN_QSS       = make_outline_btn_qss(_ACCENT,  _ACCENT_L)   # 파란색
_OUTLINE_BTN_GREEN  = make_outline_btn_qss(_SUCCESS, "#D1FAE5")   # 초록색
_OUTLINE_BTN_ORANGE = make_outline_btn_qss(_WARN,    "#FEF3C7")   # 주황색


def make_copy_button(tbl_ref, height: int = 24) -> QPushButton:
    """
    '📋 표 복사' QPushButton 을 반환합니다.
    tbl_ref : CopyableTableWidget 인스턴스 또는 callable() → CopyableTableWidget
    """
    btn = QPushButton("📋 표 복사")
    btn.setFixedHeight(height)
    btn.setStyleSheet(_COPY_BTN_QSS)

    def _do_copy():
        tbl = tbl_ref() if callable(tbl_ref) else tbl_ref
        if tbl and hasattr(tbl, "copy_selection"):
            tbl.copy_selection(include_header=True)
            apply_button_feedback(btn)

    btn.clicked.connect(_do_copy)
    return btn


# ══════════════════════════════════════════════════════════════════
# 행 강조 / 합계 셀
# ══════════════════════════════════════════════════════════════════
def bold_row(tbl, row: int, bg_color: str):
    for c in range(tbl.columnCount()):
        it = tbl.item(row, c)
        if it:
            it.setBackground(QColor(bg_color))
            f = it.font()
            f.setBold(True)
            it.setFont(f)


def set_total_pct_cell(tbl, row, col, pct_value, *,
                       item_factory, pct_formatter, err_color, decimals):
    it = item_factory(pct_formatter(pct_value))
    rounded = round(float(pct_value or 0), decimals)
    tol = 10 ** (-decimals) / 2
    if not math.isclose(rounded, 100.0, rel_tol=0.0, abs_tol=tol):
        it.setForeground(QColor(err_color))
    tbl.setItem(row, col, it)


# ══════════════════════════════════════════════════════════════════
# 차트 축 스타일
# ══════════════════════════════════════════════════════════════════
def make_main_qss(item_pad: str = "3px 6px", header_pad: str = "5px 6px") -> str:
    """aqua / plant 탭 공통 QSS. item_pad, header_pad 만 탭마다 다름."""
    return f"""
QWidget   {{ background:{_BG}; {FF_KR}; font-size:13px; color:{_TXT}; }}
QGroupBox {{
    {FF_KR}; font-size:12px; font-weight:700; color:{_TXT};
    {BD}; border-radius:10px; margin-top:12px; background:{_CARD}; padding-top:8px;
}}
QGroupBox::title {{
    subcontrol-origin:margin; subcontrol-position:top left;
    left:12px; padding:0 6px; color:{_TXT}; font-size:12px;
}}
QTableWidget {{
    {BD}; border-radius:8px; background:{_CARD};
    gridline-color:{_BORDER}; {FF_KR}; font-size:12px;
}}
QTableWidget::item          {{ padding:{item_pad}; }}
QTableWidget::item:selected {{ background:{_ACCENT_L}; color:{_ACCENT}; }}
QHeaderView::section {{
    background:{_BG}; {BD1}; border-right:1px solid {_BORDER};
    padding:{header_pad}; {FF_KR}; font-size:11px; font-weight:700; color:{_SUB};
}}
QScrollArea {{ border:none; background:transparent; }}
QScrollBar:vertical         {{ background:{_BG}; width:6px; border-radius:3px; }}
QScrollBar::handle:vertical {{ background:{_BORDER}; border-radius:3px; min-height:30px; }}
QScrollBar::handle:vertical:hover {{ background:#B0B8C8; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}
"""


def apply_light_chart_axes(ax, *, title=None, ylabel=None, xlabel=None,
                            fs=9, grid_on=False, grid_axis="y",
                            grid_color="#e6e6e6"):
    if title:
        ax.set_title(title, fontsize=max(10, fs + 1), fontweight="bold", pad=10)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=fs)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=fs)
    ax.set_axisbelow(True)
    ax.tick_params(length=0)
    for side in ("top", "right", "left", "bottom"):
        ax.spines[side].set_visible(False)
    if grid_on:
        ax.grid(axis=grid_axis, linestyle="--", alpha=0.5, color=grid_color)


# ══════════════════════════════════════════════════════════════════
# 설정 및 그래프 관련 공유 요소
# ══════════════════════════════════════════════════════════════════
_DIALOG_QSS = f"""
    QDialog {{ background: {_BG}; {FF_KR}; }}
    QLabel, QCheckBox {{ {FF_KR}; font-size: 13px; color: {_TXT}; }}
    QTabWidget::pane {{ border: 1.5px solid {_BORDER}; border-radius: 0 8px 8px 8px; background: {_CARD}; }}
    QTabBar::tab {{
        {FF_KR}; font-size: 13px; font-weight: 600; padding: 6px 16px;
        background: #E9ECEF; border: 1.5px solid {_BORDER};
        border-bottom: none; border-radius: 6px 6px 0 0;
        margin-right: 2px;
    }}
    QTabBar::tab:selected {{ background: {_CARD}; color: {_ACCENT}; font-weight: 700; border-bottom: none; }}
    QPushButton:focus, QCheckBox:focus, QRadioButton:focus, QTabBar::tab:focus {{ outline: none; }}
    QSpinBox, QDoubleSpinBox, QComboBox, QLineEdit, QListWidget {{
        border: 1.5px solid {_BORDER}; border-radius: 6px;
        padding: 4px 8px; background: white; {FF_KR}; font-size: 13px; color: {_TXT};
    }}
    QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:hover, QLineEdit:focus, QListWidget:focus {{ border-color: {_ACCENT}; }}
    QPushButton {{
        background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 {_ACCENT}, stop:1 {_ACCENT_D});
        color: white; border: none; border-radius: 7px;
        {FF_KR}; font-size: 13px; font-weight: 700; padding: 8px 18px;
    }}
    QPushButton:hover {{ background: #3B82F6; }}
""" + CHK_INDICATOR_QSS

class CanvasGroupBox(QGroupBox):
    resized_sig = Signal(int, int)
    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.resized_sig.emit(self.width() - 20, self.height() - 40)

class ResizableCanvasFrame(QFrame):
    resized_sig = Signal(int, int)
    HANDLE = 8

    def __init__(self, canvas, min_w=300, min_h=200, parent=None):
        super().__init__(parent)
        self._cv   = canvas
        self._minW = min_w
        self._minH = min_h
        self._drag_dir = None
        self._drag_origin = QPoint()
        self._orig_w = 0
        self._orig_h = 0

        self.setFrameShape(QFrame.NoFrame)
        self.setFrameShadow(QFrame.Plain)
        self.setStyleSheet(f"ResizableCanvasFrame {{ border: none; background: transparent; }}")
        inner = QVBoxLayout(self)
        inner.setContentsMargins(0, 0, 0, 0)
        inner.addWidget(canvas)
        self.setMouseTracking(True)
        canvas.setMouseTracking(True)

    def _dir_at(self, pos) -> str | None:
        H = self.HANDLE
        w, h = self.width(), self.height()
        x, y = pos.x(), pos.y()
        l = x < H; r = x > w - H; t = y < H; b = y > h - H
        if t and l: return "tl"
        if t and r: return "tr"
        if b and l: return "bl"
        if b and r: return "br"
        if l: return "l"
        if r: return "r"
        if t: return "t"
        if b: return "b"
        return None

    _CURSORS = {
        "l": Qt.SizeHorCursor, "r": Qt.SizeHorCursor,
        "t": Qt.SizeVerCursor, "b": Qt.SizeVerCursor,
        "tl": Qt.SizeFDiagCursor, "br": Qt.SizeFDiagCursor,
        "tr": Qt.SizeBDiagCursor, "bl": Qt.SizeBDiagCursor,
    }

    def mouseMoveEvent(self, event):
        if self._drag_dir:
            self._do_resize(event.globalPosition().toPoint())
        else:
            d = self._dir_at(event.pos())
            self.setCursor(self._CURSORS.get(d, Qt.ArrowCursor))
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            d = self._dir_at(event.pos())
            if d:
                self._drag_dir    = d
                self._drag_origin = event.globalPosition().toPoint()
                self._orig_w = self.width()
                self._orig_h = self.height()
                event.accept()
                return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._drag_dir:
            self._drag_dir = None
            self.setCursor(Qt.ArrowCursor)
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.resized_sig.emit(self.width(), self.height())

    def _do_resize(self, gpos: QPoint):
        dx = gpos.x() - self._drag_origin.x()
        dy = gpos.y() - self._drag_origin.y()
        d  = self._drag_dir
        new_w = self._orig_w
        new_h = self._orig_h
        if "r" in d: new_w = max(self._minW, self._orig_w + dx)
        if "l" in d: new_w = max(self._minW, self._orig_w - dx)
        if "b" in d: new_h = max(self._minH, self._orig_h + dy)
        if "t" in d: new_h = max(self._minH, self._orig_h - dy)
        self.setFixedSize(new_w, new_h)

class CanvasSizeSlider(QWidget):
    valueChanged = Signal(int)
    def __init__(self, min_val, max_val, default_val, parent=None):
        super().__init__(parent)
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)
        self._sld = QSlider(Qt.Horizontal)
        self._sld.setRange(min_val, max_val)
        self._sld.setValue(default_val)
        self._sld.setFixedHeight(22)
        self._lbl = QLabel(f"{default_val}px")
        self._lbl.setFixedWidth(45)
        self._lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lay.addWidget(self._sld, 1)
        lay.addWidget(self._lbl)
        self._sld.valueChanged.connect(self._on_change)
    def _on_change(self, v):
        self._lbl.setText(f"{v}px")
        self.valueChanged.emit(v)
    def value(self) -> int: return self._sld.value()
    def setValue(self, v: int):
        self._sld.setValue(v)
        self._lbl.setText(f"{v}px")
    def setMaximum(self, max_val: int): self._sld.setMaximum(max_val)

class _ColorPicker(QWidget):
    def __init__(self, color: str, parent=None):
        super().__init__(parent)
        self._color = color if QColor(color).isValid() else "#4472C4"
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)
        self.btn = QPushButton("색상 선택")
        self.btn.setFixedHeight(28)
        self.preview = QLabel()
        self.preview.setFixedWidth(72)
        self.preview.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.btn)
        lay.addWidget(self.preview)
        lay.addStretch()
        self.btn.clicked.connect(self._pick)
        self._paint_preview()
    def _paint_preview(self):
        self.preview.setText(self._color.upper())
        self.preview.setStyleSheet(f"border:1.5px solid {_BORDER}; border-radius:6px; padding:3px 8px; background:{self._color}; {FF_EN}; font-size:10px;")
    def _pick(self):
        c = QColorDialog.getColor(QColor(self._color), self, "색상 선택")
        if c.isValid():
            self._color = c.name()
            self._paint_preview()
    def value(self) -> str: return self._color

def _gray_palette(n: int) -> list[str]:
    """어두운 색에서 밝은 색으로 균등 그라데이션 흑백 팔레트."""
    n = max(1, n)
    if n == 1: return ["#595959"]
    dark, light = 0x38, 0xC8
    step = (light - dark) / (n - 1)
    return [f"#{int(dark + i * step):02X}{int(dark + i * step):02X}{int(dark + i * step):02X}"
            for i in range(n)]

def _series_palette(n: int, canvas=None) -> list[str]:
    mode = getattr(SETTINGS, "color_mode", "auto")
    if mode == "gray": return _gray_palette(n)
    if mode == "solid":
        base = getattr(SETTINGS, "bar_color", "#4472C4")
        return [base] * max(1, n)
    if mode == "custom" and canvas is not None:
        custom = getattr(canvas, "_series_colors", None)
        if custom: return [custom[i % len(custom)] for i in range(max(1, n))]
    n = max(1, n)
    if n <= len(PIE_COLORS): return PIE_COLORS[:n]
    colors = list(PIE_COLORS)
    extra = n - len(colors)
    for i in range(extra):
        hue = ((i * 0.61803398875) + 0.08) % 1.0
        sat = 0.52 + (0.10 if i % 2 else 0.0)
        val = 0.84 - (0.08 if i % 3 == 1 else 0.0)
        r, g, b = colorsys.hsv_to_rgb(hue, min(sat, 0.72), max(val, 0.70))
        colors.append(f"#{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}")
    return colors

def _pie_palette(n: int) -> list[str]:
    """파이차트 전용: 단색 모드에도 구분 가능한 색 반환."""
    mode = getattr(SETTINGS, "color_mode", "auto")
    if mode == "gray": return _gray_palette(n)
    n = max(1, n)
    return [PIE_COLORS[i % len(PIE_COLORS)] for i in range(n)]

def _apply_axes_common(ax, *, title: str | None = None, ylabel: str | None = None, xlabel: str | None = None, fs: int = 9):
    if title is None: title = getattr(SETTINGS, "chart_title", "")
    if ylabel is None: ylabel = getattr(SETTINGS, "y_axis_title", "")
    if title: ax.set_title(title, fontsize=max(10, fs + 1), fontweight="bold", pad=10)
    if ylabel: ax.set_ylabel(ylabel, fontsize=fs)
    if xlabel: ax.set_xlabel(xlabel, fontsize=fs)

def _apply_bar_margins(fig, canvas):
    if not canvas: return
    cfg = canvas._cfg
    left = float(cfg.get("bar_margin_left", getattr(SETTINGS, "bar_margin_left", 0.15)))
    right = float(cfg.get("bar_margin_right", getattr(SETTINGS, "bar_margin_right", 0.95)))
    bottom = float(cfg.get("bar_margin_bottom", getattr(SETTINGS, "bar_margin_bottom", 0.15)))
    top = float(cfg.get("bar_margin_top", getattr(SETTINGS, "bar_margin_top", 0.90)))
    
    left = max(0.01, min(0.5, left))
    right = max(0.5, min(0.99, right))
    bottom = max(0.01, min(0.5, bottom))
    top = max(0.5, min(0.99, top))
    fig.subplots_adjust(left=left, right=right, bottom=bottom, top=top)

def _apply_graph_scale(fig, canvas):
    if not canvas: return
    try:
        gw_cap, gh_cap = 1.10, 1.25
        gw = max(0.30, min(gw_cap, float(canvas._cfg.get("graph_width_scale", 1.0))))
        gh = max(0.30, min(gh_cap, float(canvas._cfg.get("graph_height_scale", 1.0))))
        if abs(gw - 1.0) < 0.001 and abs(gh - 1.0) < 0.001:
            return
        max_gw, max_gh = gw_cap, gh_cap
        for ax in list(fig.axes):
            pos = ax.get_position()
            if pos.width > 0:
                cx = pos.x0 + pos.width / 2.0
                max_gw = min(max_gw, ((min(cx - 0.01, 0.99 - cx) * 2.0) / pos.width))
            if pos.height > 0:
                cy = pos.y0 + pos.height / 2.0
                max_gh = min(max_gh, ((min(cy - 0.01, 0.99 - cy) * 2.0) / pos.height))
        gw = max(0.30, min(gw, max_gw))
        gh = max(0.30, min(gh, max_gh))
        canvas._cfg["graph_width_scale"] = gw
        canvas._cfg["graph_height_scale"] = gh
        for ax in list(fig.axes):
            pos = ax.get_position()
            cx = pos.x0 + pos.width / 2.0
            cy = pos.y0 + pos.height / 2.0
            nw = pos.width * gw
            nh = pos.height * gh
            x0 = max(0.01, cx - nw / 2.0)
            y0 = max(0.01, cy - nh / 2.0)
            x1 = min(0.99, cx + nw / 2.0)
            y1 = min(0.99, cy + nh / 2.0)
            ax.set_position([x0, y0, max(0.01, x1 - x0), max(0.01, y1 - y0)])
    except Exception:
        pass

class ChartColorDialog(QDialog):
    """전체 그래프 색상 모드 설정 다이얼로그."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("그래프 색상 설정")
        self.resize(380, 220)
        self.setStyleSheet(_DIALOG_QSS)
        from PySide6.QtWidgets import QButtonGroup, QColorDialog
        from PySide6.QtGui import QColor

        lay = QVBoxLayout(self)

        grp = QGroupBox("색상 모드")
        grp_lay = QVBoxLayout(grp)
        grp_lay.setSpacing(10)

        self.chk_array = QCheckBox("색배열 — 여러 색 자동 배정 (기본)")
        self.chk_solid = QCheckBox("단색 — 모든 막대를 지정 색 하나로")
        self.chk_gray  = QCheckBox("흑백 — 어두운 색~밝은 색 그라데이션")

        self._btn_grp = QButtonGroup(self)
        self._btn_grp.setExclusive(True)
        for cb in (self.chk_array, self.chk_solid, self.chk_gray):
            cb.setCheckable(True)
            self._btn_grp.addButton(cb)
            grp_lay.addWidget(cb)

        # 단색 색상 선택
        color_row = QHBoxLayout()
        self._lbl_color = QLabel("  단색 기본색:")
        self._lbl_color.setStyleSheet(f"font-size:12px;color:#555;{FF_KR};")
        self._btn_swatch = QPushButton()
        self._btn_swatch.setFixedSize(40, 22)
        self._btn_swatch.setToolTip("단색 모드에서 사용할 색 선택")
        self._solid_color = getattr(SETTINGS, "bar_color", "#4472C4")
        self._update_swatch()
        self._btn_swatch.clicked.connect(lambda: self._pick_color(QColorDialog))
        color_row.addWidget(self._lbl_color)
        color_row.addWidget(self._btn_swatch)
        color_row.addStretch()
        grp_lay.addLayout(color_row)

        lay.addWidget(grp)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

        # 현재 설정 반영
        mode = getattr(SETTINGS, "color_mode", "auto")
        if mode == "solid": self.chk_solid.setChecked(True)
        elif mode == "gray": self.chk_gray.setChecked(True)
        else: self.chk_array.setChecked(True)

    def _update_swatch(self):
        c = self._solid_color
        r, g, b = int(c[1:3], 16), int(c[3:5], 16), int(c[5:7], 16)
        lum = 0.299*r + 0.587*g + 0.114*b
        txt = "white" if lum < 140 else "#333"
        self._btn_swatch.setStyleSheet(
            f"QPushButton{{background:{c};border:1px solid #aaa;border-radius:4px;color:{txt};}}"
        )

    def _pick_color(self, QColorDialog):
        from PySide6.QtGui import QColor
        c = QColorDialog.getColor(QColor(self._solid_color), self, "단색 선택")
        if c.isValid():
            self._solid_color = c.name().upper()
            self._update_swatch()

    def apply(self):
        if self.chk_solid.isChecked():
            SETTINGS.color_mode = "solid"
            SETTINGS.bar_color  = self._solid_color
        elif self.chk_gray.isChecked():
            SETTINGS.color_mode = "gray"
        else:
            SETTINGS.color_mode = "auto"


def make_copy_graph_btn(canvas) -> "QPushButton":
    """그래프 설정 패널에 배치할 '🖼 그래프 복사' 버튼.
    클릭 시 해당 캔버스를 PNG로 클립보드에 복사합니다."""
    btn = QPushButton("🖼  그래프 복사")
    btn.setFixedHeight(28)
    btn.setStyleSheet(
        "QPushButton{background:#F0FDF4;color:#059669;border:1px solid #BBF7D0;"
        "border-radius:5px;font-size:12px;padding:2px 10px;}"
        "QPushButton:hover{background:#DCFCE7;}"
    )
    def _on_click():
        canvas._copy()
        apply_button_feedback(btn)
    btn.clicked.connect(_on_click)
    return btn


def make_color_settings_btn(refresh_fn, parent=None) -> "QPushButton":
    """그래프 설정 패널에 배치할 '🎨 색상 설정' 버튼.
    클릭 시 ChartColorDialog를 열고, 확인 후 refresh_fn()을 호출합니다."""
    from PySide6.QtWidgets import QDialog
    btn = QPushButton("🎨  색상 설정")
    btn.setFixedHeight(28)
    btn.setStyleSheet(
        "QPushButton{background:#EFF6FF;color:#2563EB;border:1px solid #BFDBFE;"
        "border-radius:5px;font-size:12px;padding:2px 10px;}"
        "QPushButton:hover{background:#DBEAFE;}"
    )
    def _on_click():
        dlg = ChartColorDialog(parent)
        if dlg.exec() == QDialog.Accepted:
            dlg.apply()
            if callable(refresh_fn):
                refresh_fn()
    btn.clicked.connect(_on_click)
    return btn


def _strip_lit_round_affixes(text: str) -> str:
    s = str(text or "").strip().replace(" ", "")
    s = s.replace("문헌조사", "문헌")
    parts = [p for p in s.split("_") if p]
    if parts and parts[0] == "문헌":
        s = parts[1] if len(parts) > 1 else ""
    s = re.sub(r"^(제|No\.?|NO\.?|no\.?)", "", s)
    s = re.sub(r"(차시|차|회|번째|번)$", "", s)
    return s


def _circled_number_value(text: str) -> int | None:
    if len(text) != 1:
        return None
    try:
        num = unicodedata.numeric(text)
    except (TypeError, ValueError):
        return None
    if num.is_integer() and 1 <= int(num) <= 50:
        name = unicodedata.name(text, "")
        if "CIRCLED" in name and "DIGIT" in name or "CIRCLED NUMBER" in name:
            return int(num)
    return None


def _roman_number_value(text: str) -> int | None:
    roman_chars = {
        "Ⅰ": "I", "Ⅱ": "II", "Ⅲ": "III", "Ⅳ": "IV", "Ⅴ": "V",
        "Ⅵ": "VI", "Ⅶ": "VII", "Ⅷ": "VIII", "Ⅸ": "IX", "Ⅹ": "X",
        "Ⅺ": "XI", "Ⅻ": "XII", "Ⅼ": "L",
    }
    s = "".join(roman_chars.get(ch, ch) for ch in str(text or "").upper())
    if not s or not re.fullmatch(r"[IVXL]+", s):
        return None
    vals = {"I": 1, "V": 5, "X": 10, "L": 50}
    total = 0
    prev = 0
    for ch in reversed(s):
        val = vals[ch]
        if val < prev:
            total -= val
        else:
            total += val
            prev = val
    def to_roman(n: int) -> str:
        pairs = ((50, "L"), (40, "XL"), (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I"))
        out = []
        for val, sym in pairs:
            while n >= val:
                out.append(sym)
                n -= val
        return "".join(out)
    return total if 1 <= total <= 50 and to_roman(total) == s else None


def _alpha_number_value(text: str) -> int | None:
    s = str(text or "")
    if len(s) == 1 and s.isalpha() and s.isascii():
        return ord(s.upper()) - ord("A") + 1
    return None


def _korean_number_value(text: str) -> int | None:
    s = str(text or "")
    if not s:
        return None
    digits = {"일": 1, "이": 2, "삼": 3, "사": 4, "오": 5, "육": 6, "칠": 7, "팔": 8, "구": 9}
    if s in digits:
        return digits[s]
    if s == "십":
        return 10
    if s.startswith("십") and len(s) == 2 and s[1] in digits:
        return 10 + digits[s[1]]
    if len(s) in (2, 3) and s[0] in digits and s[1] == "십":
        val = digits[s[0]] * 10
        if len(s) == 3:
            if s[2] not in digits:
                return None
            val += digits[s[2]]
        return val if 1 <= val <= 50 else None
    return None


def normalize_lit_title_key(value) -> str:
    """문헌제목 엑셀의 차시 표기를 저장/조회용 키로 정규화."""
    s = str(value or "").strip()
    if not s:
        return ""
    s = _strip_lit_round_affixes(s)
    m = re.search(r"\d+", s)
    if m:
        return str(int(m.group(0)))
    for parser in (_circled_number_value, _roman_number_value, _alpha_number_value, _korean_number_value):
        val = parser(s)
        if val is not None:
            return str(val)
    return s


def load_lit_title_map_from_xlsx(path: str) -> dict:
    """첫 번째 시트의 A열(차시) / B열(문헌제목)을 읽어 dict로 반환."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out = {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        raw_key = row[0] if len(row) > 0 else None
        raw_title = row[1] if len(row) > 1 else None
        key_text = str(raw_key or "").strip()
        title = str(raw_title or "").strip()
        if ri == 1:
            hk = key_text.replace(" ", "")
            ht = title.replace(" ", "")
            if any(tok in hk for tok in ("차시", "번호", "순번")) and any(tok in ht for tok in ("제목", "문헌", "출처")):
                continue
        key = normalize_lit_title_key(key_text)
        if key and title:
            out[key] = title
    return out


def lit_title_phrase_for_rounds(rns) -> str:
    """문헌제목 모드에서 단일 문헌 차시에 해당하면 '「제목」에서' 반환."""
    if getattr(SETTINGS, "lit_intro_mode", getattr(SETTINGS, "field_intro_mode", "auto")) != "lit_title":
        return ""
    keys = []
    for rn in rns or []:
        s = str(rn or "")
        if "현지" in s or "합계" in s or "종합" in s:
            continue
        key = normalize_lit_title_key(s)
        if key and key not in keys:
            keys.append(key)
    if len(keys) != 1:
        return ""
    title = str((getattr(SETTINGS, "lit_title_map", {}) or {}).get(keys[0], "") or "").strip()
    return f"「{title}」에서" if title else ""


def missing_lit_title_keys(round_names) -> list[str]:
    """문헌제목 모드에서 필요한데 제목 매칭이 없는 차시 키 목록."""
    if getattr(SETTINGS, "lit_intro_mode", getattr(SETTINGS, "field_intro_mode", "auto")) != "lit_title":
        return []
    title_map = getattr(SETTINGS, "lit_title_map", {}) or {}
    needed = []
    for rn in round_names or []:
        s = str(rn or "")
        if "현지" in s or "합계" in s or "종합" in s:
            continue
        key = normalize_lit_title_key(s)
        if key and key not in needed:
            needed.append(key)
    return [k for k in needed if not str(title_map.get(k, "") or "").strip()]


def lit_result_prefix(rns, fallback_prefix: str, comma: bool = False) -> str:
    phrase = lit_title_phrase_for_rounds(rns)
    if phrase:
        return phrase
    return f"{fallback_prefix} 결과" + ("," if comma else "")


def lit_shi_prefix(rns, fallback_prefix: str) -> str:
    phrase = lit_title_phrase_for_rounds(rns)
    return phrase if phrase else f"{fallback_prefix} 시"


class SentenceSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent); self.setWindowTitle("문장 설정"); self.resize(620, 740); self.setMinimumSize(620, 740); self.setStyleSheet(_DIALOG_QSS)
        main_lay = QVBoxLayout(self)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        content_w = QWidget()
        lay = QVBoxLayout(content_w)

        form_title = QFormLayout(); self.e_title = QLineEdit(SETTINGS.sentence_title); self.e_title.setPlaceholderText("예: -")
        form_title.addRow("문장 앞 공통:", self.e_title); lay.addLayout(form_title)
        
        self.chk_period = QCheckBox("문장 끝 마침표(.) 추가")
        self.chk_period.setChecked(getattr(SETTINGS, "sent_add_period", False))
        lay.addWidget(self.chk_period)
        
        grp_lv = QGroupBox("분류군 레벨 표시"); lv_lay = QHBoxLayout(grp_lv)
        def _chk(label, checked): cb = QCheckBox(label); cb.setChecked(checked); return cb
        self.chk_phylum = _chk("문", SETTINGS.sent_show_phylum)
        self.chk_class  = _chk("강", SETTINGS.sent_show_class)
        self.chk_order  = _chk("목", SETTINGS.sent_show_order)
        self.chk_family = _chk("과", SETTINGS.sent_show_family)
        self.chk_genus  = _chk("속", SETTINGS.sent_show_genus)
        self.chk_species= _chk("종", SETTINGS.sent_show_species)
        for cb in (self.chk_phylum, self.chk_class, self.chk_order, self.chk_family, self.chk_genus, self.chk_species): lv_lay.addWidget(cb)
        lv_lay.addStretch(); lay.addWidget(grp_lv)
        grp_pl = QGroupBox("식물 하위 분류군 표시"); pl_lay = QHBoxLayout(grp_pl)
        self.chk_var   = _chk("변종", SETTINGS.sent_show_var); self.chk_forma = _chk("품종", SETTINGS.sent_show_forma); self.chk_subsp = _chk("아종", SETTINGS.sent_show_subsp)
        self.chk_taxa  = _chk("총분류군수", getattr(SETTINGS, "sent_show_taxa", True))
        for cb in (self.chk_var, self.chk_forma, self.chk_subsp, self.chk_taxa): pl_lay.addWidget(cb)
        pl_lay.addStretch(); lay.addWidget(grp_pl)

        grp_prot = QGroupBox("법정보호종 등급 표기")
        prot_lay = QHBoxLayout(grp_prot)
        self.chk_prot_none = _chk("등급 숨김", False)
        self.chk_prot_short = _chk("축약(천/멸)", False)
        self.chk_prot_full = _chk("전체표현", False)
        self.prot_mode_group = QButtonGroup(self)
        self.prot_mode_group.setExclusive(True)
        self.prot_mode_group.addButton(self.chk_prot_none)
        self.prot_mode_group.addButton(self.chk_prot_short)
        self.prot_mode_group.addButton(self.chk_prot_full)
        cur_mode = getattr(SETTINGS, "prot_grade_mode", "short")
        if cur_mode == "none":
            self.chk_prot_none.setChecked(True)
        elif cur_mode == "full":
            self.chk_prot_full.setChecked(True)
        else:
            self.chk_prot_short.setChecked(True)
        for cb in (self.chk_prot_none, self.chk_prot_short, self.chk_prot_full):
            prot_lay.addWidget(cb)
        prot_lay.addStretch()
        lay.addWidget(grp_prot)

        grp_style = QGroupBox("차시 표현")
        style_form = QFormLayout(grp_style)
        self.cb_field_intro = QComboBox()
        self.cb_field_intro.addItem("차시 수 자동 표기", "auto")
        self.cb_field_intro.addItem("'조사 결과'로 고정", "fixed")
        cur_intro = getattr(SETTINGS, "field_intro_mode", "auto")
        self.cb_field_intro.setCurrentIndex({"auto": 0, "fixed": 1}.get(cur_intro, 0))
        self.cb_lit_intro = QComboBox()
        self.cb_lit_intro.addItem("차시 수 자동 표기", "auto")
        self.cb_lit_intro.addItem("'조사 결과'로 고정", "fixed")
        self.cb_lit_intro.addItem("문헌제목 사용", "lit_title")
        cur_lit_intro = getattr(SETTINGS, "lit_intro_mode", getattr(SETTINGS, "field_intro_mode", "auto"))
        self.cb_lit_intro.setCurrentIndex({"auto": 0, "fixed": 1, "lit_title": 2}.get(cur_lit_intro, 0))

        style_form.addRow("현지조사 차시 표현:", self.cb_field_intro)
        style_form.addRow("문헌조사 차시 표현:", self.cb_lit_intro)
        self._lit_title_map = dict(getattr(SETTINGS, "lit_title_map", {}) or {})
        self._lit_title_source_path = getattr(SETTINGS, "lit_title_source_path", "") or ""
        btn_lit_titles = QPushButton("불러오기")
        btn_lit_titles.setFixedHeight(24)
        btn_lit_titles.setMaximumWidth(78)
        btn_lit_titles.setStyleSheet(
            f"QPushButton{{background:#FFFFFF;color:{_ACCENT_D};border:1px solid {_BORDER};"
            f"border-radius:5px;{FF_KR};font-size:11px;font-weight:700;padding:2px 8px;}}"
            f"QPushButton:hover{{background:{_ACCENT_L};}}"
        )
        btn_lit_titles.clicked.connect(self._load_lit_titles)
        self.lbl_lit_titles = QLabel("")
        self.lbl_lit_titles.setStyleSheet(f"{FF_KR}; font-size:11px; color:{_SUB};")
        self._refresh_lit_title_label()
        lit_row = QHBoxLayout()
        lit_row.addWidget(btn_lit_titles)
        lit_row.addWidget(self.lbl_lit_titles, 1)
        style_form.addRow("문헌제목:", lit_row)
        lay.addWidget(grp_style)

        # ── 일괄 종결 어미 설정 ──
        grp_batch = QGroupBox("일괄 종결 어미 변경")
        batch_lay = QHBoxLayout(grp_batch)
        self.rb_m = QRadioButton("~ㅁ (예: 됨, 함, 남)")
        self.rb_eum = QRadioButton("~ㅆ음 (예: 되었음, 하였음, 났음)")
        self.rb_da = QRadioButton("~ㅆ다 (예: 되었다, 하였다, 났다)")
        
        self.bg_ending = QButtonGroup(self)
        self.bg_ending.addButton(self.rb_m)
        self.bg_ending.addButton(self.rb_eum)
        self.bg_ending.addButton(self.rb_da)
        
        batch_lay.addWidget(self.rb_m)
        batch_lay.addWidget(self.rb_eum)
        batch_lay.addWidget(self.rb_da)
        batch_lay.addStretch()
        lay.addWidget(grp_batch)
        
        self.rb_m.clicked.connect(lambda: self._apply_ending_style('m'))
        self.rb_eum.clicked.connect(lambda: self._apply_ending_style('eum'))
        self.rb_da.clicked.connect(lambda: self._apply_ending_style('da'))

        # ── 어미 설정 ──
        def le(val): return QLineEdit(val)
       
        grp_end = QGroupBox("기본 종결 어미 (조사/분석 결과)")
        form_end = QFormLayout(grp_end)
        self.e_end_field = le(SETTINGS.end_field)
        self.e_end_lit   = le(SETTINGS.end_lit)
        self.e_end_ana   = le(SETTINGS.end_ana)
        self.e_end_prb   = le(SETTINGS.end_prb)
        form_end.addRow("현지조사 종결 (예: 확인됨, 조사됨):", self.e_end_field)
        form_end.addRow("문헌조사 종결 (예: 기록됨, 문헌됨):", self.e_end_lit)
        form_end.addRow("분석결과 종결 (예: 분석됨, 산출됨):", self.e_end_ana)
        form_end.addRow("탐문조사 종결 (예: 전언됨, 청취됨):", self.e_end_prb)

        lay.addWidget(grp_end)
        lay.addStretch()
        scroll.setWidget(content_w)
        main_lay.addWidget(scroll)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_reset = QPushButton("기본값")
        btns.addButton(btn_reset, QDialogButtonBox.ResetRole)
        btn_reset.clicked.connect(self._reset_defaults)
        btn_save_default = QPushButton("문장 기본값 저장")
        btns.addButton(btn_save_default, QDialogButtonBox.ActionRole)
        btn_save_default.clicked.connect(self._save_sentence_defaults)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        main_lay.addWidget(btns)

    def _refresh_lit_title_label(self):
        count = len(getattr(self, "_lit_title_map", {}) or {})
        path = getattr(self, "_lit_title_source_path", "") or ""
        name = os.path.basename(path) if path else ""
        self.lbl_lit_titles.setText(f"{count}개 불러옴" + (f" · {name}" if name else ""))

    def _load_lit_titles(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "문헌제목 엑셀 선택",
            getattr(self, "_lit_title_source_path", "") or "",
            "Excel Files (*.xlsx *.xlsm)"
        )
        if not path:
            return
        try:
            title_map = load_lit_title_map_from_xlsx(path)
        except Exception as e:
            QMessageBox.critical(self, "문헌제목 불러오기 실패", f"문헌제목 엑셀을 읽을 수 없습니다.\n{e}")
            return
        self._lit_title_map = title_map
        self._lit_title_source_path = path
        self._refresh_lit_title_label()
        QMessageBox.information(self, "문헌제목 불러오기", f"문헌제목 {len(title_map)}개를 불러왔습니다.")

    def _needed_lit_round_names(self):
        parent = self.parent()
        parsed = getattr(parent, "_last_parsed", {}) if parent is not None else {}
        out = []
        seen = set()
        for obj in (parsed or {}).values():
            meta = getattr(obj, "meta", None)
            if not meta:
                continue
            for rn in getattr(meta, "round_names", []) or []:
                s = str(rn or "")
                if "문헌" in s and "합계" not in s and "종합" not in s and s not in seen:
                    seen.add(s)
                    out.append(s)
            for rn in getattr(meta, "field_cols", {}) or {}:
                s = str(rn or "")
                if "문헌" in s and "합계" not in s and "종합" not in s and s not in seen:
                    seen.add(s)
                    out.append(s)
        return out

    def _warn_missing_lit_titles(self):
        if self.cb_lit_intro.currentData() != "lit_title":
            return
        old_map = getattr(SETTINGS, "lit_title_map", {}) or {}
        SETTINGS.lit_title_map = dict(getattr(self, "_lit_title_map", {}) or {})
        missing = missing_lit_title_keys(self._needed_lit_round_names())
        SETTINGS.lit_title_map = old_map
        if missing:
            QMessageBox.warning(
                self,
                "문헌제목 누락",
                "다음 문헌 차시에 대응하는 제목이 없습니다.\n"
                + ", ".join(missing)
                + "\n\n누락 차시는 기존 차시 표현으로 생성됩니다."
            )

    def _reset_defaults(self):
        from config import AppSettings
        def_s = AppSettings()
        
        self.e_title.setText(def_s.sentence_title)
        self.chk_period.setChecked(getattr(def_s, "sent_add_period", False))
        self.chk_phylum.setChecked(def_s.sent_show_phylum)
        self.chk_class.setChecked(def_s.sent_show_class)
        self.chk_order.setChecked(def_s.sent_show_order)
        self.chk_family.setChecked(def_s.sent_show_family)
        self.chk_genus.setChecked(def_s.sent_show_genus)
        self.chk_species.setChecked(def_s.sent_show_species)
        self.chk_var.setChecked(def_s.sent_show_var)
        self.chk_forma.setChecked(def_s.sent_show_forma)
        self.chk_subsp.setChecked(def_s.sent_show_subsp)
        self.chk_taxa.setChecked(getattr(def_s, "sent_show_taxa", True))
        
        self.chk_prot_short.setChecked(True)
        self.cb_field_intro.setCurrentIndex({"auto": 0, "fixed": 1}.get(getattr(def_s, "field_intro_mode", "auto"), 0))
        self.cb_lit_intro.setCurrentIndex({"auto": 0, "fixed": 1, "lit_title": 2}.get(getattr(def_s, "lit_intro_mode", "auto"), 0))
        self._lit_title_map = dict(getattr(def_s, "lit_title_map", {}) or {})
        self._lit_title_source_path = getattr(def_s, "lit_title_source_path", "") or ""
        self._refresh_lit_title_label()
        
        self.e_end_field.setText(def_s.end_field)
        self.e_end_lit.setText(def_s.end_lit)
        self.e_end_ana.setText(def_s.end_ana)
        self.e_end_prb.setText(def_s.end_prb)
        
        self.rb_m.setAutoExclusive(False); self.rb_eum.setAutoExclusive(False); self.rb_da.setAutoExclusive(False)
        self.rb_m.setChecked(False); self.rb_eum.setChecked(False); self.rb_da.setChecked(False)
        self.rb_m.setAutoExclusive(True); self.rb_eum.setAutoExclusive(True); self.rb_da.setAutoExclusive(True)

    def _apply_ending_style(self, style):
        import re
        def _convert(text):
            text = re.sub(r'됨|되었음|되었다', {'m':'됨', 'eum':'되었음', 'da':'되었다'}[style], text)
            text = re.sub(r'함|하였음|하였다|했음|했다', {'m':'함', 'eum':'하였음', 'da':'하였다'}[style], text)
            text = re.sub(r'남|났음|났다', {'m':'남', 'eum':'났음', 'da':'났다'}[style], text)
            text = re.sub(r'않음|않았음|않았다', {'m':'않음', 'eum':'않았음', 'da':'않았다'}[style], text)
            text = re.sub(r'없음|없었음|없었다', {'m':'없음', 'eum':'없었음', 'da':'없었다'}[style], text)
            return text
        
        for edit in [self.e_end_field, self.e_end_lit, self.e_end_ana, self.e_end_prb]:
            edit.setText(_convert(edit.text()))

    def apply(self):
        SETTINGS.sentence_title = self.e_title.text(); SETTINGS.sent_show_phylum = self.chk_phylum.isChecked(); SETTINGS.sent_show_class = self.chk_class.isChecked(); SETTINGS.sent_show_order = self.chk_order.isChecked(); SETTINGS.sent_show_family = self.chk_family.isChecked(); SETTINGS.sent_show_genus = self.chk_genus.isChecked(); SETTINGS.sent_show_species = self.chk_species.isChecked(); SETTINGS.sent_show_var = self.chk_var.isChecked(); SETTINGS.sent_show_forma = self.chk_forma.isChecked(); SETTINGS.sent_show_subsp = self.chk_subsp.isChecked(); SETTINGS.sent_show_taxa = self.chk_taxa.isChecked(); SETTINGS.sent_add_period = self.chk_period.isChecked()

        SETTINGS.end_field = self.e_end_field.text()
        SETTINGS.end_lit   = self.e_end_lit.text()
        SETTINGS.end_ana   = self.e_end_ana.text()
        SETTINGS.end_prb   = self.e_end_prb.text()
        if self.chk_prot_none.isChecked():
            SETTINGS.prot_grade_mode = "none"
        elif self.chk_prot_full.isChecked():
            SETTINGS.prot_grade_mode = "full"
        else:
            SETTINGS.prot_grade_mode = "short"
        SETTINGS.field_intro_mode = self.cb_field_intro.currentData()
        SETTINGS.lit_intro_mode = self.cb_lit_intro.currentData()
        SETTINGS.lit_title_map = dict(getattr(self, "_lit_title_map", {}) or {})
        SETTINGS.lit_title_source_path = getattr(self, "_lit_title_source_path", "") or ""
        self._warn_missing_lit_titles()

    def _save_sentence_defaults(self):
        self.apply()
        parent = self.parent()
        if parent is not None and hasattr(parent, "_save_sentence_program_defaults"):
            if parent._save_sentence_program_defaults():
                QMessageBox.information(self, "문장 기본값 저장", "문장 기본값을 저장했습니다.")
            return
        try:
            qs = QSettings("susippi", "taxa_analyzer")
            graph_defaults = {}
            try:
                old = json.loads(qs.value("program_defaults/v1", "", str) or "{}")
                if isinstance(old, dict) and isinstance(old.get("graph_defaults"), dict):
                    graph_defaults = old.get("graph_defaults")
            except Exception:
                graph_defaults = {}
            payload = {
                "version": 1,
                "settings": SETTINGS.to_dict(),
                "graph_defaults": graph_defaults,
            }
            qs.setValue("program_defaults/v1", json.dumps(payload, ensure_ascii=False))
            qs.sync()
            QMessageBox.information(self, "문장 기본값 저장", "문장 기본값을 저장했습니다.")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"문장 기본값 저장 중 오류가 발생했습니다.\n{e}")
def _normalized_percentages(values, decimals=None):
    """
    각 값을 개별적으로 반올림한 비율(%)을 반환한다.
    합산 오차가 있을 수 있으며, 자동 보정은 하지 않는다.
    보정 추천은 suggest_ratio_fix()를 사용한다.
    """
    if decimals is None: decimals = SETTINGS.decimal
    vals = [float(v or 0) for v in values]
    total = sum(vals)
    if total <= 0: return [0.0 for _ in vals]
    return [round(v / total * 100, decimals) for v in vals]


def _apply_ratio_correction(pct_values, raw_values=None, decimals=None):
    """
    비율 리스트에 자동 보정을 적용한 뒤 (보정된_리스트, 변경_기록) 반환.
    최대잔여법과 유사하게 원본 값(raw_values)과 반올림된 비율 값의 차이의 절대값이 큰 순서대로 보정값을 개별 적용.
    """
    if decimals is None: decimals = SETTINGS.decimal
    step = round(10 ** (-decimals), decimals)

    total = round(sum(pct_values), decimals)
    diff  = round(100.0 - total, decimals)

    if abs(diff) < step * 0.5:
        return [round(v, decimals) for v in pct_values], []

    needed_steps = int(round(abs(diff) / step))
    if needed_steps == 0:
        return [round(v, decimals) for v in pct_values], []

    items = []
    if raw_values and len(raw_values) == len(pct_values):
        total_raw = sum(float(v) for v in raw_values)
        for i, (rv, pv) in enumerate(zip(raw_values, pct_values)):
            raw_pct = (float(rv) / total_raw * 100) if total_raw else 0
            rounded_pv = round(pv, decimals)
            residual = raw_pct - rounded_pv
            items.append((i, rounded_pv, residual))
    else:
        for i, pv in enumerate(pct_values):
            rounded_pv = round(pv, decimals)
            residual = pv - int(pv)
            items.append((i, rounded_pv, residual))

    # 그룹핑 기준: (반올림된 표시값, 오차의 절대값) 이 완전히 같은 항목들
    # 소수점 오차로 인한 문제를 막기 위해 오차 절대값은 6자리로 round
    import collections
    groups = collections.defaultdict(list)
    for i, rounded_pv, residual in items:
        group_key = (rounded_pv, round(abs(residual), 6))
        groups[group_key].append(i)

    # 절대값 오차가 큰 순으로 정렬 (1순위: 오차 절대값, 2순위: 반올림값 큰 순)
    sorted_groups = sorted(groups.items(), key=lambda x: (x[0][1], x[0][0]), reverse=True)

    out = [round(v, decimals) for v in pct_values]
    applied = []
    
    remaining_steps = needed_steps
    for (pv, abs_res), indices in sorted_groups:
        if remaining_steps <= 0:
            break
            
        group_size = len(indices)
        if group_size <= remaining_steps:
            for idx in indices:
                out[idx] = round(out[idx] + (step if diff > 0 else -step), decimals)
            remaining_steps -= group_size
            
            # 기록을 위해 계산 (적용된 그룹)
            old_val = pv
            new_val = round(pv + (step if diff > 0 else -step), decimals)
            applied.append({"from": old_val, "to": new_val, "count": group_size})
        else:
            # 형평성 문제: 보정해야 할 step이 항목 수보다 적으므로 적용 불가 (Skip)
            continue

    return out, applied


def format_ratio_hint(pcts, raw_values=None, decimals=None):
    """
    보정 전 비율 리스트를 받아 자동 보정을 수행하고,
    변경 내역을 문자열로 반환한다.
    합계가 이미 100.00이면 빈 문자열 반환.
    """
    if decimals is None: decimals = SETTINGS.decimal
    total = round(sum(pcts), decimals)
    tol   = 10 ** (-decimals) * 0.5

    if abs(100.0 - total) < tol:
        return ""

    _, changes = _apply_ratio_correction(pcts, raw_values, decimals)
    if not changes:
        return ""

    lines = ["보정 완료  (원래 합계: {:.{}f})".format(total, decimals)]
    for c in changes:
        lines.append("  변경: {:.{}f} → {:.{}f} ({}개)".format(
            c["from"], decimals, c["to"], decimals, c["count"]))

    return "\n".join(lines)

def normalize_tbl_pct_col(tbl: "QTableWidget", pct_col: int,
                          cnt_col: int | None = None,
                          pct_formatter=None, decimals: int | None = None):
    """테이블의 비율(%) 컬럼을 합산 100이 되도록 in-place 보정.

    Parameters
    ----------
    tbl        : 대상 QTableWidget
    pct_col    : 비율(%) 셀이 있는 열 인덱스
    cnt_col    : 종수/개체수가 있는 열 인덱스 (None 이면 기존 % 텍스트에서 파싱)
    pct_formatter : float → str 변환기 (None 이면 SETTINGS.decimal 사용)
    decimals   : 소수점 자릿수 override (None 이면 SETTINGS.decimal)
    """
    if decimals is None: decimals = SETTINGS.decimal
    if pct_formatter is None:
        fmt = lambda v: f"{v:.{decimals}f}%"
    else:
        fmt = pct_formatter

    n_data = tbl.rowCount() - 1  # 마지막 합계 행 제외
    if n_data <= 0: return

    # 분모가 될 값 수집
    if cnt_col is not None:
        raw_vals = []
        for r in range(n_data):
            it = tbl.item(r, cnt_col)
            try: raw_vals.append(float((it.text() if it else "0").replace(",", "")))
            except ValueError: raw_vals.append(0.0)
    else:
        raw_vals = []
        for r in range(n_data):
            it = tbl.item(r, pct_col)
            txt = (it.text() if it else "0").replace("%", "").replace(",", "").strip()
            try: raw_vals.append(float(txt))
            except ValueError: raw_vals.append(0.0)

    pcts = _normalized_percentages(raw_vals, decimals=decimals)

    # 자동 보정 적용 (원본 값을 넘겨서 최대잔여법 적용)
    hint = format_ratio_hint(pcts, raw_vals, decimals)
    corrected, _ = _apply_ratio_correction(pcts, raw_vals, decimals)

    for r, pct in enumerate(corrected):
        it = tbl.item(r, pct_col)
        if it: it.setText(fmt(pct))

    total_it = tbl.item(n_data, pct_col)
    if total_it:
        s = round(sum(corrected), decimals)
        total_it.setText(fmt(s))
        tol = 10 ** (-decimals) / 2
        if not math.isclose(s, 100.0, rel_tol=0.0, abs_tol=tol):
            total_it.setForeground(QColor(_ERR))
        else:
            total_it.setForeground(QColor(_TXT))

    return hint


def normalize_tbl_pct_row(tbl: "QTableWidget", pct_row: int,
                          cnt_row: int | None = None,
                          pct_formatter=None, decimals: int | None = None,
                          start_col: int = 1, end_col_offset: int = 1):
    """테이블의 비율(%) 행을 합산 100이 되도록 in-place 보정.
    
    Parameters
    ----------
    tbl        : 대상 QTableWidget
    pct_row    : 비율(%) 셀이 있는 행 인덱스
    cnt_row    : 종수/개체수가 있는 행 인덱스 (None 이면 기존 % 텍스트에서 파싱)
    start_col  : 데이터가 시작하는 열 인덱스 (기본값 1)
    end_col_offset : 마지막 열(합계 등) 제외할 개수 (기본값 1)
    """
    if decimals is None: decimals = SETTINGS.decimal
    if pct_formatter is None:
        fmt = lambda v: f"{v:.{decimals}f}%"
    else:
        fmt = pct_formatter

    n_data = tbl.columnCount() - start_col - end_col_offset
    if n_data <= 0: return

    if cnt_row is not None:
        raw_vals = []
        for c in range(n_data):
            it = tbl.item(cnt_row, start_col + c)
            try: raw_vals.append(float((it.text() if it else "0").replace(",", "")))
            except ValueError: raw_vals.append(0.0)
    else:
        raw_vals = []
        for c in range(n_data):
            it = tbl.item(pct_row, start_col + c)
            txt = (it.text() if it else "0").replace("%", "").replace(",", "").strip()
            try: raw_vals.append(float(txt))
            except ValueError: raw_vals.append(0.0)

    pcts = _normalized_percentages(raw_vals, decimals=decimals)

    # 자동 보정 적용 (원본 값을 넘겨서 최대잔여법 적용)
    hint = format_ratio_hint(pcts, raw_vals, decimals)
    corrected, _ = _apply_ratio_correction(pcts, raw_vals, decimals)

    for c, pct in enumerate(corrected):
        it = tbl.item(pct_row, start_col + c)
        if it: it.setText(fmt(pct))

    total_it = tbl.item(pct_row, start_col + n_data)
    if total_it:
        s = round(sum(corrected), decimals)
        total_it.setText(fmt(s))
        tol = 10 ** (-decimals) / 2
        if not math.isclose(s, 100.0, rel_tol=0.0, abs_tol=tol):
            total_it.setForeground(QColor(_ERR))
        else:
            total_it.setForeground(QColor(_TXT))

    return hint


def _korean_font():
    if not _MPL: return None
    cands = ["Malgun Gothic","맑은 고딕","NanumGothic","AppleGothic","Gulim"]
    avail = {f.name for f in fm.fontManager.ttflist}
    for c in cands:
        if c in avail: return c
    return None

def _apply_mpl():
    font = _korean_font()
    if font: plt.rcParams["font.family"] = font
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["font.size"] = 9
    plt.rcParams["font.weight"] = "bold"

def _get_nice_bounds(max_val: float) -> tuple[float, float]:
    if max_val <= 0: return 1.0, 0.2
    mag = 10.0 ** math.floor(math.log10(max_val))
    norm = max_val / mag
    if norm <= 1.5: step_norm = 0.2
    elif norm <= 3.0: step_norm = 0.5
    elif norm <= 5.0: step_norm = 1.0
    elif norm <= 7.5: step_norm = 2.0
    else: step_norm = 2.0
    step = step_norm * mag
    new_max = math.ceil(max_val / step) * step
    if (new_max - max_val) / (new_max if new_max > 0 else 1) < 0.12:
        new_max += step
    return new_max, step

class _Canvas(FigureCanvas if _MPL else QLabel):
    def __init__(self, w=5, h=4):
        if not _MPL:
            super().__init__("matplotlib 미설치"); self.setAlignment(Qt.AlignCenter); return
        self._fig = plt.Figure(figsize=(w, h))
        super().__init__(self._fig)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._ctx)
        self._refresh_cb = None
        self._setting_mode = "bar"
        self._cfg = {
            "dom_pct": SETTINGS.dom_pct, "bar_horiz": SETTINGS.bar_horiz,
            "bar_fontsize": SETTINGS.bar_fontsize, "bar_color": SETTINGS.bar_color, "line_color": SETTINGS.line_color,
            "grid_on": SETTINGS.grid_on, "grid_color": SETTINGS.grid_color, "x_label_rot": SETTINGS.x_label_rot,
            "marker_size": SETTINGS.marker_size, "line_width": SETTINGS.line_width, "bar_h_height": SETTINGS.bar_h_height,
            "bar_v_width": SETTINGS.bar_v_width, "div_bar_width": SETTINGS.div_bar_width, "div_bar_gap": SETTINGS.div_bar_gap,
            "bar_count_label_inside": SETTINGS.bar_count_label_inside, "pie_fontsize": SETTINGS.pie_fontsize,
            "pie_start_angle": SETTINGS.pie_start_angle, "pie_label_offset": SETTINGS.pie_label_offset, "pie_edge_width": SETTINGS.pie_edge_width,
            "pie_leader_width": SETTINGS.pie_leader_width, "pie_leader_gap": SETTINGS.pie_leader_gap, "div_color_1": SETTINGS.div_color_1, "div_color_2": SETTINGS.div_color_2,
            "div_color_3": SETTINGS.div_color_3, "axis_min": SETTINGS.axis_min, "axis_max": SETTINGS.axis_max, "axis_step": SETTINGS.axis_step,
            "graph_width_scale": 1.0, "graph_height_scale": 1.0,
        }
        self._bar_resize = False
        self._bar_start_y = 0
        self._bar_start_h = 0.0
        self._bar_start_w = 0.0
        self._bar_start_div_w = 0.0
        self._legend_drag = None
        self.mpl_connect("button_press_event", self._on_bar_press)
        self.mpl_connect("motion_notify_event", self._on_bar_move)
        self.mpl_connect("button_release_event", self._on_bar_release)

    def _on_bar_press(self, event):
        if getattr(self, "_setting_mode", "bar") == "pie": return
        if event.button != 1 or event.y is None: return
        legend_info = getattr(self, "_active_combo_legend_info", None)
        if legend_info is not None:
            try:
                legend, legend_ax, anchor_key = legend_info
                renderer = self.figure.canvas.get_renderer()
                bbox = legend.get_window_extent(renderer=renderer).padded(10)
                contains, _ = legend.contains(event)
                if bbox.contains(event.x, event.y) or contains:
                    self._legend_drag = (legend, legend_ax, anchor_key)
                    self.setCursor(Qt.SizeAllCursor)
                    return
            except Exception:
                pass
        if event.inaxes is None: return
        mode = getattr(self, "_setting_mode", "bar")
        if mode not in ("bird_combo", "diversity"): return
        self._bar_resize = True
        self._bar_start_y = event.y
        self._bar_start_h = float(self._cfg.get("bar_h_height", SETTINGS.bar_h_height))
        self._bar_start_w = float(self._cfg.get("bar_v_width", SETTINGS.bar_v_width))
        self._bar_start_div_w = float(self._cfg.get("div_bar_width", SETTINGS.div_bar_width))
        self.setCursor(Qt.SizeVerCursor)

    def _on_bar_move(self, event):
        if self._legend_drag is not None and event.x is not None and event.y is not None:
            legend, legend_ax, anchor_key = self._legend_drag
            try:
                x_ax, y_ax = legend_ax.transAxes.inverted().transform((event.x, event.y))
                anchor = (
                    max(-1.0, min(2.0, float(x_ax))),
                    max(-1.0, min(2.0, float(y_ax))),
                )
                legend.set_bbox_to_anchor(anchor, transform=legend_ax.transAxes)
                setattr(SETTINGS, anchor_key, anchor)
                self._cfg[anchor_key] = anchor
                self.figure.canvas.draw_idle()
            except Exception:
                pass
            return
        if not self._bar_resize or event.y is None: return
        dy = self._bar_start_y - event.y
        delta = dy * 0.002
        mode = getattr(self, "_setting_mode", "bar")
        if mode == "bird_combo":
            if self._cfg.get("bar_horiz", SETTINGS.bar_horiz):
                self._cfg["bar_h_height"] = max(0.20, min(1.00, self._bar_start_h + delta))
            else:
                self._cfg["bar_v_width"] = max(0.20, min(0.95, self._bar_start_w + delta))
        elif mode == "diversity":
            self._cfg["div_bar_width"] = max(0.20, min(0.90, self._bar_start_div_w + delta))
        cb = getattr(self, "_bar_changed_cb", None)
        if callable(cb):
            try:
                cb()
            except Exception:
                pass
        if self._refresh_cb: self._refresh_cb()

    def _on_bar_release(self, event):
        if self._legend_drag is not None:
            self._legend_drag = None
            self.setCursor(Qt.ArrowCursor)
            return
        if self._bar_resize:
            self._bar_resize = False
            self.setCursor(Qt.ArrowCursor)

    def draw(self):
        """위젯이 실제로 화면에 표시될 때만 렌더링 — 로딩 중 숨겨진 상태에서 발생하는 겹침 방지."""
        if not _MPL: return
        if not self.isVisible():
            return
        super().draw()

    def showEvent(self, event):
        super().showEvent(event)
        if not _MPL: return
        if not getattr(self, "_shown_once", False):
            self._shown_once = True
            from PySide6.QtCore import QTimer
            QTimer.singleShot(50, lambda: (self._refresh_cb() if self._refresh_cb else super(_Canvas, self).draw()))

    def _ctx(self, pos):
        m = QMenu(self)
        m.addAction("📋 클립보드 복사").triggered.connect(self._copy)
        m.exec(self.mapToGlobal(pos))

    def _crop_image_to_content(self, img: QImage, pad: int = 4) -> QImage:
        """복사 직전 흰 배경을 잘라 실제 그래프 요소의 최외곽만 남긴다."""
        if img.isNull():
            return img
        src = img.convertToFormat(QImage.Format_RGBA8888)
        w, h = src.width(), src.height()
        if w <= 0 or h <= 0:
            return img
        try:
            data = bytes(src.constBits())
        except Exception:
            return img
        bpl = src.bytesPerLine()
        threshold = 248
        left, top, right, bottom = w, h, -1, -1
        for y in range(h):
            row = y * bpl
            for x in range(w):
                i = row + x * 4
                r, g, b, a = data[i], data[i + 1], data[i + 2], data[i + 3]
                if a > 8 and (r < threshold or g < threshold or b < threshold):
                    if x < left: left = x
                    if x > right: right = x
                    if y < top: top = y
                    if y > bottom: bottom = y
        if right < left or bottom < top:
            return img
        left = max(0, left - pad)
        top = max(0, top - pad)
        right = min(w - 1, right + pad)
        bottom = min(h - 1, bottom + pad)
        return src.copy(QRect(left, top, right - left + 1, bottom - top + 1))

    def _copy(self):
        buf = io.BytesIO()
        self.figure.savefig(buf, format="png", dpi=150, bbox_inches="tight", pad_inches=0.03)
        buf.seek(0)
        img = QImage(); img.loadFromData(buf.read(), "PNG")
        img = self._crop_image_to_content(img, pad=4)
        QApplication.clipboard().setImage(img)

class DragPctSlider(QWidget):
    valueChanged = Signal(float)
    def __init__(self, lo=0.10, hi=1.00, value=0.60, step=0.01, parent=None):
        super().__init__(parent)
        self._lo = lo; self._hi = hi; self._step = step
        lay = QHBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(6)
        self._sld = QSlider(Qt.Horizontal)
        ticks = round((hi - lo) / step)
        self._sld.setRange(0, ticks)
        self._sld.setValue(round((value - lo) / step))
        self._sld.setFixedHeight(22)
        self._lbl = QLabel(f"{value*100:.0f}%")
        self._lbl.setFixedWidth(38)
        self._lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lay.addWidget(self._sld, 1); lay.addWidget(self._lbl)
        self._sld.valueChanged.connect(self._on_change)
    def _on_change(self, tick):
        v = round(self._lo + tick * self._step, 4)
        self._lbl.setText(f"{v*100:.0f}%")
        self.valueChanged.emit(v)
    def value(self) -> float:
        tick = self._sld.value()
        return round(self._lo + tick * self._step, 4)
    def setValue(self, v: float):
        tick = round((v - self._lo) / self._step)
        self._sld.setValue(max(0, min(self._sld.maximum(), tick)))

class GraphValueSlider(QWidget):
    valueChanged = Signal(float)
    def __init__(self, lo=0.0, hi=1.0, value=1.0, step=0.05, decimals=2, suffix="", parent=None):
        super().__init__(parent)
        self._lo = float(lo); self._hi = float(hi); self._step = float(step)
        self._decimals = int(decimals); self._suffix = suffix
        lay = QHBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(6)
        self._sld = QSlider(Qt.Horizontal)
        self._sld.setRange(0, max(1, round((self._hi - self._lo) / self._step)))
        self._lbl = QLabel()
        self._lbl.setFixedWidth(32)
        self._lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lay.addWidget(self._sld, 1); lay.addWidget(self._lbl)
        self._sld.valueChanged.connect(self._on_change)
        self.setValue(value)
    def _text(self):
        return f"{self.value():.{self._decimals}f}{self._suffix}"
    def _on_change(self, _tick):
        self._lbl.setText(self._text())
        self.valueChanged.emit(self.value())
    def value(self) -> float:
        return round(self._lo + self._sld.value() * self._step, self._decimals)
    def setValue(self, v: float):
        try: val = float(v)
        except Exception: val = self._lo
        tick = round((max(self._lo, min(self._hi, val)) - self._lo) / self._step)
        self._sld.setValue(max(0, min(self._sld.maximum(), tick)))
        self._lbl.setText(self._text())

class PieCanvas(_Canvas):
    def __init__(self):
        super().__init__(5, 5)
        if not _MPL: return
        self._txt_objs  = []
        self._tip_xy    = []
        self._drag      = None
        self._drag_orig = None
        self._drag_start = (0, 0)
        self._resize = False
        self._resize_start_y = 0
        self._resize_start_r = SETTINGS.pie_radius
        self._positions = {}
        self._data      = {}
        self._radius    = SETTINGS.pie_radius
        self._setting_mode = "pie"
        self._refresh_cb = self._redraw
        self.mpl_connect("button_press_event",   self._on_press)
        self.mpl_connect("motion_notify_event",  self._on_move)
        self.mpl_connect("button_release_event", self._on_release)

    def draw_pie(self, data: dict):
        self._data = dict(data)
        self._redraw()

    def _redraw(self):
        data = self._data
        if not _MPL or not hasattr(self, "figure") or not data: return
        _apply_mpl()
        fig = self.figure; fig.clf()
        clean  = {k: v for k, v in data.items() if isinstance(v, (int, float)) and v > 0 and v == v}
        if not clean: return
        total  = sum(clean.values())
        labels = list(clean.keys())
        values = list(clean.values())
        colors = _pie_palette(len(labels))
        fs     = self._cfg.get("pie_fontsize", SETTINGS.pie_fontsize)
        decimal = self._cfg.get("decimal", SETTINGS.decimal)
        start_angle = self._cfg.get("pie_start_angle", SETTINGS.pie_start_angle)
        label_offset = self._cfg.get("pie_label_offset", SETTINGS.pie_label_offset)
        edge_width = self._cfg.get("pie_edge_width", SETTINGS.pie_edge_width)
        ax = fig.add_axes([0.05, 0.05, 0.9, 0.9])
        r = self._radius
        wedges, _ = ax.pie(values, labels=None, autopct=None, colors=colors, startangle=start_angle, counterclock=False, radius=r, wedgeprops={"linewidth": edge_width, "edgecolor": "white"})
        ax.set_xlim(-2.2, 2.2); ax.set_ylim(-2.2, 2.2); ax.set_aspect("equal", adjustable="box")
        if getattr(self, "_show_leader_gap_guide", False):
            leader_gap = self._cfg.get("pie_leader_gap", getattr(SETTINGS, "pie_leader_gap", max(0.18, SETTINGS.pie_label_offset * 0.7)))
            guide = plt.Circle(
                (0, 0),
                r + float(leader_gap),
                fill=False,
                linestyle=(0, (4, 3)),
                linewidth=1.2,
                edgecolor="#EA580C",
                alpha=0.8,
                zorder=3,
            )
            ax.add_patch(guide)
        self._txt_objs = []; self._tip_xy = []
        for w, lbl, val in zip(wedges, labels, values):
            ang  = math.radians((w.theta1 + w.theta2) / 2)
            pct  = val  # val is already the formatted percentage extracted from the table
            text = f"{lbl}\n{pct:.{decimal}f}%"
            if text in self._positions: tx, ty = self._positions[text]
            else: tx = math.cos(ang) * (r + label_offset); ty = math.sin(ang) * (r + label_offset)
            ha = "center"
            t  = ax.text(tx, ty, text, ha=ha, va="center", fontsize=fs, fontweight="bold", color="black", zorder=5, clip_on=False)
            self._txt_objs.append(t)
            self._tip_xy.append((math.cos(ang) * (r + 0.03), math.sin(ang) * (r + 0.03)))
        
        if self._cfg.get("pie_show_leaders", getattr(SETTINGS, "pie_show_leaders", True)):
            self._redraw_leaders(ax)
        else:
            for ch in list(ax.get_children()):
                if getattr(ch, "_is_leader", False): ch.remove()
        
        self.draw()

    def _redraw_leaders(self, ax):
        def _segment_box_hit(start_xy, end_xy, box):
            sx, sy = start_xy
            ex, ey = end_xy
            dx = ex - sx
            dy = ey - sy
            cands = []

            if abs(dx) > 1e-9:
                t = (box.x0 - sx) / dx
                y = sy + t * dy
                if 0.0 <= t <= 1.0 and box.y0 <= y <= box.y1:
                    cands.append((t, box.x0, y))
                t = (box.x1 - sx) / dx
                y = sy + t * dy
                if 0.0 <= t <= 1.0 and box.y0 <= y <= box.y1:
                    cands.append((t, box.x1, y))

            if abs(dy) > 1e-9:
                t = (box.y0 - sy) / dy
                x = sx + t * dx
                if 0.0 <= t <= 1.0 and box.x0 <= x <= box.x1:
                    cands.append((t, x, box.y0))
                t = (box.y1 - sy) / dy
                x = sx + t * dx
                if 0.0 <= t <= 1.0 and box.x0 <= x <= box.x1:
                    cands.append((t, x, box.y1))

            if not cands:
                return end_xy
            cands.sort(key=lambda it: it[0])
            return (cands[0][1], cands[0][2])

        def _box_min_radius(box):
            try:
                inv = ax.transData.inverted()
                (x0, y0), (x1, y1) = inv.transform([(box.x0, box.y0), (box.x1, box.y1)])
                lo_x, hi_x = sorted((float(x0), float(x1)))
                lo_y, hi_y = sorted((float(y0), float(y1)))
                near_x = min(max(0.0, lo_x), hi_x)
                near_y = min(max(0.0, lo_y), hi_y)
                return math.hypot(near_x, near_y)
            except Exception:
                return None

        for ch in list(ax.get_children()):
            if getattr(ch, "_is_leader", False): ch.remove()
        renderer = None
        try:
            renderer = self.figure.canvas.get_renderer()
        except Exception:
            renderer = None
        for t, (tx, ty) in zip(self._txt_objs, self._tip_xy):
            lx, ly = t.get_position()
            ha = "center"
            t.set_ha(ha)
            leader_width = self._cfg.get("pie_leader_width", SETTINGS.pie_leader_width)
            leader_gap = self._cfg.get("pie_leader_gap", getattr(SETTINGS, "pie_leader_gap", max(0.18, SETTINGS.pie_label_offset * 0.7)))
            box = None
            if renderer is not None:
                try:
                    box = t.get_window_extent(renderer=renderer).expanded(1.02, 1.10)
                    min_r = _box_min_radius(box)
                    if min_r is not None and min_r <= (self._radius + leader_gap):
                        continue
                except Exception:
                    box = None
            elif math.hypot(lx, ly) <= (self._radius + leader_gap):
                continue

            end_x, end_y = lx, ly
            if renderer is not None:
                try:
                    if box is None:
                        box = t.get_window_extent(renderer=renderer).expanded(1.02, 1.10)
                    tip_disp = ax.transData.transform((tx, ty))
                    txt_ctr_disp = ax.transData.transform((lx, ly))
                    hit_disp = _segment_box_hit(tip_disp, txt_ctr_disp, box)
                    end_x, end_y = ax.transData.inverted().transform(hit_disp)
                except Exception:
                    end_x, end_y = lx, ly

            ann = ax.annotate("", xy=(tx, ty), xytext=(end_x, end_y), arrowprops=dict(arrowstyle="-", color="black", lw=leader_width, connectionstyle="arc3,rad=0"), zorder=4, annotation_clip=False)
            ann._is_leader = True

    def _find_txt(self, event):
        if not self._txt_objs: return None
        r = self.figure.canvas.get_renderer()
        for t in self._txt_objs:
            bb = t.get_window_extent(r)
            if (bb.x0 - 4 <= event.x <= bb.x1 + 4 and bb.y0 - 4 <= event.y <= bb.y1 + 4): return t
        return None

    def _on_press(self, event):
        if event.button not in (1, 3): return
        t = self._find_txt(event)
        if t is None:
            if event.inaxes is not None and event.xdata is not None and event.ydata is not None:
                if math.hypot(event.xdata, event.ydata) <= (self._radius + 0.2):
                    self._resize = True; self._resize_start_y = event.y; self._resize_start_r = self._radius; self.setCursor(Qt.SizeVerCursor)
            return
        if event.button != 1: return
        self._drag = t; self._drag_orig  = t.get_position()
        ax = t.axes
        if ax and event.x is not None:
            inv = ax.transData.inverted()
            dx, dy = inv.transform((event.x, event.y))
            self._drag_start = (dx, dy)
        else: self._drag_start = self._drag_orig
        self.setCursor(Qt.ClosedHandCursor)

    def _on_move(self, event):
        if self._resize:
            if event.y is None: return
            dy = self._resize_start_y - event.y
            self._radius = max(0.35, min(1.5, self._resize_start_r + dy * 0.01))
            cb = getattr(self, "_radius_changed_cb", None)
            if callable(cb):
                try:
                    cb(self._radius)
                except Exception:
                    pass
            self._redraw()
            return
        if self._drag is None: return
        ax = self._drag.axes
        if ax is None: return
        inv = ax.transData.inverted()
        cx, cy = inv.transform((event.x, event.y))
        sx, sy = self._drag_start
        ox, oy = self._drag_orig
        self._drag.set_position((ox + cx - sx, oy + cy - sy))
        if self._cfg.get("pie_show_leaders", getattr(SETTINGS, "pie_show_leaders", True)):
            self._redraw_leaders(ax)
        self.draw()

    def _on_release(self, event):
        if self._drag is not None:
            for t in self._txt_objs: self._positions[t.get_text()] = t.get_position()
            self._drag = None
        self._resize = False; self.setCursor(Qt.ArrowCursor)
